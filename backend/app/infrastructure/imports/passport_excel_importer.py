"""
Passport Excel Importer
=======================
Reads loosely formatted passenger spreadsheets into passport submission rows.

The source workbook is treated as untrusted input. Parsing is read-only,
bounded, deterministic, and keeps source values in ``staff_metadata`` for
auditability while exposing only recognized passport fields canonically.
"""

from __future__ import annotations

import io
import math
import re
import unicodedata
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import date, datetime
from itertools import chain, islice
from typing import Any, Final
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.domain.value_objects.passport_fields import (
    MAX_REVIEWED_FIELD_VALUE_LENGTH,
    normalize_extracted_passport_dates,
    normalize_passport_number_identity,
    normalize_passport_sex_identity,
)
from app.infrastructure.ai.passport_date_evidence import (
    normalize_passport_date_evidence,
)

MAX_PASSPORT_EXCEL_WORKSHEETS: Final[int] = 50
MAX_PASSPORT_EXCEL_ROWS: Final[int] = 10_000
MAX_PASSPORT_EXCEL_COLUMNS: Final[int] = 256
MAX_PASSPORT_EXCEL_ARCHIVE_MEMBERS: Final[int] = 2_000
MAX_PASSPORT_EXCEL_UNCOMPRESSED_BYTES: Final[int] = 50 * 1024 * 1024
MAX_PASSPORT_EXCEL_COMPRESSION_RATIO: Final[int] = 250
PASSPORT_EXCEL_COMPRESSION_RATIO_MIN_BYTES: Final[int] = 1024 * 1024
PASSPORT_EXCEL_HEADER_SCAN_ROWS: Final[int] = 25
MAX_PASSPORT_NUMBER_LENGTH: Final[int] = 32
MAX_PASSPORT_EXCEL_CELL_CHARACTERS: Final[int] = 2_048
MAX_PASSPORT_EXCEL_POPULATED_CELLS: Final[int] = 100_000
MAX_PASSPORT_EXCEL_METADATA_CHARACTERS: Final[int] = 8_000_000

_NAME_FIELDS: Final[frozenset[str]] = frozenset({"client_name", "given_names", "surname"})
_DATE_FIELDS: Final[frozenset[str]] = frozenset(
    {"date_of_birth", "date_of_issue", "date_of_expiry"}
)
_RESERVED_METADATA_KEYS: Final[frozenset[str]] = frozenset({"source_sheet", "source_zone"})
_EMPTY_CELL_VALUES: Final[frozenset[str]] = frozenset(
    {"", "-", "--", "#n/a", "n/a", "na", "none", "null"}
)
_EXCEL_ERROR_CELL_VALUES: Final[frozenset[str]] = frozenset(
    {
        "#blocked!",
        "#calc!",
        "#connect!",
        "#div/0!",
        "#field!",
        "#getting_data",
        "#name?",
        "#null!",
        "#num!",
        "#ref!",
        "#spill!",
        "#unknown!",
        "#value!",
    }
)
_MIXED_NUMERIC_DATE_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^\s*([0-9]{1,4})[\s._/-]+([0-9]{1,2})[\s._/-]+([0-9]{1,4})\s*$"
)
_OPENXML_FORMULA_TAG_PATTERN: Final[re.Pattern[bytes]] = re.compile(
    rb"<(?:[A-Za-z_][A-Za-z0-9_.-]*:)?f(?:[\s>/])"
)


def _normalized_header_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _header_identity(value: Any) -> str:
    """Collapse harmless spacing/punctuation variants into one exact key."""

    return _normalized_header_text(value).replace(" ", "")


def _canonicalize_numeric_date_separators(value: str) -> str:
    """Unify strict three-part numeric dates before shared validation."""

    match = _MIXED_NUMERIC_DATE_PATTERN.fullmatch(value)
    if match is None:
        return value
    first, second, third = match.groups()
    if (len(first) == 4 and len(third) <= 2) or (len(first) <= 2 and len(third) == 4):
        return f"{first}-{second}-{third}"
    return value


_HEADER_ALIASES: Final[dict[str, tuple[str, ...]]] = {
    "client_name": (
        "passenger name",
        "full name",
        "client name",
        "traveller name",
        "traveler name",
        "guest name",
        "recipient name",
        "staff name",
        "staffname",
        "passenger",
        "name",
    ),
    "client_email": (
        "email address",
        "client email",
        "passenger email",
        "email",
        "e-mail",
    ),
    "client_phone": (
        "whatsapp number",
        "whatsapp no",
        "mobile number",
        "mobile no",
        "phone number",
        "phone no",
        "contact number",
        "contact no",
        "client phone",
        "telephone",
        "whatsapp",
        "mobile",
        "phone",
        "contact",
    ),
    "departure_city": (
        "departure city",
        "departure hub",
        "hub",
        "city",
    ),
    "nearest_domestic_airport": (
        "nearest domestic airport",
        "nearest airport domestic",
        "domestic airport",
    ),
    "surname": (
        "surname",
        "sur name",
        "family name",
        "last name",
    ),
    "given_names": (
        "given names",
        "given name",
        "first name",
        "forenames",
        "forename",
    ),
    "passport_number": (
        "passport number",
        "passport no",
        "passport num",
        "passport #",
        "passport",
        "pp number",
        "pp no",
        "ppt number",
        "ppt no",
    ),
    "nationality": ("nationality", "citizenship"),
    "place_of_issue": (
        "place of issue",
        "place of issuance",
        "passport place of issue",
        "issue place",
    ),
    "issuing_country": (
        # Preserve legacy imports under their original meaning. These values
        # are not Place of Issue and are never AI-verified as such.
        "issuing country",
        "issue country",
        "country of issue",
    ),
    "date_of_birth": (
        "date of birth",
        "birth date",
        "birthdate",
        "dob",
    ),
    "date_of_issue": (
        "date of issue",
        "date of issuance",
        "passport issue date",
        "passport issue",
        "issue date",
        "doi",
    ),
    "date_of_expiry": (
        "date of expiry",
        "date of expiration",
        "passport expiry date",
        "passport expiration date",
        "passport expiry",
        "passport expiration",
        "expiry date",
        "expiration date",
        "expiry",
        "expiration",
        "valid until",
        "valid till",
        "doe",
    ),
    "sex": ("gender", "sex"),
    "staff_code": (
        "staff code",
        "staffcode",
        "employee code",
        "employee id",
        "staff id",
    ),
}


@dataclass(frozen=True)
class _HeaderAlias:
    field: str
    priority: int


def _build_header_alias_index() -> dict[str, _HeaderAlias]:
    index: dict[str, _HeaderAlias] = {}
    for field, aliases in _HEADER_ALIASES.items():
        for priority, alias in enumerate(aliases):
            identity = _header_identity(alias)
            existing = index.get(identity)
            if existing is not None and existing.field != field:
                raise RuntimeError(f"Excel header alias {alias!r} maps to multiple fields")
            if existing is None or priority < existing.priority:
                index[identity] = _HeaderAlias(field=field, priority=priority)
    return index


_HEADER_ALIAS_INDEX: Final[dict[str, _HeaderAlias]] = _build_header_alias_index()


class PassportExcelImportError(ValueError):
    """A safe, user-actionable workbook import failure."""


@dataclass(frozen=True)
class ImportedPassportRow:
    row_number: int
    worksheet_name: str
    client_name: str
    client_email: str | None
    client_phone: str | None
    departure_city: str | None
    nearest_domestic_airport: str | None
    confirmed_fields: dict[str, str]
    staff_metadata: dict[str, str]


class PassportExcelImporter:
    """Convert an untrusted XLSX workbook into bounded canonical row data."""

    # Retain the public alias catalog used by diagnostic tooling and tests.
    HEADER_ALIASES = _HEADER_ALIASES

    FIELD_KEYS = frozenset(
        {
            "surname",
            "given_names",
            "passport_number",
            "nationality",
            "place_of_issue",
            "issuing_country",
            "date_of_birth",
            "date_of_issue",
            "date_of_expiry",
            "sex",
            "staff_code",
        }
    )

    def import_rows(self, content: bytes) -> list[ImportedPassportRow]:
        contains_formulas = self._validate_archive(content)
        workbook = None
        formula_workbook = None
        try:
            workbook = load_workbook(
                io.BytesIO(content),
                data_only=True,
                read_only=True,
                keep_links=False,
            )
            if contains_formulas:
                # The data-only view intentionally never executes formulas, but
                # it cannot distinguish a blank cell from a formula with no
                # cached result. A second read-only view is opened only for
                # formula-bearing workbooks so canonical identity fields can
                # fail closed without doubling the normal import hot path.
                formula_workbook = load_workbook(
                    io.BytesIO(content),
                    data_only=False,
                    read_only=True,
                    keep_links=False,
                )
            if len(workbook.worksheets) > MAX_PASSPORT_EXCEL_WORKSHEETS:
                raise PassportExcelImportError(
                    "The Excel file contains too many worksheets; "
                    f"use at most {MAX_PASSPORT_EXCEL_WORKSHEETS}."
                )
            # OpenXML dimension metadata is optional. Workbooks created by
            # streaming writers can legitimately report ``None`` here; the
            # row iterator below remains the authoritative bounded check.
            if sum((worksheet.max_row or 0) for worksheet in workbook.worksheets) > (
                MAX_PASSPORT_EXCEL_ROWS
            ):
                raise PassportExcelImportError(
                    "The Excel file contains too many rows; "
                    f"use at most {MAX_PASSPORT_EXCEL_ROWS} across all worksheets."
                )

            imported: list[ImportedPassportRow] = []
            readable_sheet_count = 0
            total_rows = 0
            total_populated_cells = 0
            total_metadata_characters = 0
            for worksheet in workbook.worksheets:
                formula_worksheet = (
                    formula_workbook[worksheet.title]
                    if formula_workbook is not None
                    else None
                )
                if (worksheet.max_column or 0) > MAX_PASSPORT_EXCEL_COLUMNS:
                    raise PassportExcelImportError(
                        f"Worksheet {worksheet.title!r} contains too many columns; "
                        f"use at most {MAX_PASSPORT_EXCEL_COLUMNS}."
                    )
                row_iterator = worksheet.iter_rows(values_only=True)
                scanned_rows = list(islice(row_iterator, PASSPORT_EXCEL_HEADER_SCAN_ROWS))
                formula_row_iterator = (
                    formula_worksheet.iter_rows(values_only=False)
                    if formula_worksheet is not None
                    else None
                )
                scanned_formula_rows = (
                    list(islice(formula_row_iterator, PASSPORT_EXCEL_HEADER_SCAN_ROWS))
                    if formula_row_iterator is not None
                    else []
                )
                if formula_row_iterator is not None and len(scanned_formula_rows) != len(
                    scanned_rows
                ):
                    raise PassportExcelImportError(
                        "The Excel formula and value views are inconsistent."
                    )
                total_rows += len(scanned_rows)
                self._validate_row_count(total_rows)
                if any(
                    len(scanned_row) > MAX_PASSPORT_EXCEL_COLUMNS
                    for scanned_row in scanned_rows
                ):
                    raise PassportExcelImportError(
                        f"Worksheet {worksheet.title!r} contains too many columns; "
                        f"use at most {MAX_PASSPORT_EXCEL_COLUMNS}."
                    )
                if not scanned_rows:
                    continue
                scanned_text_rows = [
                    tuple(self._stringify(value) for value in scanned_row)
                    for scanned_row in scanned_rows
                ]
                scanned_formula_usage = [
                    self._formula_row_usage(formula_row)
                    for formula_row in scanned_formula_rows
                ]
                header_index = self._find_header_row(scanned_rows)
                if header_index is None:
                    # Do not heuristically interpret note/lookup sheets as
                    # passenger data when no exact identity header is present.
                    # Still drain and budget every row so an auxiliary sheet
                    # cannot bypass workbook-wide resource limits by omitting
                    # its optional OpenXML dimension metadata.
                    for scanned_index, text_values in enumerate(scanned_text_rows):
                        formula_columns, formula_source_characters = (
                            scanned_formula_usage[scanned_index]
                            if scanned_formula_usage
                            else (frozenset(), 0)
                        )
                        populated_cells, materialized_characters = (
                            self._row_materialized_usage(
                                text_values,
                                formula_columns=formula_columns,
                                formula_source_characters=formula_source_characters,
                            )
                        )
                        total_populated_cells += populated_cells
                        total_metadata_characters += materialized_characters
                        self._validate_materialized_budget(
                            total_populated_cells=total_populated_cells,
                            total_metadata_characters=total_metadata_characters,
                        )
                    for values in row_iterator:
                        total_rows += 1
                        self._validate_row_count(total_rows)
                        if len(values) > MAX_PASSPORT_EXCEL_COLUMNS:
                            raise PassportExcelImportError(
                                f"Worksheet {worksheet.title!r} contains too many columns; "
                                f"use at most {MAX_PASSPORT_EXCEL_COLUMNS}."
                            )
                        formula_columns, formula_source_characters = (
                            self._next_formula_row_usage(formula_row_iterator)
                        )
                        text_values = tuple(self._stringify(value) for value in values)
                        populated_cells, materialized_characters = (
                            self._row_materialized_usage(
                                text_values,
                                formula_columns=formula_columns,
                                formula_source_characters=formula_source_characters,
                            )
                        )
                        total_populated_cells += populated_cells
                        total_metadata_characters += materialized_characters
                        self._validate_materialized_budget(
                            total_populated_cells=total_populated_cells,
                            total_metadata_characters=total_metadata_characters,
                        )
                    continue

                readable_sheet_count += 1
                for scanned_index, text_values in enumerate(
                    scanned_text_rows[: header_index + 1]
                ):
                    formula_columns, formula_source_characters = (
                        scanned_formula_usage[scanned_index]
                        if scanned_formula_usage
                        else (frozenset(), 0)
                    )
                    populated_cells, materialized_characters = (
                        self._row_materialized_usage(
                            text_values,
                            formula_columns=formula_columns,
                            formula_source_characters=formula_source_characters,
                        )
                    )
                    total_populated_cells += populated_cells
                    total_metadata_characters += materialized_characters
                    self._validate_materialized_budget(
                        total_populated_cells=total_populated_cells,
                        total_metadata_characters=total_metadata_characters,
                    )
                header_row = scanned_rows[header_index]
                if scanned_formula_rows and any(
                    cell.data_type == "f" for cell in scanned_formula_rows[header_index]
                ):
                    raise PassportExcelImportError(
                        f"Worksheet {worksheet.title!r} uses a formula in its header row; "
                        "replace formula headers with plain text."
                    )
                headers = self._map_headers(
                    header_row,
                    worksheet_name=worksheet.title,
                )
                metadata_headers = self._metadata_headers(header_row)
                data_rows = chain(scanned_rows[header_index + 1 :], row_iterator)
                formula_rows = (
                    chain(
                        scanned_formula_rows[header_index + 1 :],
                        formula_row_iterator,
                    )
                    if formula_row_iterator is not None
                    else None
                )
                first_row_number = header_index + 2

                for row_number, values in enumerate(
                    data_rows,
                    start=first_row_number,
                ):
                    if row_number > len(scanned_rows):
                        total_rows += 1
                        self._validate_row_count(total_rows)
                    if len(values) > MAX_PASSPORT_EXCEL_COLUMNS:
                        raise PassportExcelImportError(
                            f"Worksheet {worksheet.title!r} contains too many columns; "
                            f"use at most {MAX_PASSPORT_EXCEL_COLUMNS}."
                        )
                    formula_columns: frozenset[int] = frozenset()
                    formula_source_characters = 0
                    if formula_rows is not None:
                        formula_row = next(formula_rows, None)
                        if formula_row is None:
                            raise PassportExcelImportError(
                                "The Excel formula and value views are inconsistent."
                            )
                        formula_columns, formula_source_characters = (
                            self._formula_row_usage(formula_row)
                        )
                    text_values = tuple(self._stringify(value) for value in values)
                    populated_cells, materialized_characters = (
                        self._row_materialized_usage(
                            text_values,
                            formula_columns=formula_columns,
                            formula_source_characters=formula_source_characters,
                            metadata_headers=metadata_headers,
                        )
                    )
                    total_populated_cells += populated_cells
                    total_metadata_characters += materialized_characters
                    self._validate_materialized_budget(
                        total_populated_cells=total_populated_cells,
                        total_metadata_characters=total_metadata_characters,
                    )
                    if not values or self._is_repeated_header_row(values, headers):
                        continue
                    if not any(text_values) and not formula_columns:
                        continue

                    mapped = self._map_row(
                        headers,
                        text_values,
                        source_row=values,
                        formula_columns=formula_columns,
                    )
                    # Passport name parts are explicit identity evidence and
                    # always outrank a generic Name/Staffname column.
                    client_name = self._name_from_parts(mapped) or mapped.get("client_name")
                    if not client_name:
                        continue

                    confirmed_fields = self._confirmed_fields(mapped)
                    staff_metadata = self._map_metadata_row(
                        metadata_headers,
                        text_values,
                        formula_columns=formula_columns,
                    )
                    if mapped.get("staff_code"):
                        staff_metadata.setdefault("staff_code", mapped["staff_code"])
                    staff_metadata["source_sheet"] = worksheet.title
                    # A zone heading is not guaranteed in third-party templates;
                    # the worksheet name remains a reliable grouping fallback.
                    staff_metadata["source_zone"] = (
                        staff_metadata.get("zone_name") or worksheet.title
                    )
                    imported.append(
                        ImportedPassportRow(
                            row_number=row_number,
                            worksheet_name=worksheet.title,
                            client_name=self._bounded_value(
                                client_name,
                                limit=255,
                                label="Passenger name",
                            )
                            or "",
                            client_email=self._normalize_email(mapped.get("client_email")),
                            client_phone=self._bounded_value(
                                mapped.get("client_phone"),
                                limit=32,
                                label="Phone number",
                            ),
                            departure_city=self._bounded_value(
                                mapped.get("departure_city"),
                                limit=120,
                                label="Departure city",
                            ),
                            nearest_domestic_airport=self._bounded_value(
                                mapped.get("nearest_domestic_airport"),
                                limit=120,
                                label="Nearest domestic airport",
                            ),
                            confirmed_fields=confirmed_fields,
                            staff_metadata=staff_metadata,
                        )
                    )
        except PassportExcelImportError:
            raise
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise PassportExcelImportError("The uploaded Excel file could not be read.") from exc
        finally:
            if workbook is not None:
                workbook.close()
            if formula_workbook is not None:
                formula_workbook.close()

        if not readable_sheet_count:
            raise PassportExcelImportError(
                "Could not find a header row in any worksheet of the Excel file"
            )
        return imported

    def _validate_archive(self, content: bytes) -> bool:
        if not content:
            raise PassportExcelImportError("The uploaded Excel file is empty.")
        try:
            with ZipFile(io.BytesIO(content)) as archive:
                members = archive.infolist()
                if len(members) > MAX_PASSPORT_EXCEL_ARCHIVE_MEMBERS:
                    raise PassportExcelImportError(
                        "The Excel file contains too many archive entries; "
                        "simplify the workbook and try again."
                    )
                total_uncompressed = sum(member.file_size for member in members)
                if total_uncompressed > MAX_PASSPORT_EXCEL_UNCOMPRESSED_BYTES:
                    raise PassportExcelImportError(
                        "The Excel file expands beyond the allowed size; "
                        "simplify the workbook and try again."
                    )
                for member in members:
                    if (
                        member.file_size > PASSPORT_EXCEL_COMPRESSION_RATIO_MIN_BYTES
                        and member.compress_size > 0
                        and member.file_size / member.compress_size
                        > MAX_PASSPORT_EXCEL_COMPRESSION_RATIO
                    ):
                        raise PassportExcelImportError(
                            "The Excel file has an unsafe compression ratio."
                        )
                return any(
                    self._archive_member_contains_formula(archive, member)
                    for member in members
                    if member.filename.casefold().startswith("xl/worksheets/")
                    and member.filename.casefold().endswith(".xml")
                )
        except (BadZipFile, OSError, RuntimeError) as exc:
            raise PassportExcelImportError("The uploaded Excel file could not be read.") from exc

    def _archive_member_contains_formula(self, archive: ZipFile, member: Any) -> bool:
        carry = b""
        with archive.open(member) as source:
            while chunk := source.read(64 * 1024):
                candidate = carry + chunk
                if _OPENXML_FORMULA_TAG_PATTERN.search(candidate):
                    return True
                carry = candidate[-128:]
        return False

    def _formula_row_usage(
        self,
        row: tuple[Any, ...] | None,
    ) -> tuple[frozenset[int], int]:
        """Return formula columns and bounded source-text size for one row."""

        if row is None:
            return frozenset(), 0
        formula_columns: set[int] = set()
        source_characters = 0
        for index, cell in enumerate(row):
            if getattr(cell, "data_type", None) != "f":
                continue
            formula_columns.add(index)
            source_characters += len(self._stringify(getattr(cell, "value", None)))
        return frozenset(formula_columns), source_characters

    def _next_formula_row_usage(
        self,
        formula_rows: Iterator[tuple[Any, ...]] | None,
    ) -> tuple[frozenset[int], int]:
        if formula_rows is None:
            return frozenset(), 0
        formula_row = next(formula_rows, None)
        if formula_row is None:
            raise PassportExcelImportError(
                "The Excel formula and value views are inconsistent."
            )
        return self._formula_row_usage(formula_row)

    def _row_materialized_usage(
        self,
        text_values: tuple[str, ...],
        *,
        formula_columns: frozenset[int] = frozenset(),
        formula_source_characters: int = 0,
        metadata_headers: dict[int, str] | None = None,
    ) -> tuple[int, int]:
        """Charge formula cells once while bounding both source and cached text."""

        populated_cells = len(formula_columns) + sum(
            bool(value)
            for index, value in enumerate(text_values)
            if index not in formula_columns
        )
        header_names = metadata_headers or {}
        materialized_characters = formula_source_characters + sum(
            len(value) + len(header_names.get(index, ""))
            for index, value in enumerate(text_values)
            if value
        )
        return populated_cells, materialized_characters

    def _validate_row_count(self, total_rows: int) -> None:
        if total_rows > MAX_PASSPORT_EXCEL_ROWS:
            raise PassportExcelImportError(
                "The Excel file contains too many rows; "
                f"use at most {MAX_PASSPORT_EXCEL_ROWS} across all worksheets."
            )

    def _validate_materialized_budget(
        self,
        *,
        total_populated_cells: int,
        total_metadata_characters: int,
    ) -> None:
        if total_populated_cells > MAX_PASSPORT_EXCEL_POPULATED_CELLS:
            raise PassportExcelImportError(
                "The Excel file contains too many populated cells; "
                f"use at most {MAX_PASSPORT_EXCEL_POPULATED_CELLS}."
            )
        if total_metadata_characters > MAX_PASSPORT_EXCEL_METADATA_CHARACTERS:
            raise PassportExcelImportError(
                "The Excel file contains too much passenger detail; "
                "remove unnecessary columns or split it into smaller files."
            )

    def _find_header_row(self, rows: list[tuple[Any, ...]]) -> int | None:
        best_index: int | None = None
        best_score: tuple[int, int, int] | None = None
        for index, row in enumerate(rows[:PASSPORT_EXCEL_HEADER_SCAN_ROWS]):
            matches = [
                match
                for value in row
                if (match := _HEADER_ALIAS_INDEX.get(_header_identity(value))) is not None
            ]
            fields = {match.field for match in matches}
            if not fields.intersection(_NAME_FIELDS):
                continue
            explicit_name_score = (
                2
                if {"given_names", "surname"}.issubset(fields)
                else 1
                if fields.intersection({"given_names", "surname"})
                else 0
            )
            score = (explicit_name_score, len(fields), len(matches))
            if best_score is None or score > best_score:
                best_score = score
                best_index = index
        return best_index

    def _map_headers(
        self,
        row: tuple[Any, ...],
        *,
        worksheet_name: str = "worksheet",
    ) -> dict[int, str]:
        selected: dict[str, tuple[int, int, str]] = {}
        for index, value in enumerate(row):
            match = _HEADER_ALIAS_INDEX.get(_header_identity(value))
            if match is None:
                continue
            source_label = self._stringify(value)
            existing = selected.get(match.field)
            if existing is None or match.priority < existing[0]:
                selected[match.field] = (match.priority, index, source_label)
                continue
            if match.priority == existing[0]:
                raise PassportExcelImportError(
                    f"Ambiguous headers in worksheet {worksheet_name!r}: "
                    f"{existing[2]!r} and {source_label!r} both map to "
                    f"{match.field.replace('_', ' ')}."
                )
        return {index: field for field, (_, index, _) in selected.items()}

    def _metadata_headers(self, row: tuple[Any, ...]) -> dict[int, str]:
        """Return stable JSON keys for every populated source column."""

        keys: dict[int, str] = {}
        used: set[str] = set(_RESERVED_METADATA_KEYS)
        for index, value in enumerate(row):
            source_header = self._stringify(value)
            if not source_header:
                continue
            base = re.sub(
                r"[^a-z0-9]+",
                "_",
                unicodedata.normalize("NFKC", source_header).casefold(),
            ).strip("_")
            if not base:
                base = f"column_{index + 1}"
            key = base[:64]
            suffix = 2
            while key in used:
                suffix_text = f"_{suffix}"
                key = f"{base[: 64 - len(suffix_text)]}{suffix_text}"
                suffix += 1
            used.add(key)
            keys[index] = key
        return keys

    def _map_row(
        self,
        headers: dict[int, str],
        row: tuple[str, ...],
        *,
        source_row: tuple[Any, ...],
        formula_columns: frozenset[int] = frozenset(),
    ) -> dict[str, str]:
        mapped: dict[str, str] = {}
        for index, field in headers.items():
            if index >= len(row):
                continue
            if index in formula_columns:
                continue
            value = row[index]
            if value and value.casefold() not in _EXCEL_ERROR_CELL_VALUES:
                mapped[field] = value
            elif field == "surname" and (
                index >= len(source_row)
                or source_row[index] is None
                or not str(source_row[index]).strip()
            ):
                # A genuinely blank surname cell is meaningful passport data.
                # Null markers such as N/A remain absent rather than clearing it.
                mapped[field] = ""
        return mapped

    def _confirmed_fields(self, mapped: dict[str, str]) -> dict[str, str]:
        confirmed: dict[str, str] = {}
        for field, raw_value in mapped.items():
            if field not in self.FIELD_KEYS:
                continue
            if field == "surname" and raw_value == "":
                confirmed[field] = ""
                continue
            if not raw_value:
                continue
            value = raw_value
            if field in _DATE_FIELDS:
                value = normalize_passport_date_evidence(
                    _canonicalize_numeric_date_separators(value),
                    field=field,
                    numeric_order="day_first",
                )
            elif field == "sex":
                value = normalize_passport_sex_identity(value)
            elif field == "passport_number":
                value = normalize_passport_number_identity(value)
                if len(value) > MAX_PASSPORT_NUMBER_LENGTH:
                    raise PassportExcelImportError(
                        "Passport number can contain at most "
                        f"{MAX_PASSPORT_NUMBER_LENGTH} letters and numbers."
                    )
            elif len(value) > MAX_REVIEWED_FIELD_VALUE_LENGTH:
                raise PassportExcelImportError(
                    f"{field.replace('_', ' ').title()} can contain at most "
                    f"{MAX_REVIEWED_FIELD_VALUE_LENGTH} characters."
                )
            if value:
                confirmed[field] = value

        return {
            key: str(value) for key, value in normalize_extracted_passport_dates(confirmed).items()
        }

    def _map_metadata_row(
        self,
        headers: dict[int, str],
        row: tuple[str, ...],
        *,
        formula_columns: frozenset[int] = frozenset(),
    ) -> dict[str, str]:
        return {
            key: value
            for index, key in headers.items()
            if index not in formula_columns
            and index < len(row)
            and (value := row[index])
        }

    def _is_repeated_header_row(
        self,
        row: tuple[Any, ...],
        headers: dict[int, str],
    ) -> bool:
        return bool(headers) and all(
            index < len(row)
            and (match := _HEADER_ALIAS_INDEX.get(_header_identity(row[index]))) is not None
            and match.field == field
            for index, field in headers.items()
        )

    def _normalize_header(self, value: Any) -> str:
        return _normalized_header_text(value)

    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float) and math.isfinite(value) and value.is_integer():
            # Excel commonly stores identifier-like numeric cells as floats.
            # Removing punctuation from "4779216.0" would otherwise corrupt a
            # passport number into "47792160" and leak .0 into codes/phones.
            value = int(value)
        raw_text = str(value)
        # Bound the actual source value before whitespace normalization. A cell
        # containing thousands of spaces must not evade per-cell and workbook
        # materialization limits merely because it collapses to an empty value.
        if len(raw_text) > MAX_PASSPORT_EXCEL_CELL_CHARACTERS:
            raise PassportExcelImportError(
                f"Excel cells can contain at most {MAX_PASSPORT_EXCEL_CELL_CHARACTERS} characters."
            )
        text = " ".join(raw_text.strip().split())
        return "" if text.casefold() in _EMPTY_CELL_VALUES else text

    def _name_from_parts(self, mapped: dict[str, str]) -> str:
        return " ".join(
            part for part in (mapped.get("given_names"), mapped.get("surname")) if part
        ).strip()

    def _normalize_email(self, value: str | None) -> str | None:
        if not value:
            return None
        normalized = value.casefold()
        return self._bounded_value(
            normalized,
            limit=255,
            label="Email address",
        )

    def _bounded_value(
        self,
        value: str | None,
        *,
        limit: int,
        label: str,
    ) -> str | None:
        if value and len(value) > limit:
            raise PassportExcelImportError(f"{label} can contain at most {limit} characters.")
        return value or None
