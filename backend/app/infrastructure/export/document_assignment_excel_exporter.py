"""Excel export for filtered document-assignment review rows."""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

_HEADER_FILL = PatternFill("solid", fgColor="123F73")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_DANGEROUS_FORMULA_PREFIXES = ("=", "+", "-", "@")


@dataclass(frozen=True, slots=True)
class DocumentAssignmentExportRow:
    passenger_name: str
    passport_number: str
    departure_city: str
    assignment_status: str
    document_count: int
    document_filenames: str
    match_statuses: str
    match_confidences: str
    delivery_statuses: str
    sent_to: str
    last_sent_at: str
    match_reasons: str


def _excel_text(value: object) -> str:
    text = str(value or "")
    return f"'{text}" if text.startswith(_DANGEROUS_FORMULA_PREFIXES) else text


def build_document_assignment_workbook(
    *,
    group_name: str,
    document_label: str,
    filter_label: str,
    search_query: str,
    rows: Sequence[DocumentAssignmentExportRow],
) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Document Assignments"
    sheet.freeze_panes = "A5"

    sheet.append(["Group", _excel_text(group_name)])
    sheet.append(["Document Type", _excel_text(document_label)])
    sheet.append(
        [
            "View",
            _excel_text(filter_label),
            "Passenger Search",
            _excel_text(search_query.strip() or "All passengers"),
        ]
    )
    headers = [
        "No.",
        "Passenger Name",
        "Passport Number",
        "Departure City",
        "Assignment Status",
        "Document Count",
        "Document Filenames",
        "Match Status",
        "Match Confidence",
        "Delivery Status",
        "Sent To",
        "Last Sent At",
        "Match Reason",
    ]
    sheet.append(headers)

    for index, row in enumerate(rows, start=1):
        sheet.append(
            [
                index,
                _excel_text(row.passenger_name),
                _excel_text(row.passport_number),
                _excel_text(row.departure_city),
                row.assignment_status,
                row.document_count,
                _excel_text(row.document_filenames),
                _excel_text(row.match_statuses),
                _excel_text(row.match_confidences),
                _excel_text(row.delivery_statuses),
                _excel_text(row.sent_to),
                _excel_text(row.last_sent_at),
                _excel_text(row.match_reasons),
            ]
        )

    for cell in sheet[4]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if rows:
        table = Table(displayName="DocumentAssignments", ref=f"A4:M{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet.auto_filter.ref = "A4:M4"

    widths = {
        "A": 8,
        "B": 30,
        "C": 20,
        "D": 20,
        "E": 20,
        "F": 16,
        "G": 42,
        "H": 24,
        "I": 22,
        "J": 24,
        "K": 26,
        "L": 24,
        "M": 48,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
