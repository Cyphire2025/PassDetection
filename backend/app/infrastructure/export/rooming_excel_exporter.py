"""Hotel-facing XLSX rooming list exporter."""

from __future__ import annotations

import io
import re
import uuid
from datetime import UTC, date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.domain.value_objects.personnel_codes import (
    prefixed_agent_employee_code,
    prefixed_staff_code,
)
from app.infrastructure.export.passport_excel_exporter import passport_age_group

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_ISO_DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_DATE_FORMAT = "DD.MM.YYYY"
_HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
_VIP_FILL = PatternFill("solid", fgColor="FDE68A")
_FIXED_GUEST_HEADERS = (
    "Age Group",
    "GIVEN NAME",
    "SURNAME",
    "GENDER",
    "PASSPORT NUM",
    "DOB",
    "DOI",
    "DOE",
    "PLACE OF ISSUE",
)


def safe_rooming_xlsx_value(value: Any) -> Any:
    """Prevent every untrusted string from becoming an Excel formula."""

    if not isinstance(value, str):
        return value
    return f"'{value}" if value.lstrip().startswith(_FORMULA_PREFIXES) else value


def _excel_date(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return value
    normalized = value.strip()
    if not _ISO_DATE_PATTERN.fullmatch(normalized):
        return value
    try:
        return date.fromisoformat(normalized)
    except ValueError:
        return value


def _gender(value: Any) -> str | None:
    normalized = " ".join(str(value or "").strip().split()).casefold()
    if normalized in {"m", "male"}:
        return "Male"
    if normalized in {"f", "female"}:
        return "Female"
    return " ".join(str(value or "").strip().split()) or None


def _upper(value: Any) -> str | None:
    return str(value).upper() if value not in (None, "") else None


class RoomingExcelExporter:
    """Build deterministic rooming and check-in workbooks."""

    def export_hotel(
        self,
        *,
        group: Any,
        hotel: Any,
        rooms: list[tuple[Any, list[Any]]],
        passenger_by_id: dict[uuid.UUID, Any],
        vip_passenger_ids: set[uuid.UUID],
        priority_fields: list[dict[str, str]],
        priority_values: dict[uuid.UUID, dict[str, str | None]],
    ) -> bytes:
        code_headers: list[str] = []
        if group.staff_code_enabled:
            code_headers.append("Staff Code")
        if group.agent_employee_code_enabled:
            code_headers.append("Agent/Employee Code")
        headers = [
            "Room Number",
            "Room Type",
            "VIP",
            *code_headers,
            *(field["label"] for field in priority_fields),
            *_FIXED_GUEST_HEADERS,
        ]

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Rooming List"
        last_column = len(headers)

        worksheet["A1"] = safe_rooming_xlsx_value(
            f"Rooming List - {hotel.hotel_name}"
        )
        worksheet["A1"].font = Font(bold=True, size=15)
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=last_column,
        )
        worksheet["A2"] = safe_rooming_xlsx_value(f"Group: {group.name}")
        worksheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=last_column,
        )
        dates = (
            f"{hotel.check_in_date or ''} to {hotel.check_out_date or ''}"
            if hotel.check_in_date or hotel.check_out_date
            else "-"
        )
        worksheet["A3"] = safe_rooming_xlsx_value(
            f"City: {hotel.city or '-'}   Stay: {dates}"
        )
        worksheet.merge_cells(
            start_row=3,
            start_column=1,
            end_row=3,
            end_column=last_column,
        )
        worksheet["A4"] = safe_rooming_xlsx_value(
            f"Generated: {datetime.now(tz=UTC).strftime('%d %b %Y %H:%M UTC')}"
        )
        worksheet.merge_cells(
            start_row=4,
            start_column=1,
            end_row=4,
            end_column=last_column,
        )

        worksheet.append([])
        worksheet.append([safe_rooming_xlsx_value(header) for header in headers])
        header_row = 6
        for cell in worksheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_count = 0
        date_columns = {
            headers.index(header) + 1 for header in ("DOB", "DOI", "DOE")
        }
        for room, assignments in rooms:
            if not assignments:
                worksheet.append(
                    [
                        safe_rooming_xlsx_value(room.room_number),
                        safe_rooming_xlsx_value(room.room_type.title()),
                        None,
                        *([None] * (len(headers) - 3)),
                    ]
                )
                row_count += 1
                continue
            for assignment in assignments:
                passenger = passenger_by_id[assignment.passenger_id]
                passport_fields = (
                    passenger.confirmed_fields or passenger.extracted_fields or {}
                )
                staff_metadata = passenger.staff_metadata or {}
                values: list[Any] = [
                    room.room_number,
                    room.room_type.title(),
                    "YES" if passenger.id in vip_passenger_ids else None,
                ]
                if group.staff_code_enabled:
                    values.append(
                        prefixed_staff_code(
                            passport_fields.get("staff_code")
                            or staff_metadata.get("staff_code")
                        )
                    )
                if group.agent_employee_code_enabled:
                    values.append(
                        prefixed_agent_employee_code(
                            passport_fields.get("agent_employee_type")
                            or staff_metadata.get("agent_employee_type"),
                            passport_fields.get("agent_employee_code")
                            or staff_metadata.get("agent_employee_code"),
                        )
                    )
                values.extend(
                    priority_values.get(passenger.id, {}).get(field["key"])
                    for field in priority_fields
                )
                values.extend(
                    [
                        passport_age_group(
                            passport_fields.get("date_of_birth"),
                            group.travel_date,
                        ),
                        _upper(passport_fields.get("given_names")),
                        _upper(passport_fields.get("surname")),
                        _gender(passport_fields.get("sex")),
                        passport_fields.get("passport_number"),
                        _excel_date(passport_fields.get("date_of_birth")),
                        _excel_date(passport_fields.get("date_of_issue")),
                        _excel_date(passport_fields.get("date_of_expiry")),
                        passport_fields.get("place_of_issue"),
                    ]
                )
                worksheet.append([safe_rooming_xlsx_value(value) for value in values])
                row_count += 1
                row_index = worksheet.max_row
                for column_index in date_columns:
                    cell = worksheet.cell(row=row_index, column=column_index)
                    if isinstance(cell.value, (date, datetime)):
                        cell.number_format = _DATE_FORMAT
                if passenger.id in vip_passenger_ids:
                    for cell in worksheet[row_index]:
                        cell.fill = _VIP_FILL
                        cell.font = Font(bold=True)

        if row_count:
            end_column = worksheet.cell(
                row=header_row,
                column=len(headers),
            ).column_letter
            table = Table(
                displayName="HotelRoomingList",
                ref=f"A{header_row}:{end_column}{header_row + row_count}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

        widths = {
            "Room Number": 16,
            "Room Type": 14,
            "VIP": 10,
            "Staff Code": 18,
            "Agent/Employee Code": 24,
            "Age Group": 14,
            "GIVEN NAME": 24,
            "SURNAME": 20,
            "GENDER": 12,
            "PASSPORT NUM": 20,
            "DOB": 16,
            "DOI": 16,
            "DOE": 16,
            "PLACE OF ISSUE": 22,
        }
        for column, header in enumerate(headers, start=1):
            worksheet.column_dimensions[
                worksheet.cell(row=header_row, column=column).column_letter
            ].width = widths.get(header, 24)
        worksheet.sheet_view.showGridLines = False
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.freeze_panes = "A7"

        for row in worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row + row_count,
        ):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def export_checkins(
        self,
        *,
        group_name: str,
        hotel_name: str,
        passengers: list[Any],
    ) -> bytes:
        headers = [
            "Group Name",
            "Hotel Name",
            "Room No",
            "Room Type",
            "Passenger Name",
            "Checked in",
            "Key Issued",
            "Welcome Kit/Letter Issued",
            "Remarks",
        ]
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Hotel Check-in"
        worksheet.append([safe_rooming_xlsx_value(header) for header in headers])
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="047857")
            cell.alignment = Alignment(horizontal="center")
        for item in passengers:
            worksheet.append(
                [
                    safe_rooming_xlsx_value(group_name),
                    safe_rooming_xlsx_value(hotel_name),
                    safe_rooming_xlsx_value(item.room_number),
                    safe_rooming_xlsx_value(item.room_type.title()),
                    safe_rooming_xlsx_value(item.passenger_name),
                    "Yes" if item.checked_in else "No",
                    "Yes" if item.key_issued else "No",
                    "Yes" if item.welcome_letter_issued else "No",
                    safe_rooming_xlsx_value(item.remarks),
                ]
            )
            if item.is_vip:
                for cell in worksheet[worksheet.max_row]:
                    cell.fill = _VIP_FILL
        if passengers:
            table = Table(
                displayName="HotelCheckins",
                ref=f"A1:I{len(passengers) + 1}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4",
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)
        for index, width in enumerate(
            [26, 26, 14, 14, 30, 14, 14, 28, 42],
            start=1,
        ):
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=index).column_letter
            ].width = width
        worksheet.freeze_panes = "A2"
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
