"""
Passport Excel Exporter
=======================
Generates agency-ready XLSX files from confirmed passport submissions.
"""

from __future__ import annotations

import io
import re
import uuid
from dataclasses import dataclass
from datetime import UTC, date, datetime
from typing import Any

import pycountry
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.domain.entities.entities import PassportSubmission
from app.domain.value_objects.personnel_codes import (
    prefixed_agent_employee_code,
    prefixed_staff_code,
)


@dataclass(frozen=True)
class _ExportColumn:
    header: str
    width: int
    enabled_flag: str | None = None
    number_format: str | None = None


_EXCEL_DATE_NUMBER_FORMAT = "DD.MM.YYYY"
_ISO_DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_PREFIX_COLUMNS = (
    _ExportColumn("Group", 24),
    _ExportColumn("Destination", 22),
    _ExportColumn("Travel/Departure Date", 22, number_format=_EXCEL_DATE_NUMBER_FORMAT),
    _ExportColumn("Return Date", 16, number_format=_EXCEL_DATE_NUMBER_FORMAT),
)
_TRAVELLER_COLUMNS = (
    _ExportColumn(
        "Agency/Dealership Name",
        28,
        "agency_dealership_name_enabled",
    ),
    _ExportColumn("Staff Code", 18, "staff_code_enabled"),
    _ExportColumn(
        "Agent/Employee Code",
        24,
        "agent_employee_code_enabled",
    ),
    _ExportColumn("Designation", 22, "designation_enabled"),
    _ExportColumn(
        "Relation with Qualifier",
        24,
        "relation_with_qualifier_enabled",
    ),
    _ExportColumn("Age Group", 14),
    _ExportColumn("Base City", 20, "base_city_enabled"),
    _ExportColumn(
        "Domestic Airport",
        26,
        "ask_nearest_domestic_airport",
    ),
    _ExportColumn("WhatsApp Email", 28),
    _ExportColumn("WhatsApp Phone", 18),
    _ExportColumn("Meal Preference", 18, "meal_preference_enabled"),
    _ExportColumn(
        "International Airport",
        28,
        "nearest_international_airport_enabled",
    ),
    _ExportColumn("SURNAME", 20),
    _ExportColumn("GIVEN NAME", 24),
    _ExportColumn("GENDER", 12),
    _ExportColumn("Passport Number", 20),
    _ExportColumn("DOB", 16, number_format=_EXCEL_DATE_NUMBER_FORMAT),
    _ExportColumn("DOI", 16, number_format=_EXCEL_DATE_NUMBER_FORMAT),
    _ExportColumn("DOE", 16, number_format=_EXCEL_DATE_NUMBER_FORMAT),
    _ExportColumn("Place of Issue", 22),
    _ExportColumn("Nationality", 22),
    _ExportColumn("Upload Email", 28),
    _ExportColumn("Upload Phone", 18),
)
_COLUMNS = _PREFIX_COLUMNS + _TRAVELLER_COLUMNS
_NAME_HISTORY_COLUMNS = (
    _ExportColumn("Old Given Name", 24),
    _ExportColumn("New Surname", 20),
    _ExportColumn("New Given Name", 24),
)

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_PENDING_ROW_FILL = PatternFill("solid", fgColor="FFF2CC")
_ZONE_SEPARATOR_BLANK_ROWS = 2


def _safe_xlsx_value(value: Any) -> Any:
    """Keep untrusted text from being interpreted as an Excel formula."""

    if not isinstance(value, str):
        return value
    if value.lstrip().startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _whatsapp_phone_value(value: Any) -> Any:
    """Remove only an explicit leading India code from WhatsApp export values."""

    if not isinstance(value, str):
        return value
    if value.startswith("'+91"):
        return value[4:].lstrip(" -")
    if value.startswith("+91"):
        return value[3:].lstrip(" -")
    return value


def _uppercase(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return str(value).upper()


def _excel_date_value(value: Any) -> Any:
    """Convert canonical dates to native Excel dates without losing legacy text."""

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return value
    normalized = value.strip()
    if not _ISO_DATE_PATTERN.fullmatch(normalized):
        return value
    try:
        return date.fromisoformat(normalized)
    except ValueError:
        return value


def _export_cell_value(column: _ExportColumn, value: Any) -> Any:
    if column.number_format:
        value = _excel_date_value(value)
    if column.header == "WhatsApp Phone":
        value = _whatsapp_phone_value(value)
    return _safe_xlsx_value(value)


def _gender_display_value(value: Any) -> str | None:
    """Normalize recognized passport sex values without mutating stored data."""

    if value in (None, ""):
        return None
    normalized = " ".join(str(value).strip().split()).casefold()
    if normalized in {"m", "male"}:
        return "Male"
    if normalized in {"f", "female"}:
        return "Female"
    # Do not invent a binary value for an unsupported or non-binary document
    # marker. Returning the trimmed source is safer than silently misgendering
    # the traveller while still normalizing every supported legacy value.
    return " ".join(str(value).strip().split()) or None


def _country_display_name(value: Any) -> str | None:
    if value in (None, ""):
        return None
    normalized = str(value).strip()
    if not normalized:
        return None
    try:
        country = pycountry.countries.lookup(normalized)
    except LookupError:
        return normalized
    return str(getattr(country, "common_name", country.name))


def _nationality_display_value(value: Any) -> str | None:
    """Format nationality for export without changing its stored source value."""

    if value in (None, ""):
        return None
    normalized = " ".join(str(value).strip().split())
    if not normalized:
        return None
    if normalized.casefold() in {"in", "ind", "india", "indian"}:
        return "Indian"
    return _country_display_name(normalized)


def _date_value(value: Any) -> date | None:
    converted = _excel_date_value(value)
    if isinstance(converted, datetime):
        return converted.date()
    return converted if isinstance(converted, date) else None


def passport_age_group(date_of_birth: Any, departure_date: Any) -> str | None:
    """Classify a traveller by completed age on the group's departure date."""

    birth = _date_value(date_of_birth)
    departure = _date_value(departure_date)
    if not birth or not departure or birth > departure:
        return None
    age = departure.year - birth.year - (
        (departure.month, departure.day) < (birth.month, birth.day)
    )
    if age < 2:
        return "Infant"
    if age < 12:
        return "Child"
    return "Adult"


class PassportExcelExporter:
    HEADERS = list(
        dict.fromkeys(
            [
                *(column.header for column in _COLUMNS),
                *(column.header for column in _NAME_HISTORY_COLUMNS),
            ]
        )
    )

    def export_group(
        self,
        submissions: list[PassportSubmission],
        *,
        group_name: str,
        group_details: dict[uuid.UUID, dict[str, Any]] | None = None,
        zone_names: dict[uuid.UUID, str] | None = None,
        additional_fields: list[dict[str, str]] | None = None,
        additional_values: dict[uuid.UUID, dict[str, str | None]] | None = None,
        whatsapp_contacts: dict[
            uuid.UUID,
            dict[str, str | None],
        ]
        | None = None,
        previous_names: dict[
            uuid.UUID,
            dict[str, str | None],
        ]
        | None = None,
        group_by_field: str | None = None,
        pending_rows: list[dict[str, Any]] | None = None,
    ) -> bytes:
        imported_fields = [
            field
            for field in (additional_fields or [])
            if (
                field.get("key") == "zone_name"
                or str(field.get("key", "")).startswith("whatsapp:")
            )
            and field.get("label")
        ]
        traveller_columns = self._enabled_traveller_columns(
            group_details,
            include_name_history=previous_names is not None,
        )
        custom_fields = self._custom_fields(
            submissions,
            group_details,
            reserved_labels={
                *(column.header.casefold() for column in _PREFIX_COLUMNS),
                *(column.header.casefold() for column in traveller_columns),
                *(str(field["label"]).casefold() for field in imported_fields),
            },
        )
        custom_question_fields = [
            field
            for field in custom_fields
            if str(field["key"]).startswith("custom:")
        ]
        custom_detail_fields = [
            field
            for field in custom_fields
            if str(field["key"]).startswith("custom_detail:")
        ]
        columns = [
            *_PREFIX_COLUMNS,
            *(
                _ExportColumn(str(field["label"]), 22)
                for field in imported_fields
            ),
            *traveller_columns,
            *(
                _ExportColumn(str(field["label"]), 22)
                for field in custom_question_fields
            ),
            *(
                _ExportColumn(str(field["label"]), 22)
                for field in custom_detail_fields
            ),
        ]
        dynamic_fields = [
            *imported_fields,
            *custom_question_fields,
            *custom_detail_fields,
        ]
        headers = [column.header for column in columns]
        label_by_key = {
            str(field["key"]): str(field["label"])
            for field in imported_fields
            if field.get("key") and field.get("label")
        }
        group_by_header = (
            "International Airport"
            if group_by_field == "international_airport"
            else label_by_key.get(group_by_field or "")
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Passport Submissions"

        worksheet["A1"] = f"Passport Export - {group_name}"
        worksheet["A1"].font = Font(bold=True, size=14)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        generated_at = datetime.now(tz=UTC).isoformat(timespec="seconds").replace("+00:00", "Z")
        worksheet["A2"] = f"Generated at {generated_at}"
        worksheet["A2"].font = Font(color="64748B")
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

        worksheet.append([])
        worksheet.append(headers)
        header_row = 4
        for cell in worksheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D4ED8")
            cell.alignment = Alignment(horizontal="center")

        prepared_rows: list[tuple[PassportSubmission, dict[str, Any]]] = []
        for submission in submissions:
            prepared_rows.append(
                (
                    submission,
                    self._submission_values(
                        submission,
                        group_name=group_name,
                        details=(group_details or {}).get(submission.group_id, {}),
                        zone_names=zone_names,
                        dynamic_fields=dynamic_fields,
                        additional_values=additional_values,
                        whatsapp_contacts=whatsapp_contacts,
                        previous_names=previous_names,
                    ),
                )
            )
        ordered_rows = sorted(
            prepared_rows,
            key=lambda item: self._submission_sort_key(
                item[0],
                item[1],
                group_by_header=group_by_header,
            ),
        )
        previous_group_key: str | None = None
        has_written_submission = False
        for submission, values in ordered_rows:
            group_key = self._group_value(values, group_by_header).casefold()
            if (
                group_by_header
                and has_written_submission
                and group_key != previous_group_key
            ):
                # Keep operational batches visually separate without mutating
                # the underlying submission or WhatsApp data.
                for _ in range(_ZONE_SEPARATOR_BLANK_ROWS):
                    worksheet.append([])
            row_values = []
            for column in columns:
                value = values[column.header]
                row_values.append(_export_cell_value(column, value))
            worksheet.append(row_values)
            row_index = worksheet.max_row
            for column_index, column in enumerate(columns, start=1):
                cell = worksheet.cell(row=row_index, column=column_index)
                if column.number_format and isinstance(cell.value, (date, datetime)):
                    cell.number_format = column.number_format
            previous_group_key = group_key
            has_written_submission = True

        if pending_rows:
            self._append_pending_rows(
                worksheet,
                columns=columns,
                pending_rows=pending_rows,
                group_by_header=group_by_header,
            )

        if ordered_rows or pending_rows:
            last_column = worksheet.cell(row=header_row, column=len(headers)).column_letter
            table_ref = f"A{header_row}:{last_column}{worksheet.max_row}"
            table = Table(displayName="PassportSubmissions", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

        for index, column in enumerate(columns, start=1):
            column_letter = worksheet.cell(row=header_row, column=index).column_letter
            worksheet.column_dimensions[column_letter].width = column.width

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @classmethod
    def _append_pending_rows(
        cls,
        worksheet: Any,
        *,
        columns: list[_ExportColumn],
        pending_rows: list[dict[str, Any]],
        group_by_header: str | None,
    ) -> None:
        """Append non-submitters beneath the shared header with durable yellow cells."""

        ordered_rows = sorted(
            pending_rows,
            key=lambda values: cls._pending_row_sort_key(
                values,
                group_by_header=group_by_header,
            ),
        )
        previous_group_key: str | None = None
        has_written_row = False
        for values in ordered_rows:
            group_key = cls._group_value(values, group_by_header).casefold()
            if group_by_header and has_written_row and group_key != previous_group_key:
                for _ in range(_ZONE_SEPARATOR_BLANK_ROWS):
                    worksheet.append([])

            row_values: list[Any] = []
            for column in columns:
                value = values.get(column.header)
                row_values.append(_export_cell_value(column, value))
            worksheet.append(row_values)
            row_index = worksheet.max_row
            for column_index, column in enumerate(columns, start=1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.fill = _PENDING_ROW_FILL
                if column.number_format and isinstance(cell.value, (date, datetime)):
                    cell.number_format = column.number_format

            previous_group_key = group_key
            has_written_row = True

    @staticmethod
    def _group_value(
        values: dict[str, Any],
        group_by_header: str | None,
    ) -> str:
        if not group_by_header:
            return ""
        return " ".join(str(values.get(group_by_header) or "").strip().split())

    @classmethod
    def _pending_row_sort_key(
        cls,
        values: dict[str, Any],
        *,
        group_by_header: str | None,
    ) -> tuple[bool, str, str, str, str]:
        group_value = cls._group_value(values, group_by_header)
        client_name = " ".join(
            str(
                values.get("New Given Name")
                or values.get("GIVEN NAME")
                or values.get("Old Given Name")
                or values.get("New Surname")
                or values.get("SURNAME")
                or ""
            ).strip().split()
        )
        phone = str(values.get("WhatsApp Phone") or "")
        email = str(values.get("WhatsApp Email") or "")
        return (
            not bool(group_value),
            group_value.casefold(),
            client_name.casefold(),
            phone,
            email.casefold(),
        )

    @staticmethod
    def _zone_name(
        submission: PassportSubmission,
        zone_names: dict[uuid.UUID, str] | None,
    ) -> str:
        if zone_names is not None and submission.id in zone_names:
            value = zone_names[submission.id]
        else:
            value = (submission.staff_metadata or {}).get("zone_name")
        return " ".join(str(value or "").strip().split())

    @classmethod
    def _submission_sort_key(
        cls,
        submission: PassportSubmission,
        values: dict[str, Any],
        *,
        group_by_header: str | None,
    ) -> tuple[bool, str, str, str]:
        group_value = cls._group_value(values, group_by_header)
        return (
            not bool(group_value),
            group_value.casefold(),
            submission.client_name.casefold(),
            str(submission.id),
        )

    @classmethod
    def _submission_values(
        cls,
        submission: PassportSubmission,
        *,
        group_name: str,
        details: dict[str, Any],
        zone_names: dict[uuid.UUID, str] | None,
        dynamic_fields: list[dict[str, str]],
        additional_values: dict[uuid.UUID, dict[str, str | None]] | None,
        whatsapp_contacts: dict[
            uuid.UUID,
            dict[str, str | None],
        ]
        | None,
        previous_names: dict[
            uuid.UUID,
            dict[str, str | None],
        ]
        | None,
    ) -> dict[str, Any]:
        fields = submission.confirmed_fields or submission.extracted_fields or {}
        staff_metadata = submission.staff_metadata or {}
        whatsapp_contact = (whatsapp_contacts or {}).get(submission.id, {})
        previous_name = (previous_names or {}).get(submission.id, {})
        values: dict[str, Any] = {
            "Group": details.get("name") or group_name,
            "Destination": details.get("destination"),
            "Travel/Departure Date": details.get("travel_date"),
            "Return Date": details.get("return_date"),
            "Agency/Dealership Name": (
                fields.get("agency_dealership_name")
                or staff_metadata.get("agency_dealership_name")
            ),
            "Base City": fields.get("base_city") or staff_metadata.get("base_city"),
            "Staff Code": prefixed_staff_code(
                fields.get("staff_code") or staff_metadata.get("staff_code")
            ),
            "Agent/Employee Code": prefixed_agent_employee_code(
                fields.get("agent_employee_type")
                or staff_metadata.get("agent_employee_type"),
                fields.get("agent_employee_code")
                or staff_metadata.get("agent_employee_code"),
            ),
            "Designation": (
                fields.get("designation") or staff_metadata.get("designation")
            ),
            "Meal Preference": (
                fields.get("meal_preference") or staff_metadata.get("meal_preference")
            ),
            "Relation with Qualifier": (
                submission.qualifier_relation_label
                if submission.qualifier_enabled_snapshot
                else None
            ),
            "Age Group": passport_age_group(
                fields.get("date_of_birth"),
                details.get("travel_date"),
            ),
            "Domestic Airport": submission.nearest_domestic_airport,
            "WhatsApp Email": whatsapp_contact.get("email"),
            "WhatsApp Phone": whatsapp_contact.get("phone"),
            "International Airport": submission.departure_city,
            "SURNAME": _uppercase(fields.get("surname")),
            "GIVEN NAME": _uppercase(fields.get("given_names")),
            "Old Given Name": _uppercase(previous_name.get("given_names")),
            "New Surname": _uppercase(fields.get("surname")),
            "New Given Name": _uppercase(fields.get("given_names")),
            "GENDER": _gender_display_value(fields.get("sex")),
            "Passport Number": fields.get("passport_number"),
            "DOB": fields.get("date_of_birth"),
            "DOI": fields.get("date_of_issue"),
            "DOE": fields.get("date_of_expiry"),
            "Place of Issue": fields.get("place_of_issue"),
            "Nationality": _nationality_display_value(fields.get("nationality")),
            "Upload Email": submission.client_email,
            "Upload Phone": submission.client_phone,
        }
        row_metadata = (additional_values or {}).get(submission.id, {})
        custom_answers = {
            f"custom:{answer.get('question_id')}": answer.get("value")
            for answer in getattr(submission, "custom_answers", []) or []
            if answer.get("question_id")
        }
        custom_detail_answers = {
            f"custom_detail:{answer.get('detail_id')}": answer.get("value")
            for answer in getattr(submission, "custom_detail_answers", []) or []
            if answer.get("detail_id")
        }
        for field in dynamic_fields:
            key = str(field["key"])
            if key == "zone_name":
                value = cls._zone_name(submission, zone_names) or None
            elif key.startswith("custom_detail:"):
                value = custom_detail_answers.get(key)
            elif key.startswith("custom:"):
                value = custom_answers.get(key)
            else:
                value = row_metadata.get(key)
            values[field["label"]] = value
        return values

    @staticmethod
    def _enabled_traveller_columns(
        group_details: dict[uuid.UUID, dict[str, Any]] | None,
        *,
        include_name_history: bool = False,
    ) -> list[_ExportColumn]:
        if group_details is None:
            enabled_columns = list(_TRAVELLER_COLUMNS)
        else:
            # Callers provide details only for groups included in the workbook.
            # Consider every included group so a pending-only group can still
            # enable and export its configured optional fields.
            relevant_details = list(group_details.values())
            if not relevant_details:
                enabled_columns = list(_TRAVELLER_COLUMNS)
            else:
                enabled_columns = [
                    column
                    for column in _TRAVELLER_COLUMNS
                    if (
                        column.enabled_flag is None
                        or any(
                            bool(details.get(column.enabled_flag, False))
                            for details in relevant_details
                        )
                    )
                ]

        if not include_name_history:
            return enabled_columns

        with_name_history: list[_ExportColumn] = []
        for column in enabled_columns:
            if column.header == "SURNAME":
                with_name_history.extend(_NAME_HISTORY_COLUMNS)
            elif column.header != "GIVEN NAME":
                with_name_history.append(column)
        return with_name_history

    @staticmethod
    def _custom_fields(
        submissions: list[PassportSubmission],
        group_details: dict[uuid.UUID, dict[str, Any]] | None,
        *,
        reserved_labels: set[str],
    ) -> list[dict[str, str]]:
        """Build stable question/detail columns, preserving configuration order."""

        definitions: list[tuple[str, str, str]] = []
        seen_keys: set[str] = set()

        def add(key: str, label: str, source_label: str) -> None:
            normalized_label = " ".join(str(label).strip().split())
            if not key or key in seen_keys or not normalized_label:
                return
            seen_keys.add(key)
            definitions.append((key, normalized_label[:120], source_label))

        for details in (group_details or {}).values():
            for question in details.get("custom_questions") or []:
                if question.get("enabled"):
                    add(
                        f"custom:{question.get('id')}",
                        str(question.get("label") or ""),
                        "Custom Question",
                    )
            for detail in details.get("custom_details") or []:
                if detail.get("enabled"):
                    add(
                        f"custom_detail:{detail.get('id')}",
                        str(detail.get("label") or ""),
                        "Custom Detail",
                    )

        for submission in submissions:
            for answer in getattr(submission, "custom_answers", []) or []:
                add(
                    f"custom:{answer.get('question_id')}",
                    str(answer.get("label") or ""),
                    "Custom Question",
                )
            for answer in getattr(submission, "custom_detail_answers", []) or []:
                add(
                    f"custom_detail:{answer.get('detail_id')}",
                    str(answer.get("label") or ""),
                    "Custom Detail",
                )

        fields: list[dict[str, str]] = []
        used_labels = set(reserved_labels)
        for key, label, source_label in definitions:
            candidate = label
            suffix_index = 1
            while candidate.casefold() in used_labels:
                suffix = (
                    f" ({source_label})"
                    if suffix_index == 1
                    else f" ({source_label} {suffix_index})"
                )
                candidate = f"{label[: max(1, 120 - len(suffix))]}{suffix}"
                suffix_index += 1
            used_labels.add(candidate.casefold())
            fields.append({"key": key, "label": candidate})
        return fields
