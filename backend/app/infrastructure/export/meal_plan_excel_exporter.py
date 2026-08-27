"""Readable XLSX export for saved category-complete meal plans."""

from __future__ import annotations

import io
import uuid
from dataclasses import dataclass
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

MEAL_ORDER = {"lunch": 0, "dinner": 1}
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


@dataclass(frozen=True, slots=True)
class MealPlanExportEntry:
    day_number: int
    meal_type: str
    category_id: uuid.UUID | None
    category_name: str
    dish_name: str
    notes: str | None = None


def _safe_excel_text(value: str | None) -> str | None:
    """Prevent user-entered labels from being interpreted as formulas."""

    if value is None:
        return None
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _entry_dish_name(entry: MealPlanExportEntry | None) -> str | None:
    return _safe_excel_text(entry.dish_name) if entry is not None else None


class MealPlanExcelExporter:
    """Create a matrix view plus a filterable dish-detail worksheet."""

    def export(
        self,
        *,
        plan_name: str,
        trip_days: int,
        start_date: date | None,
        selected_category_ids: list[uuid.UUID],
        entries: list[MealPlanExportEntry],
    ) -> bytes:
        ordered_entries = sorted(
            entries,
            key=lambda entry: (
                entry.day_number,
                MEAL_ORDER.get(entry.meal_type, 99),
                entry.category_name.casefold(),
                str(entry.category_id or ""),
            ),
        )
        categories = self._ordered_categories(
            ordered_entries,
            selected_category_ids=selected_category_ids,
        )

        workbook = Workbook()
        self._build_matrix_sheet(
            workbook.active,
            plan_name=plan_name,
            trip_days=trip_days,
            start_date=start_date,
            categories=categories,
            entries=ordered_entries,
        )
        detail_sheet = workbook.create_sheet("Dish List")
        self._build_detail_sheet(
            detail_sheet,
            start_date=start_date,
            entries=ordered_entries,
        )

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _ordered_categories(
        entries: list[MealPlanExportEntry],
        *,
        selected_category_ids: list[uuid.UUID],
    ) -> list[tuple[str, str]]:
        category_names: dict[str, str] = {}
        for entry in entries:
            key = (
                str(entry.category_id)
                if entry.category_id
                else (f"snapshot:{entry.category_name.casefold()}")
            )
            category_names.setdefault(key, entry.category_name)

        ordered_keys = [
            str(category_id)
            for category_id in selected_category_ids
            if str(category_id) in category_names
        ]
        ordered_keys.extend(
            key
            for key in sorted(
                category_names,
                key=lambda item: (category_names[item].casefold(), item),
            )
            if key not in ordered_keys
        )
        return [(key, category_names[key]) for key in ordered_keys]

    @staticmethod
    def _build_matrix_sheet(
        worksheet: Worksheet,
        *,
        plan_name: str,
        trip_days: int,
        start_date: date | None,
        categories: list[tuple[str, str]],
        entries: list[MealPlanExportEntry],
    ) -> None:
        worksheet.title = "Meal Plan"
        column_count = 3 + len(categories)
        worksheet["A1"] = _safe_excel_text(plan_name)
        worksheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max(1, column_count),
        )
        worksheet["A2"] = (
            f"{trip_days} day{'s' if trip_days != 1 else ''} | "
            f"{trip_days * 2} meals | {len(entries)} unique dishes"
        )
        worksheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=max(1, column_count),
        )
        worksheet["A3"] = (
            f"Trip starts: {start_date.strftime('%d %b %Y')}"
            if start_date
            else "Trip start date: Not set"
        )
        worksheet.merge_cells(
            start_row=3,
            start_column=1,
            end_row=3,
            end_column=max(1, column_count),
        )
        for column in range(1, column_count + 1):
            title_cell = worksheet.cell(row=1, column=column)
            title_cell.fill = PatternFill("solid", fgColor="1E3A8A")
            title_cell.border = THIN_BORDER
            for metadata_row in (2, 3):
                metadata_cell = worksheet.cell(row=metadata_row, column=column)
                metadata_cell.fill = PatternFill("solid", fgColor="EFF6FF")
                metadata_cell.border = THIN_BORDER
        worksheet.row_dimensions[1].height = 28
        worksheet.row_dimensions[2].height = 21
        worksheet.row_dimensions[3].height = 21

        header_row = 5
        headers = ["Day", "Date", "Meal", *[name for _, name in categories]]
        worksheet.append([])
        worksheet.append([_safe_excel_text(header) for header in headers])
        for cell in worksheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D4ED8")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        worksheet.row_dimensions[header_row].height = 24

        entry_lookup = {
            (
                entry.day_number,
                entry.meal_type,
                str(entry.category_id)
                if entry.category_id
                else f"snapshot:{entry.category_name.casefold()}",
            ): entry
            for entry in entries
        }
        day_row_ranges: list[tuple[int, int]] = []
        gap_rows: list[int] = []
        for day_number in range(1, trip_days + 1):
            meal_date = start_date + timedelta(days=day_number - 1) if start_date else None
            day_start_row = worksheet.max_row + 1
            for meal_type in ("lunch", "dinner"):
                worksheet.append(
                    [
                        f"Day {day_number}",
                        meal_date,
                        meal_type.title(),
                        *[
                            _entry_dish_name(
                                entry_lookup.get(
                                    (day_number, meal_type, category_key)
                                )
                            )
                            for category_key, _ in categories
                        ],
                    ]
                )
            day_end_row = worksheet.max_row
            day_row_ranges.append((day_start_row, day_end_row))
            worksheet.merge_cells(
                start_row=day_start_row,
                start_column=1,
                end_row=day_end_row,
                end_column=1,
            )
            if day_number < trip_days:
                gap_row = worksheet.max_row + 1
                worksheet.cell(row=gap_row, column=1)
                gap_rows.append(gap_row)

        last_row = worksheet.max_row

        widths = [13, 15, 13, *([26] * len(categories))]
        for column, width in enumerate(widths, start=1):
            worksheet.column_dimensions[
                worksheet.cell(row=header_row, column=column).column_letter
            ].width = width
        for day_index, (day_start_row, day_end_row) in enumerate(day_row_ranges):
            day_fill = PatternFill(
                "solid",
                fgColor="EFF6FF" if day_index % 2 == 0 else "FFFFFF",
            )
            for row in worksheet.iter_rows(
                min_row=day_start_row,
                max_row=day_end_row,
                max_col=column_count,
            ):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = THIN_BORDER
                    cell.fill = day_fill
                worksheet.row_dimensions[row[0].row].height = 23
            day_cell = worksheet.cell(row=day_start_row, column=1)
            day_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            day_cell.font = Font(bold=True)
            for row_number in range(day_start_row, day_end_row + 1):
                worksheet.cell(row=row_number, column=2).number_format = "dd mmm yyyy"

        for gap_row in gap_rows:
            worksheet.row_dimensions[gap_row].height = 8
            for cell in worksheet[gap_row][:column_count]:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
                cell.border = Border()

        worksheet.freeze_panes = "A6"
        worksheet.sheet_view.showGridLines = True
        worksheet.sheet_view.zoomScale = 90
        worksheet.print_area = f"A1:{worksheet.cell(last_row, column_count).coordinate}"
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    @staticmethod
    def _build_detail_sheet(
        worksheet: Worksheet,
        *,
        start_date: date | None,
        entries: list[MealPlanExportEntry],
    ) -> None:
        headers = ["Day", "Date", "Meal", "Category", "Dish", "Notes"]
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="047857")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        worksheet.row_dimensions[1].height = 24

        day_numbers = sorted({entry.day_number for entry in entries})
        day_row_ranges: list[tuple[int, int]] = []
        gap_rows: list[int] = []
        for day_index, day_number in enumerate(day_numbers):
            day_entries = [entry for entry in entries if entry.day_number == day_number]
            day_start_row = worksheet.max_row + 1
            for entry in day_entries:
                meal_date = (
                    start_date + timedelta(days=entry.day_number - 1) if start_date else None
                )
                worksheet.append(
                    [
                        f"Day {entry.day_number}",
                        meal_date,
                        entry.meal_type.title(),
                        _safe_excel_text(entry.category_name),
                        _safe_excel_text(entry.dish_name),
                        _safe_excel_text(entry.notes),
                    ]
                )
            day_end_row = worksheet.max_row
            day_row_ranges.append((day_start_row, day_end_row))
            worksheet.merge_cells(
                start_row=day_start_row,
                start_column=1,
                end_row=day_end_row,
                end_column=1,
            )
            if day_index < len(day_numbers) - 1:
                gap_row = worksheet.max_row + 1
                worksheet.cell(row=gap_row, column=1)
                gap_rows.append(gap_row)

        last_row = worksheet.max_row
        for column, width in enumerate([13, 15, 13, 22, 30, 40], start=1):
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=column).column_letter
            ].width = width
        for day_index, (day_start_row, day_end_row) in enumerate(day_row_ranges):
            day_fill = PatternFill(
                "solid",
                fgColor="ECFDF5" if day_index % 2 == 0 else "FFFFFF",
            )
            for row in worksheet.iter_rows(
                min_row=day_start_row,
                max_row=day_end_row,
                max_col=6,
            ):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = THIN_BORDER
                    cell.fill = day_fill
                worksheet.row_dimensions[row[0].row].height = 23
            day_cell = worksheet.cell(row=day_start_row, column=1)
            day_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            day_cell.font = Font(bold=True)
            for row_number in range(day_start_row, day_end_row + 1):
                worksheet.cell(row=row_number, column=2).number_format = "dd mmm yyyy"

        for gap_row in gap_rows:
            worksheet.row_dimensions[gap_row].height = 8
            for cell in worksheet[gap_row][:6]:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
                cell.border = Border()

        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = True
        worksheet.sheet_view.zoomScale = 90
        worksheet.print_area = f"A1:F{last_row}"
