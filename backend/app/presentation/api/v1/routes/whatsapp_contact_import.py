"""Whatsapp: contact import."""

from __future__ import annotations

import asyncio
from io import BytesIO
from itertools import islice
from typing import Any
from zipfile import BadZipFile

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.application.use_cases.whatsapp.recipient_capacity import (
    MAX_WHATSAPP_RECIPIENTS,
    WhatsAppRecipientCapacityExceeded,
    require_whatsapp_recipient_capacity,
)
from app.domain.entities.entities import User
from app.presentation.api.v1.routes.whatsapp_scope import _release_auth_transaction
from app.presentation.api.v1.routes.whatsapp_shared import (
    _WHATSAPP_CONTACT_REJECTION_REASONS,
    MAX_WHATSAPP_CONTACT_FILE_BYTES,
    MAX_WHATSAPP_EXCEL_ROWS,
    MAX_WHATSAPP_EXCEL_SHEETS,
    MAX_WHATSAPP_REJECTED_ROWS,
    PHONE_RE,
    WHATSAPP_ROLES,
    WHATSAPP_UPLOAD_READ_CHUNK_BYTES,
    _bounded_excel_raw_value,
    _excel_cell_text,
    _excel_contact_preview_response,
    _excel_fields_from_row,
    _excel_name_from_row,
    _excel_raw_name_from_row,
    _find_excel_contact_header,
    _is_repeated_excel_header,
    _merge_recipient_inputs,
    _normalize_phone,
    _row_has_contact_identity,
    _safe_imported_fields,
    _validate_excel_archive,
    _WhatsAppExcelContactParseResult,
    logger,
)
from app.presentation.api.v1.schemas.whatsapp_schemas import (
    WhatsAppContactPreviewRejectedRow,
    WhatsAppContactPreviewResponse,
    WhatsAppContactRejectionCode,
    WhatsAppRecipientInput,
)
from app.presentation.dependencies.auth import require_role

router = APIRouter()


def _append_excel_contact_rejection(
    rejected_rows: list[WhatsAppContactPreviewRejectedRow],
    rejected_counts: dict[WhatsAppContactRejectionCode, int],
    *,
    sheet_name: str,
    row_number: int,
    raw_name: str | None,
    raw_phone_number: str | None,
    imported_fields: dict[str, str],
    reason_code: WhatsAppContactRejectionCode,
) -> None:
    rejected_counts[reason_code] = rejected_counts.get(reason_code, 0) + 1
    if len(rejected_rows) >= MAX_WHATSAPP_REJECTED_ROWS:
        return
    rejected_rows.append(
        WhatsAppContactPreviewRejectedRow(
            sheet_name=sheet_name,
            row_number=row_number,
            raw_name=raw_name,
            raw_phone_number=raw_phone_number,
            imported_fields=_safe_imported_fields(imported_fields),
            reason_code=reason_code,
            reason=_WHATSAPP_CONTACT_REJECTION_REASONS[reason_code],
        )
    )


def _parse_excel_contact_bytes(
    payload: bytes,
    *,
    filename: str,
) -> _WhatsAppExcelContactParseResult:
    source_file_name = (
        filename.replace("\\", "/").rsplit("/", maxsplit=1)[-1].strip() or "contacts.xlsx"
    )
    suffix = source_file_name.rsplit(".", maxsplit=1)[-1].lower()
    suffix = f".{suffix}" if "." in filename else ".xlsx"
    if suffix not in {".xlsx", ".xlsm"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Upload an .xlsx or .xlsm contact file",
        )

    workbook = None
    try:
        _validate_excel_archive(payload)
        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=True,
        )
        worksheets = workbook.worksheets
        if len(worksheets) > MAX_WHATSAPP_EXCEL_SHEETS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "The Excel contact file contains too many worksheets; "
                    f"use at most {MAX_WHATSAPP_EXCEL_SHEETS}"
                ),
            )
        sheet_rows: list[tuple[str, list[tuple[Any, ...]]]] = []
        total_rows = 0
        for sheet in worksheets:
            remaining_rows = MAX_WHATSAPP_EXCEL_ROWS - total_rows
            rows = list(
                islice(
                    sheet.iter_rows(values_only=True),
                    remaining_rows + 1,
                )
            )
            total_rows += len(rows)
            if total_rows > MAX_WHATSAPP_EXCEL_ROWS:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        "The Excel contact file can contain at most "
                        f"{MAX_WHATSAPP_EXCEL_ROWS} rows across all worksheets"
                    ),
                )
            sheet_rows.append((sheet.title, rows))
    except HTTPException:
        raise
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded Excel contact file could not be read",
        ) from exc
    except Exception as exc:
        logger.error(
            "whatsapp_excel_contact_file_read_failed",
            extra={"error_type": type(exc).__name__},
        )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded Excel contact file could not be read",
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()

    if not any(rows for _, rows in sheet_rows):
        return _WhatsAppExcelContactParseResult(
            contacts=[],
            rejected_rows=[],
            rejected_counts={},
        )

    contacts_by_phone: dict[str, WhatsAppRecipientInput] = {}
    rejected_rows: list[WhatsAppContactPreviewRejectedRow] = []
    rejected_counts: dict[WhatsAppContactRejectionCode, int] = {}
    source_order = 0
    for sheet_index, (sheet_name, rows) in enumerate(sheet_rows):
        if not rows:
            continue
        header_match = _find_excel_contact_header(rows)
        if header_match:
            (
                header_row_index,
                phone_columns,
                name_columns,
                given_name_columns,
                surname_columns,
            ) = header_match
            header_row = rows[header_row_index]
            data_rows = rows[header_row_index + 1 :]
            first_data_row_number = header_row_index + 2
        elif sheet_index == 0:
            header_row = ()
            phone_columns = []
            name_columns = []
            given_name_columns = []
            surname_columns = []
            data_rows = rows
            first_data_row_number = 1
        else:
            # A multi-sheet workbook often contains notes or lookup sheets.
            # Never scan those heuristically for phone-like numbers.
            continue

        for row_number, row in enumerate(
            data_rows,
            start=first_data_row_number,
        ):
            row_values = list(row)
            if header_row and _is_repeated_excel_header(row_values, header_row):
                continue
            source_order += 1
            candidates: list[tuple[str | None, str, dict[str, str]]] = []
            imported_fields = (
                _excel_fields_from_row(
                    header_row=header_row,
                    row_values=row_values,
                    sheet_name=sheet_name,
                    source_file_name=source_file_name,
                    row_number=row_number,
                    source_order=source_order,
                )
                if header_row
                else _safe_imported_fields(
                    {
                        "source_file": source_file_name,
                        "source_order": str(source_order),
                        "source_sheet": sheet_name,
                        "source_row": str(row_number),
                    }
                )
            )
            name = _excel_name_from_row(
                row_values,
                name_columns=name_columns,
                given_name_columns=given_name_columns,
                surname_columns=surname_columns,
                phone_columns=phone_columns,
            )
            raw_name = _excel_raw_name_from_row(
                row_values,
                name_columns=name_columns,
                given_name_columns=given_name_columns,
                surname_columns=surname_columns,
                phone_columns=phone_columns,
            )
            if phone_columns:
                phone_values: list[str] = []
                for index in phone_columns:
                    if index >= len(row_values):
                        continue
                    phone = _bounded_excel_raw_value(
                        row_values[index],
                        max_length=64,
                    )
                    if phone:
                        phone_values.append(phone)
                if not phone_values:
                    if _row_has_contact_identity(
                        name=name,
                        imported_fields=imported_fields,
                    ):
                        _append_excel_contact_rejection(
                            rejected_rows,
                            rejected_counts,
                            sheet_name=sheet_name,
                            row_number=row_number,
                            raw_name=raw_name,
                            raw_phone_number=None,
                            imported_fields=imported_fields,
                            reason_code="missing_phone",
                        )
                    continue
                candidates.extend((name, phone, imported_fields) for phone in phone_values)
            else:
                row_text = " ".join(text for cell in row_values if (text := _excel_cell_text(cell)))
                for match in PHONE_RE.findall(row_text):
                    candidates.append((name, match, imported_fields))

            for name, phone, fields in candidates:
                normalized = _normalize_phone(phone)
                if not normalized:
                    _append_excel_contact_rejection(
                        rejected_rows,
                        rejected_counts,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        raw_name=raw_name,
                        raw_phone_number=phone,
                        imported_fields=fields,
                        reason_code="invalid_phone",
                    )
                    continue
                incoming = WhatsAppRecipientInput(
                    name=name,
                    phone_number=phone,
                    imported_fields=fields,
                )
                existing = contacts_by_phone.get(normalized)
                if not name:
                    if existing:
                        contacts_by_phone[normalized] = _merge_recipient_inputs(
                            existing,
                            incoming,
                        )
                    _append_excel_contact_rejection(
                        rejected_rows,
                        rejected_counts,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        raw_name=raw_name,
                        raw_phone_number=phone,
                        imported_fields=fields,
                        reason_code="missing_name",
                    )
                    continue
                if existing:
                    contacts_by_phone[normalized] = _merge_recipient_inputs(
                        existing,
                        incoming,
                    )
                    _append_excel_contact_rejection(
                        rejected_rows,
                        rejected_counts,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        raw_name=raw_name,
                        raw_phone_number=phone,
                        imported_fields=fields,
                        reason_code="duplicate_phone",
                    )
                    continue
                contacts_by_phone[normalized] = incoming
                try:
                    require_whatsapp_recipient_capacity(
                        active_count=0,
                        activating_count=len(contacts_by_phone),
                    )
                except WhatsAppRecipientCapacityExceeded as exc:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=(
                            "The Excel contact file can contain at most "
                            f"{MAX_WHATSAPP_RECIPIENTS} recipients"
                        ),
                    ) from exc
    return _WhatsAppExcelContactParseResult(
        contacts=list(contacts_by_phone.values()),
        rejected_rows=rejected_rows,
        rejected_counts=rejected_counts,
    )


async def _parse_excel_contact_preview(
    upload: UploadFile,
) -> _WhatsAppExcelContactParseResult:
    payload = bytearray()
    while chunk := await upload.read(WHATSAPP_UPLOAD_READ_CHUNK_BYTES):
        payload.extend(chunk)
        if len(payload) > MAX_WHATSAPP_CONTACT_FILE_BYTES:
            raise HTTPException(
                status_code=status.HTTP_413_CONTENT_TOO_LARGE,
                detail=(
                    "The Excel contact file must be "
                    f"{MAX_WHATSAPP_CONTACT_FILE_BYTES // (1024 * 1024)} MB or smaller"
                ),
            )
    filename = upload.filename or "contacts.xlsx"
    return await asyncio.to_thread(
        _parse_excel_contact_bytes,
        bytes(payload),
        filename=filename,
    )


async def _parse_excel_contacts(
    upload: UploadFile,
) -> list[WhatsAppRecipientInput]:
    result = await _parse_excel_contact_preview(upload)
    blocking_rejection_count = sum(
        count
        for reason_code, count in result.rejected_counts.items()
        if reason_code != "duplicate_phone"
    )
    if blocking_rejection_count:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"The Excel contact file contains {blocking_rejection_count} "
                "invalid contact row(s). Preview the file, correct the rejected "
                "rows, and upload it again."
            ),
        )
    return result.contacts


@router.post(
    "/contacts/preview",
    response_model=WhatsAppContactPreviewResponse,
)
async def preview_excel_contacts(
    contacts_file: UploadFile = File(...),
    current_user: User = Depends(require_role(WHATSAPP_ROLES)),
    _auth_transaction_released: None = Depends(_release_auth_transaction),
) -> WhatsAppContactPreviewResponse:
    del current_user, _auth_transaction_released
    result = await _parse_excel_contact_preview(contacts_file)
    return _excel_contact_preview_response(
        result.contacts,
        result.rejected_rows,
        rejected_count=result.rejected_count,
    )
