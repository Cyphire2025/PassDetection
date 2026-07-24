from __future__ import annotations

import io
import unittest
import uuid
from datetime import date

from openpyxl import load_workbook

from app.domain.entities.entities import (
    PassportExtractionStatus,
    PassportSubmission,
)
from app.domain.exceptions.exceptions import ValidationError
from app.domain.value_objects.passport_fields import (
    canonical_passport_fields,
    normalize_extracted_passport_dates,
    normalize_reviewed_passport_fields,
)
from app.infrastructure.export.passport_excel_exporter import PassportExcelExporter


class PassportCaptureFieldTests(unittest.TestCase):
    def test_place_of_issue_never_conflates_legacy_issuing_country(
        self,
    ) -> None:
        normalized = normalize_reviewed_passport_fields({"issuing_country": "  India  "})
        self.assertEqual(normalized, {"issuing_country": "India"})

        legacy_view = canonical_passport_fields({"issuing_country": "India"})
        self.assertEqual(
            legacy_view,
            {
                "issuing_country": "India",
            },
        )
        canonical_view = canonical_passport_fields(
            {
                "place_of_issue": "Chennai",
                "issuing_country": "India",
            }
        )
        self.assertEqual(canonical_view["place_of_issue"], "Chennai")

    def test_explicit_empty_surname_is_preserved_for_ai_and_staff_review(
        self,
    ) -> None:
        fields = normalize_reviewed_passport_fields(
            {
                "surname": "   ",
                "given_names": "MOHIT",
                "passport_number": "W6905713",
            }
        )

        self.assertIn("surname", fields)
        self.assertEqual(fields["surname"], "")
        self.assertEqual(fields["given_names"], "MOHIT")

    def test_reviewed_date_of_issue_is_optional_and_canonical(self) -> None:
        fields = normalize_reviewed_passport_fields(
            {
                "date_of_birth": "1990-01-01",
                "date_of_issue": "2020-06-15",
                "date_of_expiry": "2030-06-14",
            }
        )
        self.assertEqual(fields["date_of_issue"], "2020-06-15")

        without_issue = normalize_reviewed_passport_fields(
            {
                "date_of_birth": "1990-01-01",
                "date_of_expiry": "2030-06-14",
            }
        )
        self.assertNotIn("date_of_issue", without_issue)

    def test_impossible_reviewed_dates_are_rejected(self) -> None:
        with self.assertRaises(ValidationError):
            normalize_reviewed_passport_fields(
                {
                    "date_of_birth": "1990-01-01",
                    "date_of_issue": "1990-01-01",
                    "date_of_expiry": "2030-01-01",
                }
            )
        with self.assertRaises(ValidationError):
            normalize_reviewed_passport_fields(
                {"date_of_issue": date.today().replace(year=2200).isoformat()}
            )

    def test_invalid_extracted_issue_date_is_dropped_without_failing_result(self) -> None:
        normalized = normalize_extracted_passport_dates(
            {
                "passport_number": "P1234567",
                "date_of_issue": "not-a-date",
            }
        )
        self.assertEqual(normalized["passport_number"], "P1234567")
        self.assertNotIn("date_of_issue", normalized)

    def test_manual_review_invalidates_processing_revision_and_state(self) -> None:
        submission = PassportSubmission.create(
            group_id=uuid.uuid4(),
            agency_id=uuid.uuid4(),
            client_name="Traveller",
            client_email=None,
            image_s3_key="drafts/front.jpg",
        )
        processing_revision = submission.mark_processing()
        submission.passport_back_s3_key = "drafts/back.jpg"

        submission.submit_client_review(
            {"passport_number": "P1234567"},
            client_email="person@example.com",
            client_phone="9876543210",
        )

        self.assertGreater(submission.extraction_revision, processing_revision)
        self.assertEqual(
            submission.extraction_status,
            PassportExtractionStatus.READY_FOR_REVIEW,
        )

    def test_export_includes_domestic_airport_and_date_of_issue(self) -> None:
        submission = PassportSubmission.create(
            group_id=uuid.uuid4(),
            agency_id=uuid.uuid4(),
            client_name="Traveller",
            client_email=None,
            image_s3_key="front.jpg",
        )
        submission.submit_client_review(
            {
                "passport_number": "P1234567",
                "date_of_issue": "2020-06-15",
            },
            client_email="person@example.com",
            client_phone="9876543210",
            nearest_domestic_airport="Delhi",
        )

        content = PassportExcelExporter().export_group(
            [submission],
            group_name="Test Group",
        )
        worksheet = load_workbook(io.BytesIO(content), data_only=True).active
        headers = [cell.value for cell in worksheet[4]]
        values = {
            header: worksheet.cell(row=5, column=index + 1).value
            for index, header in enumerate(headers)
        }

        self.assertEqual(values["Nearest Domestic Airport"], "Delhi")
        issue_cell = worksheet.cell(
            row=5,
            column=headers.index("Date of Issue") + 1,
        )
        self.assertTrue(issue_cell.is_date)
        self.assertEqual(issue_cell.number_format, "DD.MM.YYYY")
        self.assertEqual(issue_cell.value.strftime("%d.%m.%Y"), "15.06.2020")


if __name__ == "__main__":
    unittest.main()
