from __future__ import annotations

import io
import uuid
from datetime import date, datetime
from zipfile import ZipFile

from openpyxl import load_workbook

from app.infrastructure.export.meal_plan_excel_exporter import (
    MealPlanExcelExporter,
    MealPlanExportEntry,
    _safe_excel_text,
)


def test_meal_plan_export_has_category_matrix_and_filterable_dish_list() -> None:
    chicken_id = uuid.uuid4()
    paneer_id = uuid.uuid4()
    entries = [
        MealPlanExportEntry(
            day_number=day,
            meal_type=meal_type,
            category_id=category_id,
            category_name=category_name,
            dish_name=f"{category_name} {day} {meal_type}",
        )
        for day in (1, 2)
        for meal_type in ("lunch", "dinner")
        for category_id, category_name in (
            (chicken_id, "Chicken"),
            (paneer_id, "Paneer"),
        )
    ]

    content = MealPlanExcelExporter().export(
        plan_name="Two Day Trip",
        trip_days=2,
        start_date=date(2026, 7, 27),
        selected_category_ids=[chicken_id, paneer_id],
        entries=entries,
    )

    with ZipFile(io.BytesIO(content)) as archive:
        table_parts = sorted(name for name in archive.namelist() if name.startswith("xl/tables/"))
        assert table_parts == []
        assert b"<autoFilter" not in archive.read("xl/worksheets/sheet1.xml")
        assert b"<autoFilter" not in archive.read("xl/worksheets/sheet2.xml")

    workbook = load_workbook(io.BytesIO(content))
    assert workbook.sheetnames == ["Meal Plan", "Dish List"]
    schedule = workbook["Meal Plan"]
    assert schedule["A1"].value == "Two Day Trip"
    assert [schedule.cell(5, column).value for column in range(1, 6)] == [
        "Day",
        "Date",
        "Meal",
        "Chicken",
        "Paneer",
    ]
    assert [schedule.cell(6, column).value for column in range(1, 6)] == [
        "Day 1",
        datetime(2026, 7, 27),
        "Lunch",
        "Chicken 1 lunch",
        "Paneer 1 lunch",
    ]
    assert schedule.auto_filter.ref is None
    assert not schedule.tables
    assert "A6:A7" in schedule.merged_cells
    assert schedule["A6"].value == "Day 1"
    assert schedule["A7"].value is None
    assert all(schedule.cell(8, column).value is None for column in range(1, 6))
    assert schedule.row_dimensions[8].height == 8
    assert "A9:A10" in schedule.merged_cells
    assert schedule["A9"].value == "Day 2"
    assert schedule.sheet_view.showGridLines is True
    assert schedule["A6"].border.left.style == "thin"
    assert schedule["A6"].fill.fgColor.rgb == "00EFF6FF"
    details = workbook["Dish List"]
    assert details.max_row == 10
    assert details.auto_filter.ref is None
    assert not details.tables
    assert "A2:A5" in details.merged_cells
    assert details["A2"].value == "Day 1"
    assert details["A3"].value is None
    assert all(details.cell(6, column).value is None for column in range(1, 7))
    assert details.row_dimensions[6].height == 8
    assert "A7:A10" in details.merged_cells
    assert details["A7"].value == "Day 2"
    assert details.sheet_view.showGridLines is True
    assert details["A2"].border.left.style == "thin"
    assert [details.cell(1, column).value for column in range(1, 7)] == [
        "Day",
        "Date",
        "Meal",
        "Category",
        "Dish",
        "Notes",
    ]


def test_excel_text_is_not_interpreted_as_a_formula() -> None:
    assert _safe_excel_text('=HYPERLINK("https://example.com")') == (
        '\'=HYPERLINK("https://example.com")'
    )
    assert _safe_excel_text("Butter Chicken") == "Butter Chicken"
    assert _safe_excel_text(None) is None
