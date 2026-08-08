from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.infrastructure.export.document_assignment_excel_exporter import (
    build_document_assignment_workbook,
)
from app.presentation.api.v1.routes.document_distribution_review_support import (
    _document_assignment_export_rows,
)


def _document(*, filename: str, delivery_status: str = "pending") -> SimpleNamespace:
    return SimpleNamespace(
        original_filename=filename,
        match_status="matched",
        match_confidence=0.94,
        match_reason="PDF text exact passenger name uniquely matched",
        delivery_status=delivery_status,
        sent_to="+919999999999" if delivery_status == "sent" else None,
        last_sent_at="2026-08-08T10:00:00Z" if delivery_status == "sent" else None,
    )


def _review_rows() -> list[SimpleNamespace]:
    return [
        SimpleNamespace(
            passenger_name="Asha Mehta",
            passport_number="P1234567",
            departure_city="Delhi",
            document=_document(filename="asha.pdf", delivery_status="sent"),
            documents=[_document(filename="asha.pdf", delivery_status="sent")],
        ),
        SimpleNamespace(
            passenger_name="Ravi Shah",
            passport_number="R7654321",
            departure_city="Mumbai",
            document=_document(filename="ravi.pdf"),
            documents=[_document(filename="ravi.pdf")],
        ),
        SimpleNamespace(
            passenger_name="Maya Singh",
            passport_number=None,
            departure_city=None,
            document=None,
            documents=[],
        ),
    ]


def test_document_assignment_export_filters_match_review_tabs() -> None:
    rows = _review_rows()

    assert len(_document_assignment_export_rows(rows, review_filter="all", search_query="")) == 3
    assert len(
        _document_assignment_export_rows(rows, review_filter="assigned", search_query="")
    ) == 2
    assert len(
        _document_assignment_export_rows(rows, review_filter="missing", search_query="")
    ) == 1
    assert len(_document_assignment_export_rows(rows, review_filter="sent", search_query="")) == 1
    assert len(
        _document_assignment_export_rows(rows, review_filter="not_sent", search_query="")
    ) == 1
    searched = _document_assignment_export_rows(
        rows,
        review_filter="all",
        search_query="  MAYA  ",
    )
    assert [row.passenger_name for row in searched] == ["Maya Singh"]


def test_document_assignment_workbook_is_a_real_filtered_xlsx() -> None:
    rows = _document_assignment_export_rows(
        _review_rows(),
        review_filter="assigned",
        search_query="",
    )
    output = build_document_assignment_workbook(
        group_name="South India Group",
        document_label="Domestic onward flight ticket",
        filter_label="Assigned",
        search_query="",
        rows=rows,
    )

    workbook = load_workbook(BytesIO(output.getvalue()))
    sheet = workbook["Document Assignments"]

    assert sheet["B1"].value == "South India Group"
    assert sheet["B3"].value == "Assigned"
    assert sheet.freeze_panes == "A5"
    assert sheet.max_row == 6
    assert [sheet.cell(row=row, column=2).value for row in (5, 6)] == [
        "Asha Mehta",
        "Ravi Shah",
    ]
    assert sheet.tables["DocumentAssignments"].ref == "A4:M6"


def test_document_assignment_workbook_neutralizes_formula_values() -> None:
    rows = _document_assignment_export_rows(
        [
            SimpleNamespace(
                passenger_name="=HYPERLINK(\"https://example.invalid\")",
                passport_number=None,
                departure_city=None,
                document=None,
                documents=[],
            )
        ],
        review_filter="all",
        search_query="",
    )
    output = build_document_assignment_workbook(
        group_name="Group",
        document_label="Visa",
        filter_label="All",
        search_query="",
        rows=rows,
    )

    workbook = load_workbook(BytesIO(output.getvalue()), data_only=False)
    assert workbook["Document Assignments"]["B5"].value.startswith("'=")
