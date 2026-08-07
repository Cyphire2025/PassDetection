from __future__ import annotations

import io
import uuid

import pytest
from openpyxl import load_workbook

from app.domain.entities.entities import PassportSubmission
from app.domain.value_objects.personnel_codes import prefixed_staff_code
from app.infrastructure.export.passport_excel_exporter import (
    PassportExcelExporter,
    _country_display_name,
    _gender_display_value,
    _nationality_display_value,
    _safe_xlsx_value,
    passport_age_group,
)

_OPTION_FLAGS = {
    "nearest_international_airport_enabled": False,
    "ask_nearest_domestic_airport": False,
    "base_city_enabled": False,
    "staff_code_enabled": False,
    "agent_employee_code_enabled": False,
    "meal_preference_enabled": False,
    "relation_with_qualifier_enabled": False,
    "designation_enabled": False,
    "agency_dealership_name_enabled": False,
}


def _submission(
    group_id: uuid.UUID,
    *,
    client_name: str = "Traveller",
    client_phone: str = "9876543210",
    fields: dict[str, str] | None = None,
) -> PassportSubmission:
    submission = PassportSubmission.create(
        group_id=group_id,
        agency_id=uuid.uuid4(),
        client_name=client_name,
        client_email=None,
        image_s3_key="front.jpg",
    )
    review_fields = {"given_names": client_name}
    review_fields.update(fields or {})
    submission.submit_client_review(
        review_fields,
        client_email="traveller@example.com",
        client_phone=client_phone,
        departure_city="Delhi",
        nearest_domestic_airport="Indira Gandhi International Airport",
    )
    return submission


def _worksheet(content: bytes):
    return load_workbook(io.BytesIO(content), data_only=False).active


def _row_values(worksheet) -> tuple[list[str], dict[str, object]]:
    headers = [cell.value for cell in worksheet[4]]
    return headers, {
        header: worksheet.cell(row=5, column=index + 1).value
        for index, header in enumerate(headers)
    }


def test_export_omits_internal_and_disabled_columns_and_formats_identity() -> None:
    group_id = uuid.uuid4()
    submission = _submission(
        group_id,
        fields={
            "surname": "vAsHiStHa",
            "given_names": "niPun kuMar",
            "passport_number": "W1234567",
            "nationality": "IND",
            "issuing_country": "IND",
            "base_city": "Delhi",
            "staff_code": "GC-42",
            "meal_preference": "Veg",
        },
    )

    content = PassportExcelExporter().export_group(
        [submission],
        group_name="Test Group",
        group_details={
            group_id: {
                "name": "Test Group",
                **_OPTION_FLAGS,
            }
        },
    )
    headers, values = _row_values(_worksheet(content))

    assert {
        "Submitted At",
        "Reviewed At",
        "Confidence",
        "Issuing Country",
        "Status",
        "Nearest International Airport",
        "Nearest Domestic Airport",
        "Domestic Airport",
        "Base City",
        "Staff Code",
        "Agent/Employee Code",
        "Agency/Dealership Name",
        "Designation",
        "Meal Preference",
        "Relation with Qualifier",
    }.isdisjoint(headers)
    assert values["SURNAME"] == "VASHISTHA"
    assert values["GIVEN NAME"] == "NIPUN KUMAR"
    assert "International Airport" not in headers
    assert values["Nationality"] == "Indian"
    assert "Place of Issue" in headers


def test_export_includes_canonical_place_of_issue() -> None:
    group_id = uuid.uuid4()
    submission = _submission(
        group_id,
        fields={
            "place_of_issue": "CHENNAI",
            "issuing_country": "IND",
        },
    )

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submission],
            group_name="Test Group",
            group_details={group_id: {"name": "Test Group", **_OPTION_FLAGS}},
        )
    )
    headers, values = _row_values(worksheet)

    assert "Place of Issue" in headers
    assert values["Place of Issue"] == "CHENNAI"


def test_agency_matched_export_places_old_given_name_before_new_names() -> None:
    group_id = uuid.uuid4()
    submission = _submission(
        group_id,
        fields={
            "surname": "Khan",
            "given_names": "Abdul Hameed",
        },
    )

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submission],
            group_name="Agency Match",
            group_details={
                group_id: {
                    "name": "Agency Match",
                    **_OPTION_FLAGS,
                    "agency_dealership_name_enabled": True,
                }
            },
            previous_names={
                submission.id: {
                    "surname": "Samsuddin",
                    "given_names": "Mubassreen",
                }
            },
            additional_fields=[
                {"key": "whatsapp:agency", "label": "Old agency"},
            ],
            additional_values={
                submission.id: {
                    "whatsapp:agency": "North Star Motors",
                }
            },
        )
    )
    headers, values = _row_values(worksheet)

    assert "SURNAME" not in headers
    assert "GIVEN NAME" not in headers
    assert "Old Surname" not in headers
    assert headers.index("Old agency") < headers.index("Old Given Name")
    assert headers.index("Old Given Name") < headers.index("New Surname")
    assert headers.index("New Surname") < headers.index("New Given Name")
    assert values["Old agency"] == "North Star Motors"
    assert values["Old Given Name"] == "MUBASSREEN"
    assert values["New Surname"] == "KHAN"
    assert values["New Given Name"] == "ABDUL HAMEED"
    assert all(
        worksheet.cell(row=5, column=headers.index(header) + 1).fill.fill_type
        is None
        for header in ("Old Given Name", "New Surname", "New Given Name")
    )


@pytest.mark.parametrize("source", ("IN", "IND", "India", "indian"))
def test_export_formats_indian_nationality_without_changing_country_labels(
    source: str,
) -> None:
    assert _nationality_display_value(source) == "Indian"
    assert _country_display_name(source if source != "indian" else "India") == "India"


def test_export_includes_only_group_options_enabled_in_the_workbook() -> None:
    first_group_id = uuid.uuid4()
    second_group_id = uuid.uuid4()
    first = _submission(first_group_id)
    second = _submission(
        second_group_id,
        fields={
            "base_city": "Mumbai",
            "staff_code": "GC-77",
            "agent_employee_type": "agent",
            "agent_employee_code": "9988",
            "meal_preference": "Jain",
        },
    )

    content = PassportExcelExporter().export_group(
        [first, second],
        group_name="Selected Groups",
        group_details={
            first_group_id: {
                "name": "First",
                **_OPTION_FLAGS,
            },
            second_group_id: {
                "name": "Second",
                **_OPTION_FLAGS,
                "base_city_enabled": True,
                "staff_code_enabled": True,
                "agent_employee_code_enabled": True,
                "meal_preference_enabled": True,
                "nearest_international_airport_enabled": True,
            },
        },
    )
    worksheet = _worksheet(content)
    headers = [cell.value for cell in worksheet[4]]
    group_column = headers.index("Group") + 1
    second_row = next(
        row
        for row in range(5, worksheet.max_row + 1)
        if worksheet.cell(row=row, column=group_column).value == "Second"
    )
    second_values = {
        header: worksheet.cell(row=second_row, column=index + 1).value
        for index, header in enumerate(headers)
    }

    assert "International Airport" in headers
    assert "Domestic Airport" not in headers
    assert "Base City" in headers
    assert "Staff Code" in headers
    assert "Agent/Employee Code" in headers
    assert "Meal Preference" in headers
    assert second_values["Base City"] == "Mumbai"
    assert second_values["Staff Code"] == "STF_GC-77"
    assert second_values["Agent/Employee Code"] == "AGT_9988"
    assert second_values["Meal Preference"] == "Jain"


def test_pending_only_group_can_enable_optional_columns() -> None:
    submitted_group_id = uuid.uuid4()
    pending_group_id = uuid.uuid4()

    content = PassportExcelExporter().export_group(
        [_submission(submitted_group_id)],
        group_name="Selected Groups",
        group_details={
            submitted_group_id: {
                "name": "Submitted Group",
                **_OPTION_FLAGS,
            },
            pending_group_id: {
                "name": "Pending Group",
                **_OPTION_FLAGS,
                "staff_code_enabled": True,
            },
        },
        pending_rows=[
            {
                "Group": "Pending Group",
                "GIVEN NAME": "Pending Traveller",
                "Staff Code": "STF_42",
            }
        ],
    )
    worksheet = _worksheet(content)
    headers = [cell.value for cell in worksheet[4]]

    assert "Staff Code" in headers
    pending_row = next(
        row
        for row in range(5, worksheet.max_row + 1)
        if worksheet.cell(
            row=row,
            column=headers.index("GIVEN NAME") + 1,
        ).value
        == "Pending Traveller"
    )
    assert worksheet.cell(
        row=pending_row,
        column=headers.index("Staff Code") + 1,
    ).value == "STF_42"


def test_export_groups_exact_zone_names_with_two_blank_rows_between_zones() -> None:
    group_id = uuid.uuid4()
    delhi_two = _submission(group_id, client_name="Zed", fields={"passport_number": "D2"})
    mumbai_two = _submission(group_id, client_name="Beta", fields={"passport_number": "M2"})
    delhi_one = _submission(group_id, client_name="Alpha", fields={"passport_number": "D1"})
    mumbai_one = _submission(group_id, client_name="Alpha", fields={"passport_number": "M1"})
    without_zone = _submission(group_id, client_name="No Zone", fields={"passport_number": "NZ"})
    zone_names = {
        delhi_two.id: "Delhi",
        delhi_one.id: "Delhi",
        mumbai_one.id: "Mumbai-1",
        mumbai_two.id: "Mumbai-2",
    }

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [mumbai_two, delhi_two, without_zone, mumbai_one, delhi_one],
            group_name="Zone Group",
            group_details={group_id: {"name": "Zone Group", **_OPTION_FLAGS}},
            zone_names=zone_names,
            additional_fields=[{"key": "zone_name", "label": "Zone Name"}],
            group_by_field="zone_name",
        )
    )
    headers = [cell.value for cell in worksheet[4]]
    zone_column = headers.index("Zone Name") + 1
    name_column = headers.index("GIVEN NAME") + 1
    data_or_gap_rows = [
        (
            worksheet.cell(row=row, column=zone_column).value,
            worksheet.cell(row=row, column=name_column).value,
        )
        for row in range(5, worksheet.max_row + 1)
    ]

    assert data_or_gap_rows == [
        ("Delhi", "ALPHA"),
        ("Delhi", "ZED"),
        (None, None),
        (None, None),
        ("Mumbai-1", "ALPHA"),
        (None, None),
        (None, None),
        ("Mumbai-2", "BETA"),
        (None, None),
        (None, None),
        (None, "NO ZONE"),
    ]


def test_overall_export_uses_one_filterable_table_and_yellow_pending_rows() -> None:
    group_id = uuid.uuid4()
    submitted_mumbai = _submission(
        group_id,
        client_name="Submitted Mumbai",
        fields={"passport_number": "M100"},
    )
    submitted_delhi = _submission(
        group_id,
        client_name="Submitted Delhi",
        fields={"passport_number": "D100"},
    )
    pending_rows = [
        {
            "GIVEN NAME": "Pending No Zone",
            "WhatsApp Email": "no-zone@example.com",
            "WhatsApp Phone": "9000000005",
        },
        {
            "GIVEN NAME": "Pending Mumbai Two",
            "Zone Name": "Mumbai-2",
            "WhatsApp Email": "mumbai-two@example.com",
            "WhatsApp Phone": "9000000004",
        },
        {
            "GIVEN NAME": "Pending Delhi Zed",
            "Zone Name": "Delhi",
            "WhatsApp Email": "delhi-zed@example.com",
            "WhatsApp Phone": "9000000002",
        },
        {
            "GIVEN NAME": "Pending Mumbai One",
            "Zone Name": "Mumbai-1",
            "WhatsApp Email": "mumbai-one@example.com",
            "WhatsApp Phone": "9000000003",
        },
        {
            "GIVEN NAME": "Pending Delhi Alpha",
            "Zone Name": "Delhi",
            "WhatsApp Email": "delhi-alpha@example.com",
            "WhatsApp Phone": "9000000001",
        },
    ]

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submitted_mumbai, submitted_delhi],
            group_name="Overall Group",
            group_details={group_id: {"name": "Overall Group", **_OPTION_FLAGS}},
            zone_names={
                submitted_mumbai.id: "Mumbai",
                submitted_delhi.id: "Delhi",
            },
            additional_fields=[{"key": "zone_name", "label": "Zone Name"}],
            group_by_field="zone_name",
            pending_rows=pending_rows,
        )
    )
    headers = [cell.value for cell in worksheet[4]]
    name_column = headers.index("GIVEN NAME") + 1
    zone_column = headers.index("Zone Name") + 1
    submitted_rows = [
        row
        for row in range(5, worksheet.max_row + 1)
        if (
            worksheet.cell(row=row, column=name_column).value
            in {"SUBMITTED DELHI", "SUBMITTED MUMBAI"}
        )
    ]

    assert [
        worksheet.cell(row=row, column=name_column).value for row in submitted_rows
    ] == ["SUBMITTED DELHI", "SUBMITTED MUMBAI"]

    pending_data_start = submitted_rows[-1] + 1
    pending_data_or_gap_rows = [
        (
            worksheet.cell(row=row, column=zone_column).value,
            worksheet.cell(row=row, column=name_column).value,
        )
        for row in range(pending_data_start, worksheet.max_row + 1)
    ]
    assert pending_data_or_gap_rows == [
        ("Delhi", "Pending Delhi Alpha"),
        ("Delhi", "Pending Delhi Zed"),
        (None, None),
        (None, None),
        ("Mumbai-1", "Pending Mumbai One"),
        (None, None),
        (None, None),
        ("Mumbai-2", "Pending Mumbai Two"),
        (None, None),
        (None, None),
        (None, "Pending No Zone"),
    ]
    pending_data_rows = [
        row
        for row in range(pending_data_start, worksheet.max_row + 1)
        if worksheet.cell(row=row, column=name_column).value
    ]
    for row in pending_data_rows:
        assert all(
            cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb == "00FFF2CC"
            for cell in worksheet[row]
        )

    assert set(worksheet.tables) == {"PassportSubmissions"}
    assert worksheet.tables["PassportSubmissions"].ref.endswith(
        str(worksheet.max_row)
    )
    assert worksheet.auto_filter.ref is None
    assert all(
        merged.min_row in {1, 2}
        for merged in worksheet.merged_cells.ranges
    )


def test_selected_export_is_unchanged_when_pending_rows_are_not_provided() -> None:
    group_id = uuid.uuid4()
    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [_submission(group_id, client_name="Selected Traveller")],
            group_name="Selected Passports",
            group_details={group_id: {"name": "Selected Passports", **_OPTION_FLAGS}},
        )
    )

    assert all(
        worksheet.cell(row=row, column=1).value != "PENDING"
        for row in range(1, worksheet.max_row + 1)
    )
    assert set(worksheet.tables) == {"PassportSubmissions"}


@pytest.mark.parametrize(
    ("person_type", "expected"),
    (("agent", "AGT_12345"), ("employee", "EMP_12345")),
)
def test_export_prefixes_agent_and_employee_codes(
    person_type: str,
    expected: str,
) -> None:
    group_id = uuid.uuid4()
    submission = _submission(
        group_id,
        fields={
            "agent_employee_type": person_type,
            "agent_employee_code": "12345",
        },
    )

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submission],
            group_name="Code Group",
            group_details={
                group_id: {
                    "name": "Code Group",
                    **_OPTION_FLAGS,
                    "agent_employee_code_enabled": True,
                }
            },
        )
    )
    _, values = _row_values(worksheet)

    assert values["Agent/Employee Code"] == expected


@pytest.mark.parametrize(
    ("source", "expected"),
    (("42", "STF_42"), ("STF_42", "STF_42"), ("STF-42", "STF_42")),
)
def test_staff_code_prefix_is_canonical_and_not_duplicated(
    source: str,
    expected: str,
) -> None:
    assert prefixed_staff_code(source) == expected


def test_export_includes_relation_snapshot_only_for_enabled_groups() -> None:
    group_id = uuid.uuid4()
    submission = _submission(group_id)
    submission.qualifier_enabled_snapshot = True
    submission.qualifier_is_self = False
    submission.qualifier_relation_code = "spouse"
    submission.qualifier_relation_label = "Spouse"

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submission],
            group_name="Qualifier Group",
            group_details={
                group_id: {
                    "name": "Qualifier Group",
                    **_OPTION_FLAGS,
                    "relation_with_qualifier_enabled": True,
                }
            },
        )
    )
    headers, values = _row_values(worksheet)

    assert "Relation with Qualifier" in headers
    assert values["Relation with Qualifier"] == "Spouse"


def test_export_neutralizes_formula_like_text_after_leading_whitespace() -> None:
    group_id = uuid.uuid4()
    submission = _submission(
        group_id,
        client_name="   =HYPERLINK(\"https://example.test\",\"click\")",
        fields={
            "surname": "+cmd",
            "given_names": "-2+3",
            "passport_number": "@SUM(A1:A2)",
        },
    )

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submission],
            group_name="Test Group",
            group_details={group_id: {"name": "Test Group", **_OPTION_FLAGS}},
        )
    )
    headers, values = _row_values(worksheet)

    for header in ("SURNAME", "GIVEN NAME", "Passport Number"):
        cell = worksheet.cell(row=5, column=headers.index(header) + 1)
        assert cell.data_type == "s"
        assert str(values[header]).startswith("'")
    assert _safe_xlsx_value(42) == 42
    assert _safe_xlsx_value("2026-07-17") == "2026-07-17"


def test_export_preserves_dynamic_header_text_without_formula_typing() -> None:
    group_id = uuid.uuid4()
    formula_label = '=HYPERLINK("https://example.test","Review")'
    submission = _submission(group_id)
    submission.custom_answers = [
        {
            "question_id": "formula-header",
            "label": formula_label,
            "answer": "ordinary answer",
        }
    ]

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submission],
            group_name="Formula Header Group",
            group_details={
                group_id: {
                    "name": "Formula Header Group",
                    **_OPTION_FLAGS,
                    "custom_questions": [
                        {
                            "id": "formula-header",
                            "label": formula_label,
                            "enabled": True,
                        }
                    ],
                }
            },
        )
    )
    header_cell = next(cell for cell in worksheet[4] if cell.value == formula_label)

    assert header_cell.value == formula_label
    assert header_cell.data_type == "s"


@pytest.mark.parametrize(
    ("source", "expected"),
    (
        ("M", "Male"),
        ("m", "Male"),
        ("Male", "Male"),
        ("male", "Male"),
        (" F ", "Female"),
        ("f", "Female"),
        ("Female", "Female"),
        ("female", "Female"),
    ),
)
def test_export_normalizes_supported_gender_values(
    source: str,
    expected: str,
) -> None:
    group_id = uuid.uuid4()
    submission = _submission(group_id, fields={"sex": source})

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submission],
            group_name="Test Group",
            group_details={group_id: {"name": "Test Group", **_OPTION_FLAGS}},
        )
    )
    _, values = _row_values(worksheet)

    assert values["GENDER"] == expected


def test_export_gender_normalization_does_not_invent_unknown_values() -> None:
    assert _gender_display_value("X") == "X"
    assert _gender_display_value(None) is None


def test_export_writes_all_visible_dates_as_native_dd_dot_mm_dot_yyyy() -> None:
    group_id = uuid.uuid4()
    submission = _submission(
        group_id,
        fields={
            "date_of_birth": "1972-08-30",
            "date_of_issue": "2023-08-10",
            "date_of_expiry": "2033-08-09",
        },
    )

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submission],
            group_name="Test Group",
            group_details={
                group_id: {
                    "name": "Test Group",
                    "travel_date": "2026-07-17",
                    "return_date": "2026-07-25",
                    **_OPTION_FLAGS,
                }
            },
        )
    )
    headers = [cell.value for cell in worksheet[4]]
    expected = {
        "Travel/Departure Date": "17.07.2026",
        "Return Date": "25.07.2026",
        "DOB": "30.08.1972",
        "DOI": "10.08.2023",
        "DOE": "09.08.2033",
    }

    for header, display_value in expected.items():
        cell = worksheet.cell(row=5, column=headers.index(header) + 1)
        assert cell.is_date
        assert cell.number_format == "DD.MM.YYYY"
        assert cell.value.strftime("%d.%m.%Y") == display_value


def test_dynamic_export_places_selected_whatsapp_fields_after_trip_columns() -> None:
    group_id = uuid.uuid4()
    submission = _submission(group_id)
    question_id = uuid.uuid4()
    submission.custom_answers = [
        {
            "question_id": str(question_id),
            "label": "Excursion",
            "value": "City tour",
        }
    ]

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submission],
            group_name="Test Group",
            group_details={
                group_id: {
                    "name": "Test Group",
                    **_OPTION_FLAGS,
                    "custom_questions": [
                        {
                            "id": str(question_id),
                            "label": "Excursion",
                            "options": ["City tour"],
                            "enabled": True,
                        }
                    ],
                }
            },
            zone_names={submission.id: "West 1"},
            additional_fields=[
                {"key": "zone_name", "label": "Zone Name"},
                {"key": "whatsapp:t_shirt_size", "label": "T Shirt Size"},
            ],
            additional_values={
                submission.id: {
                    "whatsapp:t_shirt_size": "Large",
                }
            },
            group_by_field="zone_name",
        )
    )
    headers, values = _row_values(worksheet)

    assert headers[4:6] == ["Zone Name", "T Shirt Size"]
    assert headers[-1] == "Excursion"
    assert values["Zone Name"] == "West 1"
    assert values["T Shirt Size"] == "Large"
    assert values["Excursion"] == "City tour"


def test_dynamic_export_can_omit_zone_and_group_by_another_saved_field() -> None:
    group_id = uuid.uuid4()
    beta = _submission(group_id, client_name="Beta")
    alpha = _submission(group_id, client_name="Alpha")

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [beta, alpha],
            group_name="Test Group",
            group_details={group_id: {"name": "Test Group", **_OPTION_FLAGS}},
            additional_fields=[
                {"key": "whatsapp:session", "label": "Session"},
            ],
            additional_values={
                beta.id: {"whatsapp:session": "B"},
                alpha.id: {"whatsapp:session": "A"},
            },
            group_by_field="whatsapp:session",
        )
    )
    headers = [cell.value for cell in worksheet[4]]

    assert "Zone Name" not in headers
    assert headers[4] == "Session"
    assert headers[-1] == "Upload Phone"
    assert worksheet.cell(row=5, column=headers.index("GIVEN NAME") + 1).value == "ALPHA"
    assert worksheet.cell(row=8, column=headers.index("GIVEN NAME") + 1).value == "BETA"


@pytest.mark.parametrize(
    ("whatsapp_phone", "expected"),
    (
        ("+919876543210", "9876543210"),
        ("'+919876543210", "9876543210"),
        ("+91 9876543210", "9876543210"),
        ("+91-9876543210", "9876543210"),
        ("'+91 9876543210", "9876543210"),
        ("'+91-9876543210", "9876543210"),
        ("+12025550123", "'+12025550123"),
        ("919876543210", "919876543210"),
    ),
)
def test_export_removes_only_explicit_india_code_from_whatsapp_phone(
    whatsapp_phone: str,
    expected: str,
) -> None:
    group_id = uuid.uuid4()
    submission = _submission(
        group_id,
        client_phone="+919111111111",
    )

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submission],
            group_name="Phone Normalization",
            group_details={
                group_id: {
                    "name": "Phone Normalization",
                    **_OPTION_FLAGS,
                }
            },
            whatsapp_contacts={
                submission.id: {
                    "email": "broadcast@example.com",
                    "phone": whatsapp_phone,
                }
            },
        )
    )
    _, values = _row_values(worksheet)

    assert values["WhatsApp Phone"] == expected
    assert values["Upload Phone"] == "'+919111111111"


@pytest.mark.parametrize(
    "pending_phone",
    (
        "+919222222222",
        "'+919222222222",
        "+91 9222222222",
        "'+91-9222222222",
    ),
)
def test_pending_whatsapp_phone_uses_the_same_india_code_normalization(
    pending_phone: str,
) -> None:
    group_id = uuid.uuid4()
    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [_submission(group_id)],
            group_name="Pending Phone Normalization",
            group_details={
                group_id: {
                    "name": "Pending Phone Normalization",
                    **_OPTION_FLAGS,
                }
            },
            pending_rows=[
                {
                    "GIVEN NAME": "Pending Traveller",
                    "WhatsApp Phone": pending_phone,
                    "Upload Phone": "+919333333333",
                }
            ],
        )
    )
    headers = [cell.value for cell in worksheet[4]]
    pending_row = next(
        row
        for row in range(5, worksheet.max_row + 1)
        if worksheet.cell(
            row=row,
            column=headers.index("GIVEN NAME") + 1,
        ).value
        == "Pending Traveller"
    )

    assert worksheet.cell(
        row=pending_row,
        column=headers.index("WhatsApp Phone") + 1,
    ).value == "9222222222"
    assert worksheet.cell(
        row=pending_row,
        column=headers.index("Upload Phone") + 1,
    ).value == "'+919333333333"


def test_export_uses_the_requested_exact_column_order() -> None:
    group_id = uuid.uuid4()
    question_id = uuid.uuid4()
    detail_id = uuid.uuid4()
    submission = _submission(
        group_id,
        fields={
            "agency_dealership_name": "North Agency",
            "staff_code": "77",
            "designation": "Manager",
            "base_city": "Mumbai",
            "meal_preference": "Veg",
            "surname": "Shah",
            "given_names": "Nipun",
            "sex": "M",
            "passport_number": "P1234567",
            "date_of_birth": "1990-08-15",
            "date_of_issue": "2023-01-02",
            "date_of_expiry": "2033-01-01",
            "nationality": "IND",
            "place_of_issue": "Mumbai",
        },
    )
    submission.qualifier_enabled_snapshot = True
    submission.qualifier_relation_label = "Self"
    submission.custom_answers = [
        {
            "question_id": str(question_id),
            "label": "Activity",
            "value": "Workshop",
        }
    ]
    submission.custom_detail_answers = [
        {
            "detail_id": str(detail_id),
            "label": "Badge name",
            "value": "Nipun S.",
        }
    ]

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [submission],
            group_name="Exact Order",
            group_details={
                group_id: {
                    "name": "Exact Order",
                    "destination": "Vietnam",
                    "travel_date": "2026-09-10",
                    "return_date": "2026-09-18",
                    **_OPTION_FLAGS,
                    "agency_dealership_name_enabled": True,
                    "staff_code_enabled": True,
                    "designation_enabled": True,
                    "relation_with_qualifier_enabled": True,
                    "base_city_enabled": True,
                    "ask_nearest_domestic_airport": True,
                    "meal_preference_enabled": True,
                    "nearest_international_airport_enabled": True,
                    "custom_questions": [
                        {
                            "id": str(question_id),
                            "label": "Activity",
                            "enabled": True,
                        }
                    ],
                    "custom_details": [
                        {
                            "id": str(detail_id),
                            "label": "Badge name",
                            "enabled": True,
                        }
                    ],
                }
            },
            zone_names={submission.id: "West 1"},
            additional_fields=[
                {"key": "zone_name", "label": "Zone Name"},
                {"key": "whatsapp:department", "label": "Department"},
            ],
            additional_values={
                submission.id: {"whatsapp:department": "Sales"},
            },
            whatsapp_contacts={
                submission.id: {
                    "email": "broadcast@example.com",
                    "phone": "+919000000001",
                }
            },
        )
    )
    headers, values = _row_values(worksheet)

    assert headers == [
        "Group",
        "Destination",
        "Travel/Departure Date",
        "Return Date",
        "Zone Name",
        "Department",
        "Agency/Dealership Name",
        "Staff Code",
        "Designation",
        "Relation with Qualifier",
        "Age Group",
        "Base City",
        "Domestic Airport",
        "WhatsApp Email",
        "WhatsApp Phone",
        "Meal Preference",
        "International Airport",
        "SURNAME",
        "GIVEN NAME",
        "GENDER",
        "Passport Number",
        "DOB",
        "DOI",
        "DOE",
        "Place of Issue",
        "Nationality",
        "Upload Email",
        "Upload Phone",
        "Activity",
        "Badge name",
    ]
    assert values["Age Group"] == "Adult"
    assert values["WhatsApp Email"] == "broadcast@example.com"
    assert values["WhatsApp Phone"] == "9000000001"
    assert values["Upload Email"] == "traveller@example.com"
    assert values["Upload Phone"] == "9876543210"
    assert values["Activity"] == "Workshop"
    assert values["Badge name"] == "Nipun S."
    assert "Agent/Employee Code" not in headers


@pytest.mark.parametrize(
    ("birth_date", "expected"),
    (
        ("2026-07-26", "Infant"),
        ("2024-07-27", "Infant"),
        ("2024-07-26", "Child"),
        ("2014-07-27", "Child"),
        ("2014-07-26", "Adult"),
    ),
)
def test_age_group_uses_completed_age_on_departure(
    birth_date: str,
    expected: str,
) -> None:
    assert passport_age_group(birth_date, "2026-07-26") == expected


def test_export_groups_by_fixed_international_airport() -> None:
    group_id = uuid.uuid4()
    mumbai = _submission(group_id, client_name="Mumbai Traveller")
    delhi = _submission(group_id, client_name="Delhi Traveller")
    mumbai.departure_city = "Mumbai"
    delhi.departure_city = "Delhi"

    worksheet = _worksheet(
        PassportExcelExporter().export_group(
            [mumbai, delhi],
            group_name="Airport Group",
            group_details={
                group_id: {
                    "name": "Airport Group",
                    **_OPTION_FLAGS,
                    "nearest_international_airport_enabled": True,
                }
            },
            group_by_field="international_airport",
        )
    )
    headers = [cell.value for cell in worksheet[4]]
    airport_column = headers.index("International Airport") + 1

    assert worksheet.cell(row=5, column=airport_column).value == "Delhi"
    assert worksheet.cell(row=8, column=airport_column).value == "Mumbai"
