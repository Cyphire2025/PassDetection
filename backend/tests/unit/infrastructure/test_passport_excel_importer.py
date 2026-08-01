from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

import pytest
from openpyxl import Workbook

import app.infrastructure.imports.passport_excel_importer as importer_module
from app.infrastructure.imports.passport_excel_importer import PassportExcelImporter


def _excel_bytes(
    headers: list[object],
    rows: list[list[object]],
    *,
    title: str = "Passengers",
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()


def _streaming_excel_bytes(
    headers: list[object],
    rows: list[list[object]],
) -> bytes:
    """Create a valid workbook without optional worksheet dimensions."""

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Passengers")
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    asm = workbook.active
    asm.title = "ASM"
    asm.append(
        [
            "Zone Name",
            "Staffname",
            "StaffCode",
            "Designation",
            "PASSPORT_NO",
            "DOB",
            "DOI",
            "Passport Expiry Date",
        ]
    )
    asm.append(
        [
            "ASSAM",
            "BIPLAB DAS",
            25523,
            "ACE",
            "Z4160891",
            datetime(1979, 9, 5),
            date(2021, 6, 14),
            datetime(2031, 6, 13),
        ]
    )

    delhi = workbook.create_sheet("DE1")
    delhi.append(["Zone Name", "Staff Name", "Staff Code", "Designation", "Gender"])
    delhi.append(["DELHI", "MADHVI KASHYAP", 25293, "ACE", "F"])

    ignored = workbook.create_sheet("Notes")
    ignored.append(["This sheet has no import headers"])
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()


def test_imports_every_worksheet_and_retains_staff_metadata() -> None:
    rows = PassportExcelImporter().import_rows(_workbook_bytes())

    assert len(rows) == 2
    assert rows[0].worksheet_name == "ASM"
    assert rows[0].confirmed_fields == {
        "staff_code": "25523",
        "passport_number": "Z4160891",
        "date_of_birth": "1979-09-05",
        "date_of_issue": "2021-06-14",
        "date_of_expiry": "2031-06-13",
    }
    assert rows[0].staff_metadata["zone_name"] == "ASSAM"
    assert rows[0].staff_metadata["designation"] == "ACE"
    assert rows[0].staff_metadata["source_sheet"] == "ASM"
    assert rows[1].client_name == "MADHVI KASHYAP"
    assert rows[1].confirmed_fields["sex"] == "F"
    assert rows[1].staff_metadata["gender"] == "F"
    assert rows[1].staff_metadata["staff_code"] == "25293"


def test_explicit_passport_name_parts_outrank_competing_generic_names() -> None:
    content = _excel_bytes(
        [
            "Name",
            "Staffname",
            "GIVEN_NAME",
            "SURNAME",
            "Mobile",
            "GENDER",
            "D.O.B.",
            "D_O_I",
            "D O E",
            "Passport_Number",
            "Agent Code",
        ],
        [
            [
                "Wrong Generic Name",
                "Wrong Staff Name",
                "UMESHBHAI AMRUTBHAI",
                "MAHYAVANSHI",
                9099911724,
                "FEMALE",
                "19.01.1977",
                "27/02/2018",
                "26 02 2028",
                "z 47-79216",
                "AG-42",
            ]
        ],
    )

    row = PassportExcelImporter().import_rows(content)[0]

    assert row.client_name == "UMESHBHAI AMRUTBHAI MAHYAVANSHI"
    assert row.client_phone == "9099911724"
    assert row.confirmed_fields == {
        "given_names": "UMESHBHAI AMRUTBHAI",
        "surname": "MAHYAVANSHI",
        "passport_number": "Z4779216",
        "sex": "F",
        "date_of_birth": "1977-01-19",
        "date_of_issue": "2018-02-27",
        "date_of_expiry": "2028-02-26",
    }
    assert row.staff_metadata["name"] == "Wrong Generic Name"
    assert row.staff_metadata["staffname"] == "Wrong Staff Name"
    assert row.staff_metadata["agent_code"] == "AG-42"
    assert row.staff_metadata["gender"] == "FEMALE"


def test_integral_float_identifiers_do_not_gain_decimal_digits() -> None:
    importer = PassportExcelImporter()

    passport_text = importer._stringify(4779216.0)
    staff_code_text = importer._stringify(25073.0)
    phone_text = importer._stringify(9099911724.0)

    assert passport_text == "4779216"
    assert staff_code_text == "25073"
    assert phone_text == "9099911724"
    assert importer._confirmed_fields(
        {
            "passport_number": passport_text,
            "staff_code": staff_code_text,
        }
    ) == {
        "passport_number": "4779216",
        "staff_code": "25073",
    }


@pytest.mark.parametrize(
    "excel_error",
    ["#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NUM!", "#SPILL!"],
)
def test_excel_error_cells_never_become_passport_identities(excel_error: str) -> None:
    row = PassportExcelImporter().import_rows(
        _excel_bytes(
            ["GIVEN NAME", "SURNAME", "PASSPORT_NO"],
            [["Asha", "Rao", excel_error]],
        )
    )[0]

    assert "passport_number" not in row.confirmed_fields
    assert row.staff_metadata["passport_no"] == excel_error


def test_explicitly_blank_surname_is_preserved_without_treating_null_markers_as_blank() -> None:
    rows = PassportExcelImporter().import_rows(
        _excel_bytes(
            ["GIVEN_NAME", "SURNAME", "Passport No"],
            [
                ["Asha", None, "P1234567"],
                ["Bina", "N/A", "P7654321"],
            ],
        )
    )

    assert rows[0].client_name == "Asha"
    assert rows[0].confirmed_fields["surname"] == ""
    assert "surname" not in rows[1].confirmed_fields


def test_formula_without_cached_value_is_not_mistaken_for_blank_surname() -> None:
    row = PassportExcelImporter().import_rows(
        _excel_bytes(
            ["GIVEN_NAME", "SURNAME", "Passport No"],
            [["Asha", '=CONCAT("R","ao")', "P1234567"]],
        )
    )[0]

    assert row.client_name == "Asha"
    assert "surname" not in row.confirmed_fields
    assert row.confirmed_fields["passport_number"] == "P1234567"


@pytest.mark.parametrize(
    ("birth_header", "issue_header", "expiry_header"),
    [
        ("DOB", "DOI", "DOE"),
        ("D_O_B", "D_O_I", "D_O_E"),
        ("D O B", "D O I", "D O E"),
        ("Date_Of_Birth", "Date Of Issue", "Date-Of-Expiry"),
        ("Birth Date", "Issue Date", "Expiration Date"),
    ],
)
def test_date_header_variants_use_day_first_passport_semantics(
    birth_header: str,
    issue_header: str,
    expiry_header: str,
) -> None:
    content = _excel_bytes(
        ["Given Name", "Surname", birth_header, issue_header, expiry_header],
        [["Asha", "Rao", "03.04.2000", "05.06.2020", "04.06.2030"]],
    )

    row = PassportExcelImporter().import_rows(content)[0]

    assert row.confirmed_fields["date_of_birth"] == "2000-04-03"
    assert row.confirmed_fields["date_of_issue"] == "2020-06-05"
    assert row.confirmed_fields["date_of_expiry"] == "2030-06-04"


def test_mixed_numeric_date_separators_from_real_roster_are_canonicalized() -> None:
    row = PassportExcelImporter().import_rows(
        _excel_bytes(
            ["GIVEN_NAME", "SURNAME", "DATE OF EXPIRY"],
            [["Asha", "Rao", "24-07.2035"]],
        )
    )[0]

    assert row.confirmed_fields["date_of_expiry"] == "2035-07-24"
    assert row.staff_metadata["date_of_expiry"] == "24-07.2035"


@pytest.mark.parametrize(
    ("source", "canonical"),
    [("M", "M"), ("MALE", "M"), ("F", "F"), ("Female", "F"), ("X", "X")],
)
@pytest.mark.parametrize("header", ["Gender", "SEX"])
def test_gender_aliases_store_a_canonical_passport_sex_and_raw_metadata(
    header: str,
    source: str,
    canonical: str,
) -> None:
    row = PassportExcelImporter().import_rows(
        _excel_bytes(["Staffname", header], [["A PERSON", source]])
    )[0]

    assert row.confirmed_fields["sex"] == canonical
    assert row.staff_metadata[header.casefold()] == source


def test_unknown_gender_is_not_invented_but_remains_auditable() -> None:
    row = PassportExcelImporter().import_rows(
        _excel_bytes(["Staffname", "Gender"], [["A PERSON", "Needs review"]])
    )[0]

    assert "sex" not in row.confirmed_fields
    assert row.staff_metadata["gender"] == "Needs review"


def test_all_source_columns_and_duplicate_labels_are_retained_as_metadata() -> None:
    row = PassportExcelImporter().import_rows(
        _excel_bytes(
            [
                "Given Name",
                "Surname",
                "Custom Detail",
                "Custom Detail",
                "Source Sheet",
            ],
            [["Asha", "Rao", "First", "Second", "Original source"]],
            title="Roster",
        )
    )[0]

    assert row.staff_metadata["custom_detail"] == "First"
    assert row.staff_metadata["custom_detail_2"] == "Second"
    assert row.staff_metadata["source_sheet_2"] == "Original source"
    assert row.staff_metadata["source_sheet"] == "Roster"


def test_equal_priority_duplicate_canonical_headers_fail_closed() -> None:
    content = _excel_bytes(
        ["Staffname", "DOB", "D.O.B."],
        [["A PERSON", "01.02.1990", "02.01.1990"]],
    )

    with pytest.raises(ValueError, match="Ambiguous headers.*date of birth"):
        PassportExcelImporter().import_rows(content)


def test_higher_priority_full_date_header_wins_without_losing_source_metadata() -> None:
    row = PassportExcelImporter().import_rows(
        _excel_bytes(
            ["Staffname", "DOB", "Date Of Birth"],
            [["A PERSON", "02.01.1990", "03.04.1991"]],
        )
    )[0]

    assert row.confirmed_fields["date_of_birth"] == "1991-04-03"
    assert row.staff_metadata["dob"] == "02.01.1990"
    assert row.staff_metadata["date_of_birth"] == "03.04.1991"


def test_repeated_structural_header_rows_are_not_imported_as_passengers() -> None:
    row_data = ["Asha", "Rao", "P1234567"]
    rows = PassportExcelImporter().import_rows(
        _excel_bytes(
            ["Given Name", "Surname", "Passport No"],
            [row_data, ["Given Name", "Surname", "Passport No"], row_data],
        )
    )

    assert [row.client_name for row in rows] == ["Asha Rao", "Asha Rao"]


def test_null_markers_are_not_imported_as_passport_values() -> None:
    row = PassportExcelImporter().import_rows(
        _excel_bytes(
            ["Staffname", "StaffCode", "PASSPORT_NO", "SURNAME"],
            [["A PERSON", 101, "NULL", "N/A"]],
        )
    )[0]

    assert row.confirmed_fields == {"staff_code": "101"}
    assert "passport_no" not in row.staff_metadata


def test_place_of_issue_and_legacy_issuing_country_keep_distinct_meanings() -> None:
    for header, expected_field in (
        ("Place of Issue", "place_of_issue"),
        ("Issuing Country", "issuing_country"),
    ):
        row = PassportExcelImporter().import_rows(
            _excel_bytes(["Staffname", header], [["A PERSON", "Chennai"]])
        )[0]

        assert row.confirmed_fields[expected_field] == "Chennai"
        unexpected_field = (
            "issuing_country" if expected_field == "place_of_issue" else "place_of_issue"
        )
        assert unexpected_field not in row.confirmed_fields


def test_workbook_limits_reject_excess_rows_columns_sheets_and_expansion(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    importer = PassportExcelImporter()

    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_ROWS", 2)
    with pytest.raises(ValueError, match="too many rows"):
        importer.import_rows(_excel_bytes(["Staffname"], [["ONE"], ["TWO"]]))

    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_ROWS", 10_000)
    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_COLUMNS", 2)
    with pytest.raises(ValueError, match="too many columns"):
        importer.import_rows(
            _excel_bytes(["Staffname", "Detail", "Another Detail"], [["ONE", 1, 2]])
        )

    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_COLUMNS", 256)
    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_WORKSHEETS", 1)
    workbook = Workbook()
    workbook.active.append(["Staffname"])
    workbook.active.append(["ONE"])
    workbook.create_sheet("Second")
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    with pytest.raises(ValueError, match="too many worksheets"):
        importer.import_rows(content.getvalue())

    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_WORKSHEETS", 50)
    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_UNCOMPRESSED_BYTES", 1)
    with pytest.raises(ValueError, match="expands beyond"):
        importer.import_rows(_excel_bytes(["Staffname"], [["ONE"]]))

    monkeypatch.setattr(
        importer_module,
        "MAX_PASSPORT_EXCEL_UNCOMPRESSED_BYTES",
        50 * 1024 * 1024,
    )
    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_ARCHIVE_MEMBERS", 1)
    with pytest.raises(ValueError, match="too many archive entries"):
        importer.import_rows(_excel_bytes(["Staffname"], [["ONE"]]))

    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_ARCHIVE_MEMBERS", 2_000)
    monkeypatch.setattr(importer_module, "PASSPORT_EXCEL_COMPRESSION_RATIO_MIN_BYTES", 0)
    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_COMPRESSION_RATIO", 0)
    with pytest.raises(ValueError, match="unsafe compression ratio"):
        importer.import_rows(_excel_bytes(["Staffname"], [["ONE"]]))


def test_oversized_source_and_canonical_cells_are_rejected_not_truncated(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    importer = PassportExcelImporter()

    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_CELL_CHARACTERS", 20)
    with pytest.raises(ValueError, match="Excel cells can contain at most 20"):
        importer.import_rows(
            _excel_bytes(
                ["Staffname", "Custom"],
                [["A PERSON", "x" * 21]],
            )
        )

    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_CELL_CHARACTERS", 2_048)
    with pytest.raises(ValueError, match="Given Names can contain at most 160"):
        importer.import_rows(
            _excel_bytes(
                ["GIVEN_NAME", "SURNAME"],
                [["A" * 161, "Rao"]],
            )
        )

    with pytest.raises(ValueError, match="Passenger name can contain at most 255"):
        importer.import_rows(
            _excel_bytes(
                ["Staffname"],
                [["A" * 256]],
            )
        )


@pytest.mark.parametrize("source_value", [" " * 21, "A" + (" " * 20)])
def test_source_cell_limit_is_checked_before_whitespace_normalization(
    monkeypatch: pytest.MonkeyPatch,
    source_value: str,
) -> None:
    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_CELL_CHARACTERS", 20)

    with pytest.raises(ValueError, match="Excel cells can contain at most 20"):
        PassportExcelImporter().import_rows(
            _excel_bytes(
                ["Staffname", "Custom"],
                [["A PERSON", source_value]],
            )
        )


def test_aggregate_populated_cell_and_metadata_character_budgets_are_enforced(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    importer = PassportExcelImporter()

    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_POPULATED_CELLS", 2)
    with pytest.raises(ValueError, match="too many populated cells"):
        importer.import_rows(
            _excel_bytes(
                ["Staffname", "Detail", "Other"],
                [["A PERSON", "one", "two"]],
            )
        )

    monkeypatch.setattr(
        importer_module,
        "MAX_PASSPORT_EXCEL_POPULATED_CELLS",
        100_000,
    )
    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_METADATA_CHARACTERS", 10)
    with pytest.raises(ValueError, match="too much passenger detail"):
        importer.import_rows(
            _excel_bytes(
                ["Staffname", "Detail"],
                [["A PERSON", "a detailed value"]],
            )
        )


def test_repeated_header_rows_are_charged_to_workbook_budgets(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(importer_module, "MAX_PASSPORT_EXCEL_POPULATED_CELLS", 3)

    with pytest.raises(ValueError, match="too many populated cells"):
        PassportExcelImporter().import_rows(
            _excel_bytes(
                ["GIVEN NAME", "SURNAME"],
                [["GIVEN NAME", "SURNAME"]],
            )
        )


def test_importer_handles_intended_1500_passenger_scale() -> None:
    rows = [
        [f"Given {index}", f"Surname {index}", f"P{index:07d}", "M" if index % 2 else "F"]
        for index in range(1, 1501)
    ]

    imported = PassportExcelImporter().import_rows(
        _excel_bytes(
            ["GIVEN_NAME", "SURNAME", "Passport No", "Gender"],
            rows,
        )
    )

    assert len(imported) == 1500
    assert imported[-1].client_name == "Given 1500 Surname 1500"
    assert imported[-1].confirmed_fields["passport_number"] == "P0001500"


def test_importer_streams_workbooks_without_optional_dimension_metadata() -> None:
    imported = PassportExcelImporter().import_rows(
        _streaming_excel_bytes(
            ["GIVEN NAME", "SURNAME", "PASSPORT_NO", "GENDER"],
            [["Asha", "Rao", "Z1234567", "Female"]],
        )
    )

    assert len(imported) == 1
    assert imported[0].client_name == "Asha Rao"
    assert imported[0].confirmed_fields["passport_number"] == "Z1234567"
    assert imported[0].confirmed_fields["sex"] == "F"


def test_streaming_workbook_header_cannot_bypass_column_limit() -> None:
    headers = ["GIVEN NAME", *(f"Custom {index}" for index in range(256))]

    with pytest.raises(ValueError, match="too many columns"):
        PassportExcelImporter().import_rows(
            _streaming_excel_bytes(headers, [["Asha"]])
        )


def test_ignored_streaming_sheet_cannot_bypass_global_row_limit() -> None:
    workbook = Workbook(write_only=True)
    passengers = workbook.create_sheet("Passengers")
    passengers.append(["GIVEN NAME", "SURNAME"])
    passengers.append(["Asha", "Rao"])
    notes = workbook.create_sheet("Notes")
    for index in range(10_001):
        notes.append([f"Note {index}"])
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    with pytest.raises(ValueError, match="too many rows"):
        PassportExcelImporter().import_rows(content.getvalue())
