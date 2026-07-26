from __future__ import annotations

import io
import uuid
from datetime import date
from types import SimpleNamespace

from openpyxl import load_workbook

from app.infrastructure.export.rooming_excel_exporter import RoomingExcelExporter


def test_rooming_export_has_exact_guest_order_values_vip_style_and_formula_safety() -> None:
    passenger_id = uuid.uuid4()
    group = SimpleNamespace(
        name="=Unsafe Group",
        staff_code_enabled=True,
        agent_employee_code_enabled=True,
        travel_date=date(2026, 8, 1),
    )
    hotel = SimpleNamespace(
        hotel_name="+Unsafe Hotel",
        city="@Unsafe City",
        check_in_date=date(2026, 8, 1),
        check_out_date=date(2026, 8, 3),
    )
    room = SimpleNamespace(room_number="-1", room_type="single")
    assignment = SimpleNamespace(passenger_id=passenger_id)
    passenger = SimpleNamespace(
        id=passenger_id,
        confirmed_fields={
            "staff_code": "42",
            "agent_employee_type": "employee",
            "agent_employee_code": "7",
            "given_names": "-Formula Given",
            "surname": "Tester",
            "sex": "F",
            "passport_number": "=P123",
            "date_of_birth": "2000-01-02",
            "date_of_issue": "2020-02-03",
            "date_of_expiry": "2030-02-03",
            "place_of_issue": "@Delhi",
        },
        extracted_fields={},
        staff_metadata={},
    )
    content = RoomingExcelExporter().export_hotel(
        group=group,
        hotel=hotel,
        rooms=[(room, [assignment])],
        passenger_by_id={passenger_id: passenger},
        vip_passenger_ids={passenger_id},
        priority_fields=[
            {"key": "whatsapp:department", "label": "Department", "source": "whatsapp"},
            {"key": "custom:zone", "label": "Zone", "source": "custom_question"},
        ],
        priority_values={
            passenger_id: {
                "whatsapp:department": "=Finance",
                "custom:zone": "North",
            }
        },
    )
    worksheet = load_workbook(io.BytesIO(content)).active
    headers = [cell.value for cell in worksheet[6]]

    assert headers == [
        "Room Number",
        "Room Type",
        "VIP",
        "Staff Code",
        "Agent/Employee Code",
        "Department",
        "Zone",
        "Age Group",
        "GIVEN NAME",
        "SURNAME",
        "GENDER",
        "PASSPORT NUM",
        "DOB",
        "DOI",
        "DOE",
        "PLACE OF ISSUE",
    ]
    values = dict(zip(headers, (cell.value for cell in worksheet[7]), strict=True))
    assert values["Staff Code"] == "STF_42"
    assert values["Agent/Employee Code"] == "EMP_7"
    assert values["Department"] == "'=Finance"
    assert values["Age Group"] == "Adult"
    assert values["GIVEN NAME"] == "'-FORMULA GIVEN"
    assert values["GENDER"] == "Female"
    assert values["PASSPORT NUM"] == "'=P123"
    assert values["PLACE OF ISSUE"] == "'@Delhi"
    assert worksheet["A1"].data_type != "f"
    assert worksheet["A2"].data_type != "f"
    assert worksheet["A3"].data_type != "f"
    assert all(cell.fill.fgColor.rgb == "00FDE68A" for cell in worksheet[7])


def test_checkin_export_sanitizes_all_external_string_cells() -> None:
    passenger = SimpleNamespace(
        room_number="=101",
        room_type="twin",
        passenger_name="+Guest",
        checked_in=True,
        key_issued=False,
        welcome_letter_issued=False,
        remarks="@SUM(A1:A2)",
        is_vip=True,
    )
    content = RoomingExcelExporter().export_checkins(
        group_name="-Group",
        hotel_name="=Hotel",
        passengers=[passenger],
    )
    worksheet = load_workbook(io.BytesIO(content)).active

    assert [cell.value for cell in worksheet[2]][:5] == [
        "'-Group",
        "'=Hotel",
        "'=101",
        "Twin",
        "'+Guest",
    ]
    assert worksheet["I2"].value == "'@SUM(A1:A2)"


def test_rooming_export_omits_code_columns_when_group_did_not_ask() -> None:
    passenger_id = uuid.uuid4()
    group = SimpleNamespace(
        name="No Codes",
        staff_code_enabled=False,
        agent_employee_code_enabled=False,
        travel_date=date(2026, 8, 1),
    )
    hotel = SimpleNamespace(
        hotel_name="Hotel",
        city=None,
        check_in_date=None,
        check_out_date=None,
    )
    room = SimpleNamespace(room_number="1", room_type="twin")
    assignment = SimpleNamespace(passenger_id=passenger_id)
    passenger = SimpleNamespace(
        id=passenger_id,
        confirmed_fields={"sex": "M"},
        extracted_fields={},
        staff_metadata={},
    )
    content = RoomingExcelExporter().export_hotel(
        group=group,
        hotel=hotel,
        rooms=[(room, [assignment])],
        passenger_by_id={passenger_id: passenger},
        vip_passenger_ids=set(),
        priority_fields=[],
        priority_values={passenger_id: {}},
    )
    headers = [cell.value for cell in load_workbook(io.BytesIO(content)).active[6]]

    assert "Staff Code" not in headers
    assert "Agent/Employee Code" not in headers


def test_rooming_export_separates_primary_sections_and_bands_complete_rooms() -> None:
    passenger_ids = [uuid.UUID(int=index) for index in range(1, 8)]
    group = SimpleNamespace(
        name="Priority Sections",
        staff_code_enabled=False,
        agent_employee_code_enabled=False,
        travel_date=date(2026, 8, 1),
    )
    hotel = SimpleNamespace(
        hotel_name="Hotel",
        city=None,
        check_in_date=None,
        check_out_date=None,
    )
    passengers = {
        passenger_id: SimpleNamespace(
            id=passenger_id,
            confirmed_fields={
                "given_names": f"Guest {index}",
                "sex": "F" if index < 6 else "M",
            },
            extracted_fields={},
            staff_metadata={},
        )
        for index, passenger_id in enumerate(passenger_ids, start=1)
    }
    rooms = [
        (
            SimpleNamespace(room_number="1", room_type="single"),
            [SimpleNamespace(passenger_id=passenger_ids[0])],
        ),
        (
            SimpleNamespace(room_number="2", room_type="twin"),
            [
                SimpleNamespace(passenger_id=passenger_ids[1]),
                SimpleNamespace(passenger_id=passenger_ids[2]),
            ],
        ),
        (
            SimpleNamespace(room_number="3", room_type="twin"),
            [
                SimpleNamespace(passenger_id=passenger_ids[3]),
                SimpleNamespace(passenger_id=passenger_ids[4]),
            ],
        ),
        (
            SimpleNamespace(room_number="4", room_type="twin"),
            [
                SimpleNamespace(passenger_id=passenger_ids[5]),
                SimpleNamespace(passenger_id=passenger_ids[6]),
            ],
        ),
    ]
    priority_values = {
        passenger_id: {
            "whatsapp:zone": "Gujarat" if index <= 5 else "Odisha",
            "whatsapp:branch": "A",
        }
        for index, passenger_id in enumerate(passenger_ids, start=1)
    }

    content = RoomingExcelExporter().export_hotel(
        group=group,
        hotel=hotel,
        rooms=rooms,
        passenger_by_id=passengers,
        vip_passenger_ids={passenger_ids[0]},
        priority_fields=[
            {"key": "whatsapp:zone", "label": "Zone", "source": "whatsapp"},
            {
                "key": "whatsapp:branch",
                "label": "Branch",
                "source": "whatsapp",
            },
        ],
        priority_values=priority_values,
    )
    worksheet = load_workbook(io.BytesIO(content)).active

    assert worksheet["A7"].value == "1"
    assert worksheet["A8"].value == "2"
    assert worksheet["A9"].value == "2"
    assert worksheet["A10"].value == "3"
    assert worksheet["A11"].value == "3"
    assert all(worksheet.cell(row=12, column=column).value is None for column in range(1, 15))
    assert all(worksheet.cell(row=13, column=column).value is None for column in range(1, 15))
    assert all(
        worksheet.cell(row=12, column=column).fill.fill_type is None
        for column in range(1, 15)
    )
    assert all(
        worksheet.cell(row=13, column=column).fill.fill_type is None
        for column in range(1, 15)
    )
    assert worksheet["A14"].value == "4"

    assert worksheet["A7"].fill.fgColor.rgb == "00FDE68A"
    first_twin_fill = worksheet["A8"].fill.fgColor.rgb
    second_twin_fill = worksheet["A10"].fill.fgColor.rgb
    assert first_twin_fill == "00DDEBF7"
    assert worksheet["A9"].fill.fgColor.rgb == first_twin_fill
    assert second_twin_fill == "00BDD7EE"
    assert worksheet["A11"].fill.fgColor.rgb == second_twin_fill
    assert worksheet["A14"].fill.fgColor.rgb == "00DDEBF7"
    assert worksheet.tables["HotelRoomingList"].tableStyleInfo.showRowStripes is False
