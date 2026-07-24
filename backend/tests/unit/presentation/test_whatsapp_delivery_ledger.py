from __future__ import annotations

import json
import sys
import uuid
from datetime import UTC, datetime
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock

import pytest
from fastapi import HTTPException, Request, UploadFile
from openpyxl import Workbook
from sqlalchemy import UniqueConstraint, select

from app.application.use_cases.whatsapp.message_templates import render_message
from app.domain.entities.entities import UserRole
from app.infrastructure.database.models import (
    WhatsAppBroadcastRecipientModel,
    WhatsAppMessageLogModel,
    WhatsAppRecipientMessageStateModel,
)
from app.infrastructure.whatsapp.worker_runtime import (
    _load_sendable_recipient,
    _resolve_log_template_snapshot,
    _set_message_state,
    mark_whatsapp_batch_failed,
    run_whatsapp_broadcast,
)
from app.presentation.api.v1.routes.whatsapp import (
    WhatsAppContactPreviewResponse,
    WhatsAppRecipientInput,
    WhatsAppResendRequest,
    _activate_recipient_models,
    _apply_provider_status_to_delivery_state,
    _decode_legacy_template_snapshot,
    _excel_contact_preview_response,
    _extract_status_error,
    _normalized_recipient_inputs,
    _parse_excel_contact_preview,
    _parse_excel_contacts,
    _provider_status_state_predicates,
    _recipient_delivery_counts,
    _recipient_response,
    _template_snapshot_from_log,
    delete_broadcast_group,
    get_broadcast_batch_status,
    preview_excel_contacts,
    receive_whatsapp_webhook,
    resend_recipient_message,
)
from app.presentation.api.v1.routes.whatsapp import (
    router as whatsapp_router,
)


def test_delivery_ledger_schema_is_generic_and_unique_per_recipient_type() -> None:
    constraints = WhatsAppRecipientMessageStateModel.__table__.constraints
    unique_columns = {
        tuple(column.name for column in constraint.columns)
        for constraint in constraints
        if isinstance(constraint, UniqueConstraint)
    }

    assert ("recipient_id", "message_type") in unique_columns
    assert WhatsAppRecipientMessageStateModel.__table__.c.message_type.type.length == 64
    assert WhatsAppMessageLogModel.__table__.c.message_type.type.length == 64
    assert "removed_at" in WhatsAppBroadcastRecipientModel.__table__.c
    assert "imported_fields" in WhatsAppBroadcastRecipientModel.__table__.c


def test_webhook_delivery_error_redacts_raw_meta_message() -> None:
    error = _extract_status_error(
        {
            "errors": [
                {
                    "code": 131026,
                    "message": "Raw provider recipient detail",
                    "details": "Sensitive upstream diagnostic",
                }
            ]
        }
    )

    assert error == (
        "WHATSAPP_PROVIDER_DELIVERY_FAILED: "
        "Meta reported that this message was not delivered (131026)"
    )
    assert "Raw provider" not in error
    assert "Sensitive upstream" not in error


def test_recipient_checklist_only_marks_provider_accepted_states_as_sent() -> None:
    recipient_id = uuid.uuid4()
    now = datetime.now(tz=UTC)
    recipient = SimpleNamespace(
        id=recipient_id,
        name="Aarav",
        phone_number="9876543210",
        normalized_phone_number="+919876543210",
    )
    states = [
        SimpleNamespace(
            recipient_id=recipient_id,
            message_type="welcome",
            status="delivered",
            submitted_at=now,
            status_updated_at=now,
        ),
        SimpleNamespace(
            recipient_id=recipient_id,
            message_type="passport_link",
            status="failed",
            submitted_at=None,
            status_updated_at=now,
        ),
    ]

    response = _recipient_response(recipient, states)

    assert response.sent_message_types == ["welcome"]
    assert [item.message_type for item in response.message_statuses] == [
        "passport_link",
        "welcome",
    ]
    assert {item.message_type: item.already_sent for item in response.message_statuses} == {
        "passport_link": False,
        "welcome": True,
    }


@pytest.mark.asyncio
async def test_eligibility_counts_failed_as_retryable_and_active_claim_as_skipped() -> None:
    recipient_ids = [uuid.uuid4() for _ in range(5)]
    recipients = [SimpleNamespace(id=recipient_id) for recipient_id in recipient_ids]
    result = MagicMock()
    result.all.return_value = [
        (recipient_ids[0], "submitted"),
        (recipient_ids[1], "processing"),
        (recipient_ids[2], "failed"),
        (recipient_ids[3], "delivery_unknown"),
    ]
    session = AsyncMock()
    session.execute.return_value = result

    eligible, already_sent, in_progress, uncertain = await _recipient_delivery_counts(
        session,
        recipients=recipients,
        message_type="welcome",
    )

    assert eligible == 2  # failed + never attempted
    assert already_sent == 1
    assert in_progress == 1
    assert uncertain == 1


def test_readding_removed_phone_reactivates_same_row_and_preserves_identity() -> None:
    recipient_id = uuid.uuid4()
    group = SimpleNamespace(id=uuid.uuid4(), agency_id=uuid.uuid4())
    existing = SimpleNamespace(
        id=recipient_id,
        name="Old Name",
        phone_number="9876543210",
        normalized_phone_number="+919876543210",
        imported_fields={"old": "value"},
        removed_at=datetime.now(tz=UTC),
        suppressed_by_roster_resolution_id=None,
    )
    normalized = _normalized_recipient_inputs(
        [
            WhatsAppRecipientInput(
                name="Updated Name",
                phone_number="+91 98765 43210",
                imported_fields={"email": "updated@example.com"},
            )
        ]
    )
    session = MagicMock()

    _activate_recipient_models(
        session=session,
        group=group,
        existing_by_phone={existing.normalized_phone_number: existing},
        normalized_contacts=normalized,
        now=datetime.now(tz=UTC),
    )

    assert existing.id == recipient_id
    assert existing.removed_at is None
    assert existing.name == "Updated Name"
    assert existing.imported_fields == {"email": "updated@example.com"}
    session.add.assert_not_called()


def test_manual_reactivation_does_not_erase_imported_fields() -> None:
    group = SimpleNamespace(id=uuid.uuid4(), agency_id=uuid.uuid4())
    existing = SimpleNamespace(
        id=uuid.uuid4(),
        name="Imported Name",
        phone_number="9876543210",
        normalized_phone_number="+919876543210",
        imported_fields={
            "email": "imported@example.com",
            "staff_code": "GC42",
        },
        removed_at=datetime.now(tz=UTC),
        suppressed_by_roster_resolution_id=None,
    )
    normalized = _normalized_recipient_inputs(
        [
            WhatsAppRecipientInput(
                name="Edited Name",
                phone_number="+91 98765 43210",
            )
        ]
    )

    _activate_recipient_models(
        session=MagicMock(),
        group=group,
        existing_by_phone={existing.normalized_phone_number: existing},
        normalized_contacts=normalized,
        now=datetime.now(tz=UTC),
    )

    assert existing.name == "Edited Name"
    assert existing.imported_fields == {
        "email": "imported@example.com",
        "staff_code": "GC42",
    }


def test_readding_replaced_phone_requires_roster_restore() -> None:
    group = SimpleNamespace(id=uuid.uuid4(), agency_id=uuid.uuid4())
    resolution_id = uuid.uuid4()
    removed_at = datetime.now(tz=UTC)
    existing = SimpleNamespace(
        id=uuid.uuid4(),
        name="Replaced Traveller",
        phone_number="9876543210",
        normalized_phone_number="+919876543210",
        imported_fields={"staff_code": "GC42"},
        removed_at=removed_at,
        suppressed_by_roster_resolution_id=resolution_id,
    )
    normalized = _normalized_recipient_inputs(
        [
            WhatsAppRecipientInput(
                name="Attempted Reactivation",
                phone_number="+91 98765 43210",
            )
        ]
    )

    with pytest.raises(HTTPException) as exc_info:
        _activate_recipient_models(
            session=MagicMock(),
            group=group,
            existing_by_phone={existing.normalized_phone_number: existing},
            normalized_contacts=normalized,
            now=datetime.now(tz=UTC),
        )

    assert exc_info.value.status_code == 409
    assert "marked as replaced" in str(exc_info.value.detail)
    assert existing.removed_at == removed_at
    assert existing.name == "Replaced Traveller"


@pytest.mark.asyncio
async def test_worker_acceptance_makes_inflight_provider_success_authoritative() -> None:
    batch_id = uuid.uuid4()
    log = SimpleNamespace(
        recipient_id=uuid.uuid4(),
        message_type="future_template_type",
    )
    session = AsyncMock()

    await _set_message_state(
        session,
        log=log,
        expected_batch_id=batch_id,
        state_status="submitted",
        submitted=True,
    )

    statement = session.execute.await_args.args[0]
    compiled = statement.compile()
    assert "recipient_id" in str(statement.whereclause)
    assert "message_type" in str(statement.whereclause)
    assert "status" in str(statement.whereclause)
    assert batch_id in compiled.params.values()
    assert "submitted" in compiled.params.values()
    assert any(isinstance(value, datetime) for value in compiled.params.values())


@pytest.mark.asyncio
async def test_worker_failure_releases_ledger_claim_for_safe_retry() -> None:
    batch_id = uuid.uuid4()
    log = SimpleNamespace(
        recipient_id=uuid.uuid4(),
        message_type="welcome",
    )
    session = AsyncMock()

    await _set_message_state(
        session,
        log=log,
        expected_batch_id=batch_id,
        state_status="failed",
        release_claim=True,
    )

    statement = session.execute.await_args.args[0]
    compiled = statement.compile()
    assert "failed" in compiled.params.values()
    assert None in compiled.params.values()


@pytest.mark.asyncio
async def test_worker_success_exits_retry_loop_and_remains_submitted(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    batch_id = uuid.uuid4()
    group_id = uuid.uuid4()
    recipient_id = uuid.uuid4()
    log = SimpleNamespace(
        id=uuid.uuid4(),
        batch_id=batch_id,
        broadcast_group_id=group_id,
        recipient_id=recipient_id,
        message_type="welcome",
        template_name="welcome_template",
        status="queued",
        status_updated_at=datetime.now(tz=UTC),
        provider_message_id=None,
        error_message=None,
        created_at=datetime.now(tz=UTC),
    )
    group = SimpleNamespace(
        id=group_id,
        name="Vietnam 2026",
        organizing_company_name="Bluechip",
    )
    recipient = SimpleNamespace(
        id=recipient_id,
        name="Aarav",
        normalized_phone_number="+919876543210",
    )

    logs_result = MagicMock()
    logs_result.scalars.return_value.all.return_value = [log]
    group_result = MagicMock()
    group_result.scalar_one_or_none.return_value = group
    support_result = MagicMock()
    support_result.scalars.return_value.all.return_value = []
    log_claim_result = MagicMock()
    log_claim_result.scalar_one_or_none.return_value = log.id
    state_claim_result = MagicMock()
    state_claim_result.scalar_one_or_none.return_value = uuid.uuid4()

    session = AsyncMock()
    session.execute.side_effect = [
        logs_result,
        group_result,
        support_result,
        log_claim_result,
        state_claim_result,
    ]

    class _AsyncContext:
        def __init__(self, value: object) -> None:
            self.value = value

        async def __aenter__(self) -> object:
            return self.value

        async def __aexit__(self, *args: object) -> None:
            return None

    client = SimpleNamespace()
    send_template = AsyncMock(return_value="wamid.success")
    load_recipient = AsyncMock(return_value=(recipient, None))
    set_state = AsyncMock()
    monkeypatch.setattr(
        "app.infrastructure.whatsapp.worker_runtime.AsyncSessionFactory",
        lambda: _AsyncContext(session),
    )
    monkeypatch.setattr(
        "app.infrastructure.whatsapp.worker_runtime.httpx.AsyncClient",
        lambda **_kwargs: _AsyncContext(client),
    )
    monkeypatch.setattr(
        "app.infrastructure.whatsapp.worker_runtime.send_whatsapp_template",
        send_template,
    )
    monkeypatch.setattr(
        "app.infrastructure.whatsapp.worker_runtime._load_sendable_recipient",
        load_recipient,
    )
    monkeypatch.setattr(
        "app.infrastructure.whatsapp.worker_runtime._set_message_state",
        set_state,
    )

    await run_whatsapp_broadcast(
        batch_id=str(batch_id),
        message_type="welcome",
        message_content="Welcome aboard",
        passport_link=None,
        header_image_id="media-123",
    )

    assert send_template.await_count == 1
    send_kwargs = send_template.await_args.kwargs
    assert send_kwargs["message_type"] == "welcome"
    assert send_kwargs["header_parameters"] == ["media-123"]
    assert send_kwargs["parameters"] == ["Welcome aboard"]
    assert "Aarav" not in " ".join(send_kwargs["parameters"])
    assert "Bluechip" not in " ".join(send_kwargs["parameters"])
    assert load_recipient.await_count == 2
    assert log.status == "submitted"
    assert log.provider_message_id == "wamid.success"
    assert set_state.await_args.kwargs["state_status"] == "submitted"
    assert set_state.await_args.kwargs["submitted"] is True


@pytest.mark.asyncio
async def test_worker_guard_rejects_removed_recipient_before_provider_call() -> None:
    recipient = SimpleNamespace(
        id=uuid.uuid4(),
        removed_at=datetime.now(tz=UTC),
    )
    recipient_result = MagicMock()
    recipient_result.scalar_one_or_none.return_value = recipient
    session = AsyncMock()
    session.execute.return_value = recipient_result

    sendable, reason = await _load_sendable_recipient(
        session,
        log=SimpleNamespace(
            recipient_id=recipient.id,
            broadcast_group_id=uuid.uuid4(),
            message_type="welcome",
        ),
        expected_batch_id=uuid.uuid4(),
    )

    assert sendable is None
    assert reason == "WhatsApp recipient was removed"
    assert session.execute.await_count == 2


@pytest.mark.asyncio
async def test_worker_guard_requires_current_processing_batch() -> None:
    batch_id = uuid.uuid4()
    recipient = SimpleNamespace(
        id=uuid.uuid4(),
        agency_id=uuid.uuid4(),
        broadcast_group_id=uuid.uuid4(),
        normalized_phone_number="+919876543210",
        removed_at=None,
    )
    recipient_result = MagicMock()
    recipient_result.scalar_one_or_none.return_value = recipient
    state_result = MagicMock()
    state_result.scalar_one_or_none.return_value = SimpleNamespace(
        batch_id=batch_id,
        status="processing",
    )
    group_result = MagicMock()
    group_result.scalar_one_or_none.return_value = SimpleNamespace(id=uuid.uuid4())
    session = AsyncMock()
    active_replacement_result = MagicMock()
    active_replacement_result.scalar_one_or_none.return_value = None
    session.execute.side_effect = [
        group_result,
        recipient_result,
        active_replacement_result,
        state_result,
    ]

    sendable, reason = await _load_sendable_recipient(
        session,
        log=SimpleNamespace(
            recipient_id=recipient.id,
            broadcast_group_id=group_result.scalar_one_or_none.return_value.id,
            message_type="welcome",
        ),
        expected_batch_id=batch_id,
    )

    assert sendable is recipient
    assert reason is None


@pytest.mark.asyncio
async def test_worker_guard_rejects_later_same_phone_replacement_row() -> None:
    batch_id = uuid.uuid4()
    group_id = uuid.uuid4()
    recipient = SimpleNamespace(
        id=uuid.uuid4(),
        agency_id=uuid.uuid4(),
        broadcast_group_id=group_id,
        normalized_phone_number="+919876543210",
        removed_at=None,
    )
    group_result = MagicMock()
    group_result.scalar_one_or_none.return_value = SimpleNamespace(id=group_id)
    recipient_result = MagicMock()
    recipient_result.scalar_one_or_none.return_value = recipient
    replacement_result = MagicMock()
    replacement_result.scalar_one_or_none.return_value = uuid.uuid4()
    session = AsyncMock()
    session.execute.side_effect = [
        group_result,
        recipient_result,
        replacement_result,
    ]

    sendable, reason = await _load_sendable_recipient(
        session,
        log=SimpleNamespace(
            recipient_id=recipient.id,
            broadcast_group_id=group_id,
            message_type="welcome",
        ),
        expected_batch_id=batch_id,
    )

    assert sendable is None
    assert reason == "WhatsApp recipient was replaced in a linked passport group"
    assert session.execute.await_count == 3


def test_late_failed_receipt_does_not_regress_accepted_checklist() -> None:
    original_time = datetime(2026, 7, 18, tzinfo=UTC)
    state = SimpleNamespace(
        status="delivered",
        batch_id=uuid.uuid4(),
        submitted_at=original_time,
        status_updated_at=original_time,
        provider_status_at=original_time,
        updated_at=original_time,
    )

    _apply_provider_status_to_delivery_state(
        state,
        provider_status="failed",
        provider_status_at=datetime(2026, 7, 19, tzinfo=UTC),
        now=datetime(2026, 7, 19, tzinfo=UTC),
    )

    assert state.status == "delivered"
    assert state.batch_id is not None
    assert state.submitted_at == original_time


def test_out_of_order_accepted_receipt_never_moves_status_backwards() -> None:
    original_time = datetime(2026, 7, 18, tzinfo=UTC)
    state = SimpleNamespace(
        status="read",
        batch_id=uuid.uuid4(),
        submitted_at=original_time,
        status_updated_at=original_time,
        provider_status_at=original_time,
        updated_at=original_time,
    )

    _apply_provider_status_to_delivery_state(
        state,
        provider_status="sent",
        provider_status_at=datetime(2026, 7, 19, tzinfo=UTC),
        now=datetime(2026, 7, 19, tzinfo=UTC),
    )

    assert state.status == "read"


def test_failed_receipt_before_acceptance_releases_retry_claim() -> None:
    batch_id = uuid.uuid4()
    state = SimpleNamespace(
        status="processing",
        batch_id=batch_id,
        submitted_at=None,
        status_updated_at=datetime(2026, 7, 18, tzinfo=UTC),
        provider_status_at=None,
        updated_at=datetime(2026, 7, 18, tzinfo=UTC),
    )

    _apply_provider_status_to_delivery_state(
        state,
        provider_status="failed",
        provider_status_at=datetime(2026, 7, 19, tzinfo=UTC),
        now=datetime(2026, 7, 19, tzinfo=UTC),
    )

    assert state.status == "failed"
    assert state.batch_id == batch_id


def test_current_failed_receipt_releases_submitted_but_not_delivered() -> None:
    submitted_at = datetime(2026, 7, 18, tzinfo=UTC)
    provider_event_at = datetime(2026, 7, 19, tzinfo=UTC)
    batch_id = uuid.uuid4()
    state = SimpleNamespace(
        status="submitted",
        batch_id=batch_id,
        submitted_at=submitted_at,
        status_updated_at=submitted_at,
        provider_status_at=None,
        updated_at=submitted_at,
    )

    _apply_provider_status_to_delivery_state(
        state,
        provider_status="failed",
        provider_status_at=provider_event_at,
        now=provider_event_at,
    )

    assert state.status == "failed"
    assert state.batch_id == batch_id
    assert state.provider_status_at == provider_event_at


def test_older_provider_event_is_ignored() -> None:
    latest_event_at = datetime(2026, 7, 19, tzinfo=UTC)
    state = SimpleNamespace(
        status="delivered",
        batch_id=uuid.uuid4(),
        submitted_at=datetime(2026, 7, 18, tzinfo=UTC),
        status_updated_at=latest_event_at,
        provider_status_at=latest_event_at,
        updated_at=latest_event_at,
    )

    _apply_provider_status_to_delivery_state(
        state,
        provider_status="read",
        provider_status_at=datetime(2026, 7, 18, tzinfo=UTC),
        now=datetime(2026, 7, 20, tzinfo=UTC),
    )

    assert state.status == "delivered"
    assert state.provider_status_at == latest_event_at


def test_late_accepted_receipt_promotes_ledger_across_retry_batches() -> None:
    log = SimpleNamespace(
        recipient_id=uuid.uuid4(),
        message_type="welcome",
        batch_id=uuid.uuid4(),
    )

    statement = select(WhatsAppRecipientMessageStateModel).where(
        *_provider_status_state_predicates(log, provider_status="delivered")
    )

    assert "batch_id" not in str(statement.whereclause)


def test_late_failed_receipt_remains_scoped_to_its_original_batch() -> None:
    log = SimpleNamespace(
        recipient_id=uuid.uuid4(),
        message_type="welcome",
        batch_id=uuid.uuid4(),
    )

    statement = select(WhatsAppRecipientMessageStateModel).where(
        *_provider_status_state_predicates(log, provider_status="failed")
    )

    compiled = statement.compile()
    assert "batch_id" in str(statement.whereclause)
    assert log.batch_id in compiled.params.values()


@pytest.mark.asyncio
async def test_batch_status_reports_ambiguous_outcomes_separately_from_failures() -> None:
    now = datetime.now(tz=UTC)
    stale_time = now.replace(year=now.year - 1)
    rows = [
        (
            SimpleNamespace(
                status="delivery_unknown",
                status_updated_at=now,
                provider_message_id=None,
                error_message="Provider outcome is unknown",
            ),
            SimpleNamespace(id=uuid.uuid4(), normalized_phone_number="+919999999991"),
        ),
        (
            SimpleNamespace(
                status="processing",
                status_updated_at=stale_time,
                provider_message_id=None,
                error_message=None,
            ),
            SimpleNamespace(id=uuid.uuid4(), normalized_phone_number="+919999999992"),
        ),
        (
            SimpleNamespace(
                status="failed",
                status_updated_at=now,
                provider_message_id=None,
                error_message="Recipient number rejected",
            ),
            SimpleNamespace(id=uuid.uuid4(), normalized_phone_number="+919999999993"),
        ),
    ]
    result = MagicMock()
    result.all.return_value = rows
    session = AsyncMock()
    session.execute.return_value = result

    response = await get_broadcast_batch_status(
        batch_id=uuid.uuid4(),
        current_user=SimpleNamespace(role=UserRole.SUPER_ADMIN, agency_id=None),
        session=session,
    )

    assert response.queued == 0
    assert response.sent == 0
    assert response.failed == 1
    assert response.delivery_unknown == 2
    assert [item.status for item in response.results] == [
        "delivery_unknown",
        "stalled",
        "failed",
    ]


def test_manual_recipient_rejects_non_phone_characters_and_excess_digits() -> None:
    for invalid_number in (
        "call-me-9876543210",
        "1234567890123456789012345678901",
    ):
        with pytest.raises(HTTPException) as exc_info:
            _normalized_recipient_inputs(
                [
                    WhatsAppRecipientInput(
                        name="Invalid",
                        phone_number=invalid_number,
                    )
                ]
            )

        assert exc_info.value.status_code == 400


@pytest.mark.asyncio
async def test_excel_contact_upload_enforces_compressed_byte_limit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        "app.presentation.api.v1.routes.whatsapp.MAX_WHATSAPP_CONTACT_FILE_BYTES",
        4,
    )
    upload = UploadFile(
        file=BytesIO(b"12345"),
        filename="contacts.xlsx",
    )

    with pytest.raises(HTTPException) as exc_info:
        await _parse_excel_contacts(upload)

    assert exc_info.value.status_code == 413


@pytest.mark.asyncio
async def test_excel_contact_upload_detects_headers_below_title_row() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([None, "GC Staff List", None, None])
    sheet.append([None, "S.no", "NAME", "Phone"])
    sheet.append([None, 1, "Aarav Sharma", 9873361557])
    sheet.append([None, 2, "Meera Patel", 9355926411])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    payload.seek(0)
    upload = UploadFile(file=payload, filename="contacts.xlsx")

    contacts = await _parse_excel_contacts(upload)
    normalized = _normalized_recipient_inputs(contacts)

    assert [(contact.name, contact.phone_number) for contact in contacts] == [
        ("Aarav Sharma", "9873361557"),
        ("Meera Patel", "9355926411"),
    ]
    assert set(normalized) == {
        "+919873361557",
        "+919355926411",
    }


@pytest.mark.asyncio
async def test_excel_contact_upload_distinguishes_contact_name_from_phone() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Contact Name", "Mobile Number"])
    sheet.append(["Aarav Sharma", 9873361557])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    payload.seek(0)

    contacts = await _parse_excel_contacts(
        UploadFile(file=payload, filename="contacts.xlsx"),
    )

    assert [(contact.name, contact.phone_number) for contact in contacts] == [
        ("Aarav Sharma", "9873361557"),
    ]


@pytest.mark.asyncio
async def test_excel_contact_upload_preserves_bounded_fields_across_sheets() -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "North"
    first.append(
        [
            "Staffname",
            "Mobile",
            "Email",
            "StaffCode",
            "PASSPORT_NO",
            "Agent Company",
        ]
    )
    first.append(
        [
            "Irfan Khan",
            9355926412,
            "IRFAN@EXAMPLE.COM",
            25523,
            "Z4160891",
            "Bluechip",
        ]
    )
    second = workbook.create_sheet("South")
    second.append(["NAME", "Phone", "Custom Field", "Empty Field"])
    second.append(["Meera Patel", 9355926411, "Useful value", "NULL"])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    payload.seek(0)

    contacts = await _parse_excel_contacts(
        UploadFile(file=payload, filename="contacts.xlsx"),
    )

    assert [contact.name for contact in contacts] == [
        "Irfan Khan",
        "Meera Patel",
    ]
    assert contacts[0].imported_fields == {
        "name": "Irfan Khan",
        "phone_number": "+919355926412",
        "email": "irfan@example.com",
        "staff_code": "25523",
        "passport_number": "Z4160891",
        "agent_company": "Bluechip",
        "source_file": "contacts.xlsx",
        "source_order": "1",
        "source_sheet": "North",
        "source_row": "2",
    }
    assert contacts[1].imported_fields == {
        "name": "Meera Patel",
        "phone_number": "+919355926411",
        "custom_field": "Useful value",
        "source_file": "contacts.xlsx",
        "source_order": "2",
        "source_sheet": "South",
        "source_row": "2",
    }


@pytest.mark.asyncio
async def test_excel_contact_name_prefers_staffname_over_zone_name() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Region",
            "Zonecode",
            "Zone Name",
            "BdmCode",
            "Brname",
            "Staffname",
            "StaffCode",
            "SURNAME",
            "GIVEN_NAME",
            "Email",
            "Mobile",
        ]
    )
    sheet.append(
        [
            "T03",
            "Z017",
            "ASSAM",
            0,
            "GUWAHATI",
            "BIPLAB DAS",
            25523,
            "DAS",
            "BIPLOB",
            "biplab@example.com",
            9435734892,
        ]
    )
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    payload.seek(0)

    contacts = await _parse_excel_contacts(
        UploadFile(file=payload, filename="zone-list.xlsx"),
    )

    assert contacts[0].name == "BIPLAB DAS"
    assert contacts[0].name != contacts[0].imported_fields["zone_name"]


@pytest.mark.asyncio
async def test_excel_contact_name_composes_given_names_and_surname() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Zone Name", "GIVEN_NAME", "SURNAME", "Mobile"])
    sheet.append(["ASSAM", "Irfan", "Khan", 9355926412])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    payload.seek(0)

    contacts = await _parse_excel_contacts(
        UploadFile(file=payload, filename="passengers.xlsx"),
    )

    assert contacts[0].name == "Irfan Khan"


@pytest.mark.asyncio
async def test_excel_contact_upload_converts_unexpected_parser_error_to_400(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        "app.presentation.api.v1.routes.whatsapp._validate_excel_archive",
        lambda _payload: None,
    )

    def raise_malformed_workbook(*_args: object, **_kwargs: object) -> None:
        raise KeyError("missing OOXML workbook relationship")

    monkeypatch.setattr(
        "app.presentation.api.v1.routes.whatsapp.load_workbook",
        raise_malformed_workbook,
    )

    with pytest.raises(HTTPException) as exc_info:
        await _parse_excel_contacts(
            UploadFile(file=BytesIO(b"malformed"), filename="contacts.xlsx"),
        )

    assert exc_info.value.status_code == 400
    assert exc_info.value.detail == "The uploaded Excel contact file could not be read"


@pytest.mark.asyncio
async def test_excel_contact_preview_returns_named_normalized_recipients() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([None, "GCT Staff Mobile Number Detail", None, None])
    sheet.append([None, "S.no", "NAME", "Phone"])
    sheet.append([None, 1, "  Aarav   Sharma  ", 9873361557])
    sheet.append([None, 2, "Meera Patel", "+91 93559 26411"])
    sheet.append([None, 3, "Duplicate Aarav", "+91-9873361557"])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    payload.seek(0)

    response = await preview_excel_contacts(
        contacts_file=UploadFile(file=payload, filename="contacts.xlsx"),
        current_user=SimpleNamespace(),
    )

    assert response.model_dump() == {
        "recipient_count": 2,
        "accepted_count": 2,
        "recipients": [
            {
                "name": "Aarav Sharma",
                "phone_number": "+919873361557",
                "imported_fields": {
                    "name": "Aarav Sharma",
                    "phone_number": "+919873361557",
                    "s_no": "1",
                    "name_2": "Duplicate Aarav",
                    "s_no_2": "3",
                    "duplicate_conflicting_fields": "name, s_no",
                    "source_file": "contacts.xlsx",
                    "source_order": "1",
                    "source_sheet": "Sheet",
                    "source_row": "3",
                },
            },
            {
                "name": "Meera Patel",
                "phone_number": "+919355926411",
                "imported_fields": {
                    "name": "Meera Patel",
                    "phone_number": "+919355926411",
                    "s_no": "2",
                    "source_file": "contacts.xlsx",
                    "source_order": "2",
                    "source_sheet": "Sheet",
                    "source_row": "4",
                },
            },
        ],
        "rejected_count": 1,
        "rejected_rows": [
            {
                "sheet_name": "Sheet",
                "row_number": 5,
                "raw_name": "Duplicate Aarav",
                "raw_phone_number": "+91-9873361557",
                "imported_fields": {
                    "s_no": "3",
                    "name": "Duplicate Aarav",
                    "phone_number": "+919873361557",
                    "source_file": "contacts.xlsx",
                    "source_order": "3",
                    "source_sheet": "Sheet",
                    "source_row": "5",
                },
                "reason_code": "duplicate_phone",
                "reason": (
                    "This WhatsApp number is already listed; its extra details "
                    "were merged into the first accepted contact."
                ),
            }
        ],
        "rejected_rows_truncated": False,
        "omitted_rejected_count": 0,
    }


@pytest.mark.asyncio
async def test_excel_contact_preview_returns_valid_rows_and_actionable_rejections() -> None:
    workbook = Workbook()
    north = workbook.active
    north.title = "North"
    north.append(["Name", "Mobile", "Email"])
    north.append(["Aarav Sharma", 9873361557, "aarav@example.com"])
    north.append(["Invalid Phone", 919726092, "invalid@example.com"])
    north.append(["Missing Phone", None, "missing-phone@example.com"])
    north.append([None, 9355926411, "missing-name@example.com"])
    north.append(["Duplicate Aarav", "+91-9873361557", "alternate@example.com"])
    north.append(["Name", "Mobile", "Email"])
    south = workbook.create_sheet("South")
    south.append(["Name", "Mobile"])
    south.append(["Meera Patel", 9355926412])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    payload.seek(0)

    response = await preview_excel_contacts(
        contacts_file=UploadFile(file=payload, filename="contacts.xlsx"),
        current_user=SimpleNamespace(),
    )

    assert response.recipient_count == 2
    assert response.accepted_count == 2
    assert response.rejected_count == 4
    assert [
        (
            row.sheet_name,
            row.row_number,
            row.raw_name,
            row.raw_phone_number,
            row.reason_code,
        )
        for row in response.rejected_rows
    ] == [
        ("North", 3, "Invalid Phone", "919726092", "invalid_phone"),
        ("North", 4, "Missing Phone", None, "missing_phone"),
        ("North", 5, None, "9355926411", "missing_name"),
        (
            "North",
            6,
            "Duplicate Aarav",
            "+91-9873361557",
            "duplicate_phone",
        ),
    ]
    assert "10-digit Indian mobile number" in response.rejected_rows[0].reason
    assert response.recipients[0].imported_fields["email_2"] == "alternate@example.com"
    assert response.recipients[0].imported_fields["duplicate_conflicting_fields"] == ("email, name")
    assert [recipient.name for recipient in response.recipients] == [
        "Aarav Sharma",
        "Meera Patel",
    ]


@pytest.mark.asyncio
async def test_excel_contact_preview_skips_structural_header_variants() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Region", "Name", "Mobile"])
    sheet.append(["North", "Aarav Sharma", 9873361557])
    sheet.append(["B", "Name", "Mobile"])
    sheet.append(["North", "Invalid Phone", 919726092])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    payload.seek(0)

    response = await preview_excel_contacts(
        contacts_file=UploadFile(file=payload, filename="contacts.xlsx"),
        current_user=SimpleNamespace(),
    )

    assert response.accepted_count == 1
    assert response.rejected_count == 1
    assert response.rejected_rows[0].row_number == 4
    assert response.rejected_rows[0].reason_code == "invalid_phone"


@pytest.mark.asyncio
async def test_excel_contact_preview_returns_all_rejected_rows_without_a_server_error() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Mobile"])
    sheet.append(["Invalid Phone", 805527415])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    payload.seek(0)

    response = await preview_excel_contacts(
        contacts_file=UploadFile(file=payload, filename="contacts.xlsx"),
        current_user=SimpleNamespace(),
    )

    assert response.recipient_count == 0
    assert response.accepted_count == 0
    assert response.recipients == []
    assert response.rejected_count == 1
    assert response.rejected_rows[0].reason_code == "invalid_phone"


@pytest.mark.asyncio
async def test_excel_contact_direct_upload_remains_strict_but_allows_merged_duplicates() -> None:
    invalid_workbook = Workbook()
    invalid_sheet = invalid_workbook.active
    invalid_sheet.append(["Name", "Mobile"])
    invalid_sheet.append(["Valid", 9873361557])
    invalid_sheet.append(["Invalid", 919726092])
    invalid_payload = BytesIO()
    invalid_workbook.save(invalid_payload)
    invalid_workbook.close()
    invalid_payload.seek(0)

    with pytest.raises(HTTPException) as exc_info:
        await _parse_excel_contacts(
            UploadFile(file=invalid_payload, filename="contacts.xlsx"),
        )

    assert exc_info.value.status_code == 400
    assert "1 invalid contact row" in str(exc_info.value.detail)

    duplicate_workbook = Workbook()
    duplicate_sheet = duplicate_workbook.active
    duplicate_sheet.append(["Name", "Mobile", "Email"])
    duplicate_sheet.append(["Aarav", 9873361557, "first@example.com"])
    duplicate_sheet.append(["Aarav Sharma", "+91-9873361557", "second@example.com"])
    duplicate_payload = BytesIO()
    duplicate_workbook.save(duplicate_payload)
    duplicate_workbook.close()
    duplicate_payload.seek(0)

    contacts = await _parse_excel_contacts(
        UploadFile(file=duplicate_payload, filename="contacts.xlsx"),
    )

    assert len(contacts) == 1
    assert contacts[0].imported_fields["email_2"] == "second@example.com"


@pytest.mark.asyncio
async def test_excel_contact_preview_bounds_rejected_row_payload(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        "app.presentation.api.v1.routes.whatsapp.MAX_WHATSAPP_REJECTED_ROWS",
        1,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Mobile"])
    sheet.append(["Invalid One", 919726092])
    sheet.append(["Invalid Two", 805527415])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    payload.seek(0)

    result = await _parse_excel_contact_preview(
        UploadFile(file=payload, filename="contacts.xlsx"),
    )
    response = _excel_contact_preview_response(
        result.contacts,
        result.rejected_rows,
        rejected_count=result.rejected_count,
    )

    assert response.rejected_count == 2
    assert len(response.rejected_rows) == 1
    assert response.rejected_rows_truncated is True
    assert response.omitted_rejected_count == 1


def test_excel_contact_preview_rejects_empty_or_unnamed_contacts() -> None:
    with pytest.raises(HTTPException) as empty_error:
        _excel_contact_preview_response([])

    assert empty_error.value.status_code == 400
    assert "No recipients were found" in str(empty_error.value.detail)

    with pytest.raises(HTTPException) as unnamed_error:
        _excel_contact_preview_response(
            [
                WhatsAppRecipientInput(
                    name=None,
                    phone_number="9873361557",
                )
            ]
        )

    assert unnamed_error.value.status_code == 400
    assert "Missing names for 1 contact" in str(unnamed_error.value.detail)


@pytest.mark.asyncio
async def test_excel_contact_preview_rejects_invalid_file_type() -> None:
    upload = UploadFile(
        file=BytesIO(b"name,phone\nAarav,9873361557"),
        filename="contacts.csv",
    )

    with pytest.raises(HTTPException) as exc_info:
        await preview_excel_contacts(
            contacts_file=upload,
            current_user=SimpleNamespace(),
        )

    assert exc_info.value.status_code == 400
    assert exc_info.value.detail == "Upload an .xlsx or .xlsm contact file"


def test_excel_contact_preview_route_is_role_gated_and_has_stable_contract() -> None:
    route = next(item for item in whatsapp_router.routes if item.path == "/contacts/preview")

    assert route.methods == {"POST"}
    assert route.response_model is WhatsAppContactPreviewResponse
    assert set(WhatsAppContactPreviewResponse.model_fields) == {
        "recipient_count",
        "accepted_count",
        "recipients",
        "rejected_count",
        "rejected_rows",
        "rejected_rows_truncated",
        "omitted_rejected_count",
    }
    assert [dependency.call.__name__ for dependency in route.dependant.dependencies] == [
        "_check_role"
    ]


@pytest.mark.asyncio
async def test_excel_contact_upload_stops_after_recipient_row_limit() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "WhatsApp Phone"])
    for index in range(501):
        sheet.append([f"Recipient {index}", f"9190000{index:05d}"])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    payload.seek(0)
    upload = UploadFile(file=payload, filename="contacts.xlsx")

    with pytest.raises(HTTPException) as exc_info:
        await _parse_excel_contacts(upload)

    assert exc_info.value.status_code == 400
    assert "at most 500 recipients" in str(exc_info.value.detail)


def test_explicit_resend_schema_freezes_payload_and_has_single_active_guard() -> None:
    columns = WhatsAppMessageLogModel.__table__.c
    indexes = {index.name: index for index in WhatsAppMessageLogModel.__table__.indexes}

    assert "header_parameter_values" in columns
    assert "template_parameter_values" in columns
    assert "is_explicit_resend" in columns
    assert indexes["uq_whatsapp_active_explicit_resend"].unique is True


def _legacy_welcome_rendered(message_content: str, support_contacts: str) -> str:
    return (
        "Dear Delegates\n\n"
        "Greetings from Global Connect Travels.\n\n"
        f"{message_content}\n\n"
        "This is an automated notification sent individually to you. Replies to this "
        "WhatsApp message are not monitored and will not be treated as support requests."
        "\n\n"
        "For assistance, please contact:\n"
        f"{support_contacts}\n\n"
        "Regards,\n"
        "Team Global Connect Travels"
    )


def test_legacy_welcome_snapshot_is_byte_exact_and_keeps_original_support() -> None:
    rendered = _legacy_welcome_rendered(
        'This message is regarding your upcoming trip to "THAILAND".',
        "Aman: +919876543211",
    )

    header, body = _decode_legacy_template_snapshot(
        message_type="welcome",
        rendered_message=rendered,
    )

    assert header == []
    assert body == [
        'This message is regarding your upcoming trip to "THAILAND".',
        "Aman: +919876543211",
    ]


def test_legacy_passport_snapshot_survives_later_group_and_support_edits() -> None:
    rendered = render_message(
        message_type="passport_link",
        group_name="Thailand 2026",
        support_contacts="Original Support: +919876543211",
        message_content="Upload clear copies and review every detail.",
        passport_link="https://tech.example/passports/secure-token",
    )
    legacy_log = SimpleNamespace(
        message_type="passport_link",
        rendered_message=rendered,
        header_parameter_values=None,
        template_parameter_values=None,
    )

    header, body = _template_snapshot_from_log(legacy_log)

    assert header == []
    assert body == [
        (
            "Please use the secure link below to submit your travel documents required for "
            "your trip to Thailand 2026."
        ),
        "https://tech.example/passports/secure-token",
        "Upload clear copies and review every detail.",
        "Original Support: +919876543211",
    ]


def test_legacy_decoder_fails_closed_when_saved_render_is_modified() -> None:
    rendered = _legacy_welcome_rendered(
        "Welcome to the trip.",
        "Aman: +919876543211",
    ).replace("Team Global Connect Travels", "Unknown sender")

    with pytest.raises(ValueError, match="footer layout"):
        _decode_legacy_template_snapshot(
            message_type="welcome",
            rendered_message=rendered,
        )


def test_worker_prefers_frozen_template_snapshot_over_current_group_values() -> None:
    log = SimpleNamespace(
        header_parameter_values=[],
        template_parameter_values=[
            "Original trip message",
            "Original Support: +919876543211",
        ],
        is_explicit_resend=True,
    )

    header, body = _resolve_log_template_snapshot(
        log=log,
        message_type="welcome",
        fallback_header_parameters=[],
        fallback_parameters=["Changed trip message", "Changed support"],
    )

    assert header == []
    assert body == [
        "Original trip message",
        "Original Support: +919876543211",
    ]


@pytest.mark.asyncio
async def test_explicit_resend_never_mutates_baseline_delivery_ledger() -> None:
    session = AsyncMock()
    log = SimpleNamespace(
        is_explicit_resend=True,
        recipient_id=uuid.uuid4(),
        message_type="welcome",
    )

    await _set_message_state(
        session,
        log=log,
        expected_batch_id=uuid.uuid4(),
        state_status="failed",
        release_claim=True,
    )

    session.execute.assert_not_awaited()


@pytest.mark.asyncio
async def test_explicit_resend_worker_guard_uses_log_claim_not_baseline_state() -> None:
    batch_id = uuid.uuid4()
    recipient = SimpleNamespace(
        id=uuid.uuid4(),
        agency_id=uuid.uuid4(),
        broadcast_group_id=uuid.uuid4(),
        normalized_phone_number="+919876543210",
        removed_at=None,
    )
    group_result = MagicMock()
    group_result.scalar_one_or_none.return_value = SimpleNamespace(id=uuid.uuid4())
    recipient_result = MagicMock()
    recipient_result.scalar_one_or_none.return_value = recipient
    claim_result = MagicMock()
    claim_result.scalar_one_or_none.return_value = uuid.uuid4()
    active_replacement_result = MagicMock()
    active_replacement_result.scalar_one_or_none.return_value = None
    session = AsyncMock()
    session.execute.side_effect = [
        group_result,
        recipient_result,
        active_replacement_result,
        claim_result,
    ]

    sendable, reason = await _load_sendable_recipient(
        session,
        log=SimpleNamespace(
            id=uuid.uuid4(),
            recipient_id=recipient.id,
            broadcast_group_id=group_result.scalar_one_or_none.return_value.id,
            message_type="welcome",
            is_explicit_resend=True,
        ),
        expected_batch_id=batch_id,
    )

    assert sendable is recipient
    assert reason is None
    assert session.execute.await_count == 4


def test_explicit_resend_route_is_role_gated_and_returns_send_contract() -> None:
    route = next(
        item
        for item in whatsapp_router.routes
        if item.path == "/groups/{group_id}/recipients/{recipient_id}/resend"
    )

    assert route.methods == {"POST"}
    assert route.response_model.__name__ == "WhatsAppSendResponse"
    assert [dependency.call.__name__ for dependency in route.dependant.dependencies] == [
        "_check_role",
        "get_db_session",
    ]


@pytest.mark.asyncio
async def test_resend_endpoint_queues_one_edited_message_with_current_template(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    group_id = uuid.uuid4()
    recipient_id = uuid.uuid4()
    agency_id = uuid.uuid4()
    group = SimpleNamespace(
        id=group_id,
        agency_id=agency_id,
        name="Thailand",
        recipient_opt_in_confirmed_at=datetime.now(tz=UTC),
    )
    recipient = SimpleNamespace(
        id=recipient_id,
        agency_id=agency_id,
        broadcast_group_id=group_id,
        name="Aarav Sharma",
        normalized_phone_number="+919876543210",
        removed_at=None,
    )
    state = SimpleNamespace(status="delivered")
    source_log = SimpleNamespace(
        id=uuid.uuid4(),
        status="delivered",
        message_type="welcome",
        template_name="legacy_welcome_v2",
        rendered_message=render_message(
            message_type="welcome",
            group_name="Thailand",
            support_contacts="Aman: +919876543211",
            message_content="Original trip message",
        ),
        header_parameter_values=["media-original"],
        template_parameter_values=["Original trip message"],
    )

    def scalar_result(value: object) -> MagicMock:
        result = MagicMock()
        result.scalar_one_or_none.return_value = value
        return result

    session = AsyncMock()
    session.add = MagicMock()
    session.execute.side_effect = [
        scalar_result(group),
        scalar_result(recipient),
        scalar_result(None),
        scalar_result(state),
        MagicMock(),
        MagicMock(),
        scalar_result(None),
        scalar_result(source_log),
        MagicMock(),
    ]
    audit_record = AsyncMock()
    queue_message = MagicMock()
    monkeypatch.setattr(
        "app.presentation.api.v1.routes.whatsapp.AuditLogRepository.record",
        audit_record,
    )
    monkeypatch.setitem(
        sys.modules,
        "app.infrastructure.whatsapp.tasks",
        SimpleNamespace(
            process_whatsapp_broadcast=SimpleNamespace(apply_async=queue_message),
        ),
    )
    settings = SimpleNamespace(
        whatsapp_access_token="token",
        whatsapp_phone_number_id="phone-id",
        whatsapp_welcome_template_name="welcome_v3",
        whatsapp_passport_link_template_name="passport_template",
    )
    monkeypatch.setattr(
        "app.presentation.api.v1.routes.whatsapp.get_settings",
        lambda: settings,
    )
    current_user = SimpleNamespace(
        id=uuid.uuid4(),
        role=UserRole.SUPER_ADMIN,
        agency_id=None,
        email="admin@example.com",
    )

    response = await resend_recipient_message(
        group_id=group_id,
        recipient_id=recipient_id,
        body=WhatsAppResendRequest(
            message_type="welcome",
            message_content="Edited trip message",
        ),
        request=Request({"type": "http", "client": ("127.0.0.1", 1234)}),
        current_user=current_user,
        session=session,
    )

    assert response.queued == 1
    assert response.results[0].recipient_id == recipient_id
    queued_log = session.add.call_args.args[0]
    assert queued_log.is_explicit_resend is True
    assert queued_log.template_name == "welcome_v3"
    assert queued_log.template_parameter_values == ["Edited trip message"]
    assert queued_log.header_parameter_values == ["media-original"]
    assert state.status == "delivered"
    queue_message.assert_called_once()
    audit_record.assert_awaited_once()


@pytest.mark.asyncio
async def test_resend_endpoint_blocks_delivery_unknown_explicit_attempt() -> None:
    group_id = uuid.uuid4()
    recipient_id = uuid.uuid4()
    group = SimpleNamespace(
        id=group_id,
        agency_id=uuid.uuid4(),
        recipient_opt_in_confirmed_at=datetime.now(tz=UTC),
    )
    recipient = SimpleNamespace(
        id=recipient_id,
        agency_id=group.agency_id,
        broadcast_group_id=group_id,
        normalized_phone_number="+919876543210",
        removed_at=None,
    )

    def scalar_result(value: object) -> MagicMock:
        result = MagicMock()
        result.scalar_one_or_none.return_value = value
        return result

    session = AsyncMock()
    session.execute.side_effect = [
        scalar_result(group),
        scalar_result(recipient),
        scalar_result(None),
        scalar_result(SimpleNamespace(status="delivered")),
        MagicMock(),
        MagicMock(),
        scalar_result(SimpleNamespace(status="delivery_unknown")),
    ]
    current_user = SimpleNamespace(
        id=uuid.uuid4(),
        role=UserRole.SUPER_ADMIN,
        agency_id=None,
        email="admin@example.com",
    )

    with pytest.raises(HTTPException) as exc_info:
        await resend_recipient_message(
            group_id=group_id,
            recipient_id=recipient_id,
            body=WhatsAppResendRequest(message_type="welcome"),
            request=Request({"type": "http", "client": ("127.0.0.1", 1234)}),
            current_user=current_user,
            session=session,
        )

    assert exc_info.value.status_code == 409
    assert "unknown delivery outcome" in str(exc_info.value.detail)
    session.add.assert_not_called()


@pytest.mark.asyncio
async def test_resend_endpoint_rejects_removed_or_missing_recipient() -> None:
    group_id = uuid.uuid4()
    group_result = MagicMock()
    group_result.scalar_one_or_none.return_value = SimpleNamespace(
        id=group_id,
        agency_id=uuid.uuid4(),
        recipient_opt_in_confirmed_at=datetime.now(tz=UTC),
    )
    recipient_result = MagicMock()
    recipient_result.scalar_one_or_none.return_value = None
    session = AsyncMock()
    session.execute.side_effect = [group_result, recipient_result]
    current_user = SimpleNamespace(
        role=UserRole.SUPER_ADMIN,
        agency_id=None,
    )

    with pytest.raises(HTTPException) as exc_info:
        await resend_recipient_message(
            group_id=group_id,
            recipient_id=uuid.uuid4(),
            body=WhatsAppResendRequest(message_type="passport_link"),
            request=Request({"type": "http", "client": ("127.0.0.1", 1234)}),
            current_user=current_user,
            session=session,
        )

    assert exc_info.value.status_code == 404
    assert exc_info.value.detail == "WhatsApp recipient not found"


@pytest.mark.asyncio
async def test_resend_queue_failure_does_not_release_baseline_sent_state(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    group_id = uuid.uuid4()
    recipient_id = uuid.uuid4()
    agency_id = uuid.uuid4()
    group = SimpleNamespace(
        id=group_id,
        agency_id=agency_id,
        name="Thailand",
        recipient_opt_in_confirmed_at=datetime.now(tz=UTC),
    )
    recipient = SimpleNamespace(
        id=recipient_id,
        agency_id=agency_id,
        broadcast_group_id=group_id,
        name="Aarav Sharma",
        normalized_phone_number="+919876543210",
        removed_at=None,
    )
    baseline_state = SimpleNamespace(status="delivered")
    source_log = SimpleNamespace(
        id=uuid.uuid4(),
        status="delivered",
        message_type="welcome",
        template_name="welcome_template",
        rendered_message=render_message(
            message_type="welcome",
            group_name="Thailand",
            support_contacts="Aman: +919876543211",
            message_content="Original trip message",
        ),
        header_parameter_values=["media-original"],
        template_parameter_values=["Original trip message"],
    )

    def scalar_result(value: object) -> MagicMock:
        result = MagicMock()
        result.scalar_one_or_none.return_value = value
        return result

    session = AsyncMock()
    session.add = MagicMock()
    session.execute.side_effect = [
        scalar_result(group),
        scalar_result(recipient),
        scalar_result(None),
        scalar_result(baseline_state),
        MagicMock(),
        MagicMock(),
        scalar_result(None),
        scalar_result(source_log),
        MagicMock(),
    ]
    monkeypatch.setattr(
        "app.presentation.api.v1.routes.whatsapp.AuditLogRepository.record",
        AsyncMock(),
    )
    queue_message = MagicMock(side_effect=RuntimeError("broker unavailable"))
    monkeypatch.setitem(
        sys.modules,
        "app.infrastructure.whatsapp.tasks",
        SimpleNamespace(
            process_whatsapp_broadcast=SimpleNamespace(apply_async=queue_message),
        ),
    )
    monkeypatch.setattr(
        "app.presentation.api.v1.routes.whatsapp.get_settings",
        lambda: SimpleNamespace(
            whatsapp_access_token="token",
            whatsapp_phone_number_id="phone-id",
            whatsapp_welcome_template_name="welcome_template",
            whatsapp_passport_link_template_name="passport_template",
        ),
    )
    current_user = SimpleNamespace(
        id=uuid.uuid4(),
        role=UserRole.SUPER_ADMIN,
        agency_id=None,
        email="admin@example.com",
    )

    with pytest.raises(HTTPException) as exc_info:
        await resend_recipient_message(
            group_id=group_id,
            recipient_id=recipient_id,
            body=WhatsAppResendRequest(message_type="welcome"),
            request=Request({"type": "http", "client": ("127.0.0.1", 1234)}),
            current_user=current_user,
            session=session,
        )

    assert exc_info.value.status_code == 503
    assert baseline_state.status == "delivered"
    queued_log = session.add.call_args.args[0]
    assert queued_log.is_explicit_resend is True
    assert queued_log.status == "failed"
    assert session.commit.await_count == 2


@pytest.mark.asyncio
async def test_explicit_resend_webhook_updates_log_without_loading_baseline_state(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    log = SimpleNamespace(
        is_explicit_resend=True,
        status="submitted",
        status_updated_at=datetime.now(tz=UTC),
        provider_status_at=None,
        error_message=None,
    )
    logs_result = MagicMock()
    logs_result.scalars.return_value.all.return_value = [log]
    session = AsyncMock()
    session.execute.return_value = logs_result
    monkeypatch.setattr(
        "app.presentation.api.v1.routes.whatsapp.get_settings",
        lambda: SimpleNamespace(
            whatsapp_app_secret="",
            is_production=False,
        ),
    )
    payload = {
        "entry": [
            {
                "changes": [
                    {
                        "value": {
                            "statuses": [
                                {
                                    "id": "wamid.resend",
                                    "status": "failed",
                                    "timestamp": "1784419200",
                                }
                            ]
                        }
                    }
                ]
            }
        ]
    }
    request = SimpleNamespace(body=AsyncMock(return_value=json.dumps(payload).encode("utf-8")))

    response = await receive_whatsapp_webhook(
        request=request,
        x_hub_signature_256=None,
        session=session,
    )

    assert response.processed_statuses == 1
    assert log.status == "failed"
    assert session.execute.await_count == 1
    session.commit.assert_awaited_once()


@pytest.mark.asyncio
async def test_mark_failed_skips_baseline_ledger_for_explicit_resend(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    log = SimpleNamespace(
        is_explicit_resend=True,
        status="processing",
        status_updated_at=datetime.now(tz=UTC),
        error_message=None,
        recipient_id=uuid.uuid4(),
        message_type="welcome",
        batch_id=uuid.uuid4(),
    )
    logs_result = MagicMock()
    logs_result.scalars.return_value.all.return_value = [log]
    session = AsyncMock()
    session.execute.return_value = logs_result

    class AsyncContext:
        async def __aenter__(self) -> object:
            return session

        async def __aexit__(self, *args: object) -> None:
            return None

    monkeypatch.setattr(
        "app.infrastructure.whatsapp.worker_runtime.AsyncSessionFactory",
        lambda: AsyncContext(),
    )

    await mark_whatsapp_batch_failed(
        batch_id=str(log.batch_id),
        error_message="worker failed",
    )

    assert log.status == "delivery_unknown"
    assert session.execute.await_count == 1
    session.commit.assert_awaited_once()


@pytest.mark.asyncio
async def test_group_delete_blocks_processing_explicit_resend() -> None:
    group_id = uuid.uuid4()
    group_result = MagicMock()
    group_result.scalar_one_or_none.return_value = SimpleNamespace(id=group_id)
    baseline_processing_count = MagicMock()
    baseline_processing_count.scalar_one.return_value = 0
    explicit_processing_count = MagicMock()
    explicit_processing_count.scalar_one.return_value = 1
    active_replacement_result = MagicMock()
    active_replacement_result.scalar_one_or_none.return_value = None
    session = AsyncMock()
    session.execute.side_effect = [
        group_result,
        active_replacement_result,
        baseline_processing_count,
        explicit_processing_count,
    ]
    current_user = SimpleNamespace(
        role=UserRole.SUPER_ADMIN,
        agency_id=None,
    )

    with pytest.raises(HTTPException) as exc_info:
        await delete_broadcast_group(
            group_id=group_id,
            current_user=current_user,
            session=session,
        )

    assert exc_info.value.status_code == 409
    assert "provider request is currently in progress" in str(exc_info.value.detail)


@pytest.mark.asyncio
async def test_group_delete_blocks_active_passport_replacement() -> None:
    group_id = uuid.uuid4()
    group_result = MagicMock()
    group_result.scalar_one_or_none.return_value = SimpleNamespace(id=group_id)
    active_replacement_result = MagicMock()
    active_replacement_result.scalar_one_or_none.return_value = uuid.uuid4()
    session = AsyncMock()
    session.execute.side_effect = [
        group_result,
        active_replacement_result,
    ]
    current_user = SimpleNamespace(
        role=UserRole.SUPER_ADMIN,
        agency_id=None,
    )

    with pytest.raises(HTTPException) as exc_info:
        await delete_broadcast_group(
            group_id=group_id,
            current_user=current_user,
            session=session,
        )

    assert exc_info.value.status_code == 409
    assert "marked as replaced" in str(exc_info.value.detail)
    assert session.execute.await_count == 2
