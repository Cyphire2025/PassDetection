from __future__ import annotations

import io
import uuid
from datetime import UTC, date, datetime
from unittest.mock import AsyncMock, MagicMock

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.application.use_cases.whatsapp.group_submission_matching import (
    RecipientFieldSet,
    SubmissionMatchRow,
)
from app.domain.entities.entities import (
    ClientGroup,
    GroupStatus,
    User,
    UserRole,
)
from app.presentation.api.v1.routes import passports as passports_route
from app.presentation.api.v1.schemas.passport_schemas import (
    ExportSelectedGroupsRequest,
    ExportSelectedPassportsRequest,
)

NOW = datetime(2026, 7, 23, 12, tzinfo=UTC)


def _group(
    *,
    name: str,
    agency_id: uuid.UUID,
    staff_code_enabled: bool = False,
    international_airport_enabled: bool = False,
) -> ClientGroup:
    return ClientGroup(
        id=uuid.uuid4(),
        name=name,
        token=f"token-{uuid.uuid4()}",
        agency_id=agency_id,
        status=GroupStatus.ACTIVE,
        created_by_user_id=uuid.uuid4(),
        created_at=NOW,
        destination="Vietnam",
        travel_date=date(2026, 8, 12),
        return_date=date(2026, 8, 15),
        staff_code_enabled=staff_code_enabled,
        departure_cities=(
            ["Delhi", "Mumbai"] if international_airport_enabled else []
        ),
        nearest_international_airport_enabled=international_airport_enabled,
    )


def _pending_match(
    *,
    name: str,
    phone: str,
    zone: str,
    staff_code: str | None = None,
    extra_fields: dict[str, str] | None = None,
) -> SubmissionMatchRow:
    recipient_id = uuid.uuid4()
    imported_fields = {"Zone Name": zone}
    if staff_code is not None:
        imported_fields["Staff Code"] = staff_code
    imported_fields.update(extra_fields or {})
    return SubmissionMatchRow(
        status="not_submitted",
        match_basis=None,
        normalized_phone=phone,
        recipient_ids=(recipient_id,),
        submission_ids=(),
        broadcast_ids=(uuid.uuid4(),),
        broadcast_names=("Recipients",),
        recipient_names=(name,),
        submission_names=(),
        updated_at=NOW,
        recipient_fields=(
            RecipientFieldSet(
                recipient_id=recipient_id,
                fields=imported_fields,
            ),
        ),
    )


@pytest.mark.asyncio
async def test_selected_groups_export_combines_pending_only_groups(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    agency_id = uuid.uuid4()
    mumbai_group = _group(name="Mumbai Group", agency_id=agency_id)
    delhi_group = _group(
        name="Delhi Group",
        agency_id=agency_id,
        staff_code_enabled=True,
    )
    groups = [mumbai_group, delhi_group]

    group_result = MagicMock()
    group_result.scalars.return_value.all.return_value = groups
    submission_result = MagicMock()
    submission_result.scalars.return_value.all.return_value = []
    session = AsyncMock(spec=AsyncSession)
    session.execute.side_effect = [group_result, submission_result]

    match_rows = {
        mumbai_group.id: [
            _pending_match(
                name="Mumbai Pending",
                phone="+919000000002",
                zone="Mumbai-2",
                extra_fields={
                    "Region": "West",
                    "Mumbai Only": "Mumbai value",
                },
            )
        ],
        delhi_group.id: [
            _pending_match(
                name="Delhi Pending",
                phone="+919000000001",
                zone="Delhi",
                staff_code="25290",
                extra_fields={
                    "Region": "North",
                    "Delhi Only": "Delhi value",
                },
            )
        ],
    }
    monkeypatch.setattr(
        passports_route.ClientGroupRepository,
        "_to_entity",
        staticmethod(lambda model: model),
    )
    matcher = AsyncMock(return_value=match_rows)
    monkeypatch.setattr(
        passports_route,
        "_export_whatsapp_match_rows",
        matcher,
    )

    response = await passports_route.export_selected_groups(
        body=ExportSelectedGroupsRequest(
            group_ids=[mumbai_group.id, delhi_group.id],
            supplemental_fields=[
                "whatsapp:delhi_only",
                "zone_name",
                "whatsapp:mumbai_only",
                "whatsapp:region",
            ],
            group_by_field="zone_name",
        ),
        current_user=User(
            id=uuid.uuid4(),
            email="admin@example.test",
            hashed_password="hash",
            full_name="Admin",
            role=UserRole.AGENCY_ADMIN,
            agency_id=agency_id,
        ),
        session=session,
    )
    chunks: list[bytes] = []
    async for chunk in response.body_iterator:
        chunks.append(chunk if isinstance(chunk, bytes) else chunk.encode())
    content = b"".join(chunks)
    worksheet = load_workbook(io.BytesIO(content), data_only=False).active
    headers = [cell.value for cell in worksheet[4]]
    assert headers[4:8] == [
        "Zone Name",
        "Region",
        "Mumbai Only",
        "Delhi Only",
    ]
    name_column = headers.index("GIVEN NAME") + 1
    zone_column = headers.index("Zone Name") + 1
    staff_code_column = headers.index("Staff Code") + 1
    pending_rows = [
        row
        for row in range(5, worksheet.max_row + 1)
        if worksheet.cell(row=row, column=name_column).value
        in {"DELHI PENDING", "MUMBAI PENDING"}
    ]

    assert session.execute.await_count == 2
    group_sql = str(session.execute.await_args_list[0].args[0])
    submission_sql = str(session.execute.await_args_list[1].args[0])
    assert "client_groups.status !=" in group_sql
    assert "client_groups.deleted_at IS NULL" in group_sql
    assert "client_groups.status !=" in submission_sql
    assert "client_groups.deleted_at IS NULL" in submission_sql
    matcher.assert_awaited_once_with(session, [], groups=groups)
    assert [
        worksheet.cell(row=row, column=name_column).value
        for row in pending_rows
    ] == ["DELHI PENDING", "MUMBAI PENDING"]
    assert [
        worksheet.cell(row=row, column=zone_column).value
        for row in pending_rows
    ] == ["Delhi", "Mumbai-2"]
    assert worksheet.cell(
        row=pending_rows[0],
        column=staff_code_column,
    ).value == "25290"
    assert worksheet.cell(
        row=pending_rows[0],
        column=headers.index("Region") + 1,
    ).value == "North"
    assert worksheet.cell(
        row=pending_rows[1],
        column=headers.index("Mumbai Only") + 1,
    ).value == "Mumbai value"
    assert all(
        cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb == "00FFF2CC"
        for row in pending_rows
        for cell in worksheet[row]
    )
    assert set(worksheet.tables) == {"PassportSubmissions"}
    assert worksheet.tables["PassportSubmissions"].ref.endswith(
        str(worksheet.max_row)
    )


@pytest.mark.asyncio
async def test_selected_passport_export_excludes_retained_deleted_groups() -> None:
    agency_id = uuid.uuid4()
    empty_result = MagicMock()
    empty_result.scalars.return_value.all.return_value = []
    session = AsyncMock(spec=AsyncSession)
    session.execute.return_value = empty_result

    with pytest.raises(HTTPException) as exc_info:
        await passports_route.export_selected_passports(
            body=ExportSelectedPassportsRequest(
                submission_ids=[uuid.uuid4()],
            ),
            current_user=User(
                id=uuid.uuid4(),
                email="admin@example.test",
                hashed_password="hash",
                full_name="Admin",
                role=UserRole.AGENCY_ADMIN,
                agency_id=agency_id,
            ),
            session=session,
        )

    assert exc_info.value.status_code == 404
    export_sql = str(session.execute.await_args.args[0])
    assert "client_groups.status !=" in export_sql
    assert "client_groups.deleted_at IS NULL" in export_sql


def test_combined_catalog_places_common_fields_before_group_specific_fields() -> None:
    agency_id = uuid.uuid4()
    north_group = _group(name="North", agency_id=agency_id)
    south_group = _group(name="South", agency_id=agency_id)
    rows_by_group = {
        north_group.id: [
            _pending_match(
                name="North Traveller",
                phone="+919000000001",
                zone="North",
                extra_fields={
                    "Branch": "Delhi",
                    "Category": "A",
                    "North Only": "N",
                },
            )
        ],
        south_group.id: [
            _pending_match(
                name="South Traveller",
                phone="+919000000002",
                zone="South",
                extra_fields={
                    "Branch": "Mumbai",
                    "Category": "B",
                    "South Only": "S",
                },
            )
        ],
    }

    catalog = passports_route._combined_export_field_catalog(
        [north_group, south_group],
        rows_by_group,
        [],
    )

    assert [field["key"] for field in catalog] == [
        "zone_name",
        "whatsapp:branch",
        "whatsapp:category",
        "whatsapp:north_only",
        "whatsapp:south_only",
    ]
    assert catalog[0]["selected_by_default"] is True


def test_combined_catalog_makes_cross_group_labels_unique() -> None:
    merged = passports_route._merge_export_field_catalogs(
        [
            [
                {
                    "key": "whatsapp:first",
                    "label": "Shared label",
                    "source": "whatsapp",
                    "selected_by_default": False,
                }
            ],
            [
                {
                    "key": "whatsapp:second",
                    "label": "Shared label",
                    "source": "whatsapp",
                    "selected_by_default": False,
                }
            ],
        ]
    )

    assert [field["label"] for field in merged] == [
        "Shared label",
        "Shared label (WhatsApp)",
    ]


def test_combined_catalog_keeps_group_specific_zone_first() -> None:
    merged = passports_route._merge_export_field_catalogs(
        [
            [
                {
                    "key": "whatsapp:shared",
                    "label": "Shared",
                    "source": "whatsapp",
                    "selected_by_default": False,
                },
                {
                    "key": "whatsapp:first_only",
                    "label": "First Only",
                    "source": "whatsapp",
                    "selected_by_default": False,
                },
            ],
            [
                {
                    "key": "zone_name",
                    "label": "Zone Name",
                    "source": "whatsapp",
                    "selected_by_default": True,
                },
                {
                    "key": "whatsapp:shared",
                    "label": "Shared",
                    "source": "whatsapp",
                    "selected_by_default": False,
                },
            ],
        ]
    )

    assert [field["key"] for field in merged] == [
        "zone_name",
        "whatsapp:shared",
        "whatsapp:first_only",
    ]


@pytest.mark.asyncio
async def test_selected_groups_field_options_use_the_merged_catalog(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    agency_id = uuid.uuid4()
    north_group = _group(name="North", agency_id=agency_id)
    south_group = _group(
        name="South",
        agency_id=agency_id,
        international_airport_enabled=True,
    )

    group_result = MagicMock()
    group_result.scalars.return_value.all.return_value = [
        south_group,
        north_group,
    ]
    submission_result = MagicMock()
    submission_result.scalars.return_value.all.return_value = []
    session = AsyncMock(spec=AsyncSession)
    session.execute.side_effect = [group_result, submission_result]
    rows_by_group = {
        north_group.id: [
            _pending_match(
                name="North Traveller",
                phone="+919000000004",
                zone="North",
                extra_fields={"Branch": "Delhi", "North Only": "N"},
            )
        ],
        south_group.id: [
            _pending_match(
                name="South Traveller",
                phone="+919000000005",
                zone="South",
                extra_fields={"Branch": "Mumbai", "South Only": "S"},
            )
        ],
    }

    monkeypatch.setattr(
        passports_route.ClientGroupRepository,
        "_to_entity",
        staticmethod(lambda model: model),
    )
    monkeypatch.setattr(
        passports_route,
        "_export_whatsapp_match_rows",
        AsyncMock(return_value=rows_by_group),
    )

    response = await passports_route.get_selected_groups_export_fields(
        body=ExportSelectedGroupsRequest(
            group_ids=[north_group.id, south_group.id],
        ),
        current_user=User(
            id=uuid.uuid4(),
            email="admin@example.test",
            hashed_password="hash",
            full_name="Admin",
            role=UserRole.AGENCY_ADMIN,
            agency_id=agency_id,
        ),
        session=session,
    )

    assert response.group_ids == [north_group.id, south_group.id]
    assert [field.key for field in response.fields] == [
        "zone_name",
        "whatsapp:branch",
        "whatsapp:north_only",
        "whatsapp:south_only",
    ]
    assert response.default_selected_fields == ["zone_name"]
    assert response.default_group_by_field == "zone_name"
    assert response.grouping_fields[0].key == "international_airport"
    assert response.grouping_fields[0].fixed is True


@pytest.mark.asyncio
async def test_selected_groups_export_omits_zone_when_not_selected(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    agency_id = uuid.uuid4()
    group = _group(name="No Zone Export", agency_id=agency_id)

    group_result = MagicMock()
    group_result.scalars.return_value.all.return_value = [group]
    submission_result = MagicMock()
    submission_result.scalars.return_value.all.return_value = []
    session = AsyncMock(spec=AsyncSession)
    session.execute.side_effect = [group_result, submission_result]

    monkeypatch.setattr(
        passports_route.ClientGroupRepository,
        "_to_entity",
        staticmethod(lambda model: model),
    )
    monkeypatch.setattr(
        passports_route,
        "_export_whatsapp_match_rows",
        AsyncMock(
            return_value={
                group.id: [
                    _pending_match(
                        name="Pending Traveller",
                        phone="+919000000003",
                        zone="West",
                    )
                ]
            }
        ),
    )

    response = await passports_route.export_selected_groups(
        body=ExportSelectedGroupsRequest(
            group_ids=[group.id],
            supplemental_fields=[],
            group_by_field="none",
        ),
        current_user=User(
            id=uuid.uuid4(),
            email="admin@example.test",
            hashed_password="hash",
            full_name="Admin",
            role=UserRole.AGENCY_ADMIN,
            agency_id=agency_id,
        ),
        session=session,
    )
    chunks: list[bytes] = []
    async for chunk in response.body_iterator:
        chunks.append(chunk if isinstance(chunk, bytes) else chunk.encode())
    worksheet = load_workbook(
        io.BytesIO(b"".join(chunks)),
        data_only=False,
    ).active
    headers = [cell.value for cell in worksheet[4]]

    assert "Zone Name" not in headers
    assert headers[4] == "Age Group"
