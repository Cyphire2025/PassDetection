from __future__ import annotations

import io
import uuid

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security.jwt import create_access_token
from app.infrastructure.database.models import AgencyModel, UserModel


@pytest.mark.asyncio
async def test_menu_library_generates_a_saved_plan_without_repeated_dishes(
    client: AsyncClient,
    db_session: AsyncSession,
) -> None:
    agency_id = uuid.uuid4()
    user_id = uuid.uuid4()
    db_session.add(
        AgencyModel(
            id=agency_id,
            name="Menu Test Agency",
            email=f"{agency_id}@example.com",
        )
    )
    db_session.add(
        UserModel(
            id=user_id,
            email=f"{user_id}@example.com",
            hashed_password="not-used-in-this-test",
            full_name="Menu Planner",
            role="agency_admin",
            agency_id=agency_id,
            is_active=True,
        )
    )
    await db_session.commit()
    access_token, _ = create_access_token(
        user_id=user_id,
        role="agency_admin",
        agency_id=agency_id,
    )
    headers = {"Authorization": f"Bearer {access_token}"}

    chicken_response = await client.post(
        "/api/v1/menu/categories",
        json={"name": "Chicken"},
        headers=headers,
    )
    paneer_response = await client.post(
        "/api/v1/menu/categories",
        json={"name": "Paneer"},
        headers=headers,
    )
    assert chicken_response.status_code == 201, chicken_response.text
    assert paneer_response.status_code == 201, paneer_response.text
    chicken_id = chicken_response.json()["id"]
    paneer_id = paneer_response.json()["id"]

    created_dishes: dict[str, list[str]] = {chicken_id: [], paneer_id: []}
    for category_id, dish_name in (
        (chicken_id, "Butter Chicken"),
        (chicken_id, "Chicken Tikka"),
        (chicken_id, "Chicken Lababdar"),
        (chicken_id, "Kadai Chicken"),
        (chicken_id, "Chicken Curry"),
        (paneer_id, "Shahi Paneer"),
        (paneer_id, "Paneer Tikka"),
        (paneer_id, "Paneer Lababdar"),
        (paneer_id, "Kadai Paneer"),
        (paneer_id, "Matar Paneer"),
    ):
        dish_response = await client.post(
            f"/api/v1/menu/categories/{category_id}/dishes",
            json={"name": dish_name},
            headers=headers,
        )
        assert dish_response.status_code == 201
        created_dishes[category_id].append(dish_response.json()["id"])

    plan_response = await client.post(
        "/api/v1/menu/plans/generate",
        json={
            "name": "Two Day Test Trip",
            "trip_days": 2,
            "category_ids": [chicken_id, paneer_id],
        },
        headers=headers,
    )
    assert plan_response.status_code == 201
    plan = plan_response.json()
    entries = [
        dish for day in plan["days"] for meal in (day["lunch"], day["dinner"]) for dish in meal
    ]
    assert len(entries) == 8
    assert len({entry["dish_id"] for entry in entries}) == 8
    for day in plan["days"]:
        assert {entry["category_id"] for entry in day["lunch"]} == {
            chicken_id,
            paneer_id,
        }
        assert {entry["category_id"] for entry in day["dinner"]} == {
            chicken_id,
            paneer_id,
        }

    workspace_response = await client.get("/api/v1/menu", headers=headers)
    assert workspace_response.status_code == 200
    workspace = workspace_response.json()
    assert workspace["active_dishes"] == 10
    assert workspace["max_trip_days_without_repeats"] == 2
    assert workspace["plans"][0]["id"] == plan["id"]

    chicken_entry = next(
        entry for entry in plan["days"][0]["lunch"] if entry["category_id"] == chicken_id
    )
    used_chicken_ids = {entry["dish_id"] for entry in entries if entry["category_id"] == chicken_id}
    unused_chicken_id = next(
        dish_id for dish_id in created_dishes[chicken_id] if dish_id not in used_chicken_ids
    )
    update_response = await client.patch(
        f"/api/v1/menu/plans/{plan['id']}/entries/{chicken_entry['id']}",
        json={"dish_id": unused_chicken_id},
        headers=headers,
    )
    assert update_response.status_code == 200

    cross_category_response = await client.patch(
        f"/api/v1/menu/plans/{plan['id']}/entries/{chicken_entry['id']}",
        json={"dish_id": created_dishes[paneer_id][0]},
        headers=headers,
    )
    assert cross_category_response.status_code == 409
    assert "one dish from each selected category" in (cross_category_response.json()["detail"])

    export_response = await client.get(
        f"/api/v1/menu/plans/{plan['id']}/export.xlsx",
        headers=headers,
    )
    assert export_response.status_code == 200
    assert export_response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(io.BytesIO(export_response.content))
    assert workbook.sheetnames == ["Meal Plan", "Dish List"]
    assert [workbook["Meal Plan"].cell(5, column).value for column in range(1, 6)] == [
        "Day",
        "Date",
        "Meal",
        "Chicken",
        "Paneer",
    ]

    other_agency_id = uuid.uuid4()
    other_user_id = uuid.uuid4()
    db_session.add(
        AgencyModel(
            id=other_agency_id,
            name="Other Menu Agency",
            email=f"{other_agency_id}@example.com",
        )
    )
    db_session.add(
        UserModel(
            id=other_user_id,
            email=f"{other_user_id}@example.com",
            hashed_password="not-used-in-this-test",
            full_name="Other Menu Planner",
            role="agency_admin",
            agency_id=other_agency_id,
            is_active=True,
        )
    )
    await db_session.commit()
    other_token, _ = create_access_token(
        user_id=other_user_id,
        role="agency_admin",
        agency_id=other_agency_id,
    )
    isolated_workspace = await client.get(
        "/api/v1/menu",
        headers={"Authorization": f"Bearer {other_token}"},
    )
    assert isolated_workspace.status_code == 200
    assert isolated_workspace.json()["categories"] == []
    assert isolated_workspace.json()["plans"] == []


@pytest.mark.asyncio
async def test_menu_generation_refuses_to_repeat_when_dishes_are_insufficient(
    client: AsyncClient,
    db_session: AsyncSession,
) -> None:
    agency_id = uuid.uuid4()
    user_id = uuid.uuid4()
    db_session.add(
        AgencyModel(
            id=agency_id,
            name="Small Menu Agency",
            email=f"{agency_id}@example.com",
        )
    )
    db_session.add(
        UserModel(
            id=user_id,
            email=f"{user_id}@example.com",
            hashed_password="not-used-in-this-test",
            full_name="Small Menu Planner",
            role="agency_staff",
            agency_id=agency_id,
            is_active=True,
        )
    )
    await db_session.commit()
    access_token, _ = create_access_token(
        user_id=user_id,
        role="agency_staff",
        agency_id=agency_id,
    )
    headers = {"Authorization": f"Bearer {access_token}"}

    category_response = await client.post(
        "/api/v1/menu/categories",
        json={"name": "Vegetarian"},
        headers=headers,
    )
    category_id = category_response.json()["id"]
    for dish_name in ("Dal Tadka", "Jeera Rice", "Mixed Veg"):
        await client.post(
            f"/api/v1/menu/categories/{category_id}/dishes",
            json={"name": dish_name},
            headers=headers,
        )

    response = await client.post(
        "/api/v1/menu/plans/generate",
        json={
            "name": "Impossible Plan",
            "trip_days": 2,
            "category_ids": [category_id],
        },
        headers=headers,
    )

    assert response.status_code == 422
    assert response.json()["detail"] == (
        "Each selected category needs 4 active dishes for a 2-day lunch and dinner "
        "plan. Vegetarian: add 1 dish."
    )
